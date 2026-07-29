#!/usr/bin/env python3
# ============================================================
#  売上管理シート（Excel）の生成
# ------------------------------------------------------------
#  out/dashboard-data.json（Kintone から書き出したもの）を読み、
#  Chatwork に送れる Excel ファイルを作ります。
#
#  実行:
#    python3 scripts/buildSalesSheet.py
#
#  シート構成:
#    1. 日次サマリー   … 日ごとの売上・アクセス・転換率・販売個数
#    2. 商品別サマリー … 商品ごとのチャネル別販売個数
#    3. 商品×日 個数   … 商品と日付のクロス集計
#    4. 月次サマリー   … 月ごとの集計
#
#  ★合計・前日比・構成比はすべて Excel の数式で入れています。
#    数字を1つ直せば全部が計算し直されるようにするためです。
# ============================================================
import json
import os
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, 'out', 'dashboard-data.json')
OUT = os.path.join(ROOT, 'out', '売上管理シート.xlsx')

FONT = 'Arial'
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
HEAD_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color='808080')
TOTAL_FILL = PatternFill('solid', fgColor='D9E2F3')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(bottom=THIN)

YEN = '¥#,##0;(¥#,##0);-'
NUM = '#,##0;(#,##0);-'
PCT = '0.0%;(0.0%);-'
PCT2 = '0.00%;(0.00%);-'

WEEKDAY = ['月', '火', '水', '木', '金', '土', '日']

CHANNELS = [('rakuten', '楽天'), ('amazon', 'Amazon'), ('shopify', 'Shopify'), ('tiktok', 'TikTok')]


def header(ws, row, labels, widths=None):
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def weekday_of(iso):
    y, m, d = (int(x) for x in iso.split('-'))
    return WEEKDAY[date(y, m, d).weekday()]


def build_daily(wb, data):
    ws = wb.create_sheet('日次サマリー')
    cols = ['日付', '曜日', '楽天', 'Amazon', '自社サイト', '合計', '前日比',
            '楽天アクセス', '楽天転換率', 'お気に入り', 'Amazonアクセス', 'Amazon転換率', '販売個数']
    header(ws, 1, cols, [11, 5, 12, 12, 12, 13, 9, 12, 10, 10, 13, 12, 10])

    r = 2
    for d in data['daily']:
        ws.cell(row=r, column=1, value=d['date']).font = BODY
        ws.cell(row=r, column=2, value=weekday_of(d['date'])).font = BODY
        for i, key in enumerate(['rakuten', 'amazon', 'own']):
            c = ws.cell(row=r, column=3 + i, value=d['sales'][key] or 0)
            c.font = BODY
            c.number_format = YEN
        # 合計は数式（元の数字を直したら自動で合う）
        c = ws.cell(row=r, column=6, value=f'=SUM(C{r}:E{r})')
        c.font = BOLD
        c.number_format = YEN
        # 前日比。1行目と、前日が0の場合は空欄にする
        if r > 2:
            c = ws.cell(row=r, column=7, value=f'=IF(F{r-1}=0,"",F{r}/F{r-1}-1)')
            c.number_format = PCT
            c.font = BODY

        vals = [d['access']['rakuten'], d['cvr']['rakuten'], d['favorites'],
                d['access']['amazon'], d['cvr']['amazon']]
        fmts = [NUM, PCT2, NUM, NUM, PCT2]
        for i, (v, f) in enumerate(zip(vals, fmts)):
            c = ws.cell(row=r, column=8 + i)
            # 転換率は「1.43」で入っているので、Excelの%表示に合わせて 100 で割る
            c.value = (v / 100 if f == PCT2 and v is not None else v)
            c.number_format = f
            c.font = BODY

        units = sum(d['units'].values()) if d['units'] else 0
        c = ws.cell(row=r, column=13, value=units)
        c.number_format = NUM
        c.font = BODY

        for col in range(1, 14):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    last = r - 1
    # ── 合計行 ──
    ws.cell(row=r, column=1, value='合計').font = BOLD
    for col, letter in [(3, 'C'), (4, 'D'), (5, 'E'), (6, 'F')]:
        c = ws.cell(row=r, column=col, value=f'=SUM({letter}2:{letter}{last})')
        c.font = BOLD
        c.number_format = YEN
        c.fill = TOTAL_FILL
    c = ws.cell(row=r, column=13, value=f'=SUM(M2:M{last})')
    c.font = BOLD
    c.number_format = NUM
    c.fill = TOTAL_FILL
    for col in [1, 2, 7, 8, 9, 10, 11, 12]:
        ws.cell(row=r, column=col).fill = TOTAL_FILL

    # ── 平均行（転換率は単純平均。空欄は除いて平均する）──
    r += 1
    ws.cell(row=r, column=1, value='平均').font = BOLD
    for col, letter in [(3, 'C'), (4, 'D'), (5, 'E'), (6, 'F')]:
        c = ws.cell(row=r, column=col, value=f'=IFERROR(AVERAGE({letter}2:{letter}{last}),0)')
        c.font = BOLD
        c.number_format = YEN
    for col, letter, fmt in [(8, 'H', NUM), (9, 'I', PCT2), (10, 'J', NUM),
                             (11, 'K', NUM), (12, 'L', PCT2), (13, 'M', NUM)]:
        c = ws.cell(row=r, column=col, value=f'=IFERROR(AVERAGE({letter}2:{letter}{last}),0)')
        c.font = BOLD
        c.number_format = fmt

    r += 2
    ws.cell(row=r, column=1, value='※ 合計・前日比・平均はすべて数式です。上の数字を直すと自動で計算し直されます。').font = NOTE
    r += 1
    ws.cell(row=r, column=1, value=f"※ 出典: {data['source']}（Kintoneは読み取りのみ）").font = NOTE
    return last


def build_products(wb, data):
    ws = wb.create_sheet('商品別サマリー')
    cols = ['商品名', 'カテゴリ', '楽天', 'Amazon', 'Shopify', 'TikTok', '合計個数', '構成比']
    header(ws, 1, cols, [26, 14, 10, 10, 10, 10, 11, 9])

    rows = []
    for p in data['products']:
        by = {}
        for key, _ in CHANNELS:
            by[key] = sum((p['units'].get(key) or {}).values())
        total = sum(by.values())
        if total > 0:
            rows.append((p['name'], p['group'], by, total))
    rows.sort(key=lambda x: -x[3])

    r = 2
    for name, group, by, _ in rows:
        ws.cell(row=r, column=1, value=name).font = BODY
        ws.cell(row=r, column=2, value=group).font = BODY
        for i, (key, _) in enumerate(CHANNELS):
            c = ws.cell(row=r, column=3 + i, value=by[key])
            c.number_format = NUM
            c.font = BODY
        c = ws.cell(row=r, column=7, value=f'=SUM(C{r}:F{r})')
        c.font = BOLD
        c.number_format = NUM
        r += 1

    last = r - 1
    # 構成比は合計行を参照する（合計行の位置が決まってから入れる）
    total_row = r
    for rr in range(2, last + 1):
        c = ws.cell(row=rr, column=8, value=f'=IF($G${total_row}=0,"",G{rr}/$G${total_row})')
        c.number_format = PCT
        c.font = BODY
        for col in range(1, 9):
            ws.cell(row=rr, column=col).border = BORDER

    ws.cell(row=total_row, column=1, value='合計').font = BOLD
    for col, letter in [(3, 'C'), (4, 'D'), (5, 'E'), (6, 'F'), (7, 'G')]:
        c = ws.cell(row=total_row, column=col, value=f'=SUM({letter}2:{letter}{last})')
        c.font = BOLD
        c.number_format = NUM
        c.fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=2).fill = TOTAL_FILL
    ws.cell(row=total_row, column=8).fill = TOTAL_FILL

    r = total_row + 2
    ws.cell(row=r, column=1, value='※ 商品名は資料によって書き方が違うため、確実に同じと分かるものだけまとめています。').font = NOTE
    r += 1
    ws.cell(row=r, column=1, value='※ 楽天の商品ページ名「多機能PC」系は、販売個数のどの商品に当たるか確証が無いため未統合です。要確認。').font = NOTE


def build_matrix(wb, data):
    ws = wb.create_sheet('商品×日 個数')
    dates = [d['date'] for d in data['daily']]

    rows = []
    for p in data['products']:
        per_day = {}
        for key, _ in CHANNELS:
            for dt, q in (p['units'].get(key) or {}).items():
                per_day[dt] = per_day.get(dt, 0) + q
        total = sum(per_day.values())
        if total > 0:
            rows.append((p['name'], per_day, total))
    rows.sort(key=lambda x: -x[2])

    header(ws, 1, ['商品名'] + [d[5:] for d in dates] + ['合計'],
           [24] + [7] * len(dates) + [10])

    r = 2
    for name, per_day, _ in rows:
        ws.cell(row=r, column=1, value=name).font = BODY
        for i, dt in enumerate(dates):
            v = per_day.get(dt)
            c = ws.cell(row=r, column=2 + i, value=v if v else None)
            c.number_format = NUM
            c.font = BODY
        first = get_column_letter(2)
        lastc = get_column_letter(1 + len(dates))
        c = ws.cell(row=r, column=2 + len(dates), value=f'=SUM({first}{r}:{lastc}{r})')
        c.font = BOLD
        c.number_format = NUM
        r += 1

    last = r - 1
    ws.cell(row=r, column=1, value='合計').font = BOLD
    for i in range(len(dates) + 1):
        letter = get_column_letter(2 + i)
        c = ws.cell(row=r, column=2 + i, value=f'=SUM({letter}2:{letter}{last})')
        c.font = BOLD
        c.number_format = NUM
        c.fill = TOTAL_FILL
    ws.cell(row=r, column=1).fill = TOTAL_FILL

    r += 2
    ws.cell(row=r, column=1, value='※ 空欄はその日の販売がゼロ、または販売個数の記載が無い日です。').font = NOTE


def build_monthly(wb, data):
    ws = wb.create_sheet('月次サマリー')
    header(ws, 1, ['月', '記録日数', '楽天', 'Amazon', '自社サイト', '合計', '日商平均', '販売個数'],
           [10, 10, 14, 14, 14, 15, 14, 11])

    months = {}
    for d in data['daily']:
        ym = d['date'][:7]
        m = months.setdefault(ym, {'days': 0, 'rakuten': 0, 'amazon': 0, 'own': 0, 'units': 0})
        m['days'] += 1
        m['rakuten'] += d['sales']['rakuten'] or 0
        m['amazon'] += d['sales']['amazon'] or 0
        m['own'] += d['sales']['own'] or 0
        m['units'] += sum(d['units'].values()) if d['units'] else 0

    r = 2
    for ym in sorted(months):
        m = months[ym]
        ws.cell(row=r, column=1, value=ym).font = BODY
        c = ws.cell(row=r, column=2, value=m['days'])
        c.font = BODY
        c.number_format = NUM
        for i, key in enumerate(['rakuten', 'amazon', 'own']):
            c = ws.cell(row=r, column=3 + i, value=m[key])
            c.font = BODY
            c.number_format = YEN
        c = ws.cell(row=r, column=6, value=f'=SUM(C{r}:E{r})')
        c.font = BOLD
        c.number_format = YEN
        # 日商平均。記録日数が違う月どうしを比べられるようにする
        c = ws.cell(row=r, column=7, value=f'=IF(B{r}=0,"",F{r}/B{r})')
        c.font = BOLD
        c.number_format = YEN
        c = ws.cell(row=r, column=8, value=m['units'])
        c.font = BODY
        c.number_format = NUM
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='※ 月によって記録日数が違います。月どうしの比較は「日商平均」でご覧ください。').font = NOTE
    r += 1
    ws.cell(row=r, column=1, value='　（累計どうしを比べると、日数の差がそのまま増減に見えてしまいます）').font = NOTE
    r += 2
    if data.get('issues'):
        ws.cell(row=r, column=1, value='【要確認】データの記入について').font = BOLD
        r += 1
        for s in data['issues']:
            ws.cell(row=r, column=1, value='・' + s).font = BODY
            r += 1


def main():
    if not os.path.exists(DATA):
        print(f'データがありません: {DATA}\n  先に npm run dashboard:data を実行してください。')
        sys.exit(1)

    with open(DATA, encoding='utf-8') as f:
        data = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)

    build_daily(wb, data)
    build_products(wb, data)
    build_matrix(wb, data)
    build_monthly(wb, data)

    # openpyxl は数式に計算結果を持たせられない（値は空のまま保存される）。
    # このフラグを立てておくと、Excel / Googleスプレッドシートが
    # ファイルを開いた瞬間に全数式を計算し直してくれる。
    wb.calculation.fullCalcOnLoad = True

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f'✅ 作成しました: {OUT}')
    print(f'   期間 {data["period"]["from"]} 〜 {data["period"]["to"]} / {len(data["daily"])}日分')


if __name__ == '__main__':
    main()
