#!/usr/bin/env python3
# ============================================================
#  朝礼用 売上進捗シート（イベント加重）を Excel で生成
# ------------------------------------------------------------
#  入力: out/chorei-progress.json（choreiSheetData.js が作る）
#  出力: out/売上進捗シート.xlsx
# ============================================================
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
data = json.load(open(os.path.join(ROOT, 'out', 'chorei-progress.json'), encoding='utf-8'))

month = data['month']; days = data['days']; upTo = data['upTo']
N = int(upTo[8:10]); DIM = 31
MAIN = data['targets']['main']; STRETCH = data['targets']['stretch']
events = {int(k): v for k, v in data.get('events', {}).items()}

AR='Arial'
H=Font(name=AR,bold=True,size=15); BOLD=Font(name=AR,bold=True); NORM=Font(name=AR,size=10)
SMALL=Font(name=AR,size=9,color='666666'); WHITE_B=Font(name=AR,bold=True,color='FFFFFF')
RED_B=Font(name=AR,bold=True,color='C00000'); BLUE=Font(name=AR,color='0000FF')
NAVY=PatternFill('solid',fgColor='1F3864'); GRAY=PatternFill('solid',fgColor='F2F2F2')
EVE=PatternFill('solid',fgColor='DDEBF7'); TSF=PatternFill('solid',fgColor='FDE9D9')
thin=Side(style='thin',color='BFBFBF'); B=Border(left=thin,right=thin,top=thin,bottom=thin)
YEN='\\¥#,##0'; PCT='0.0%'; DIFF='\\¥#,##0;[Red]\\¥-#,##0'

wb = Workbook()

# --- 日別実績 ---
ws2 = wb.active; ws2.title='日別実績'
ws2['A1']=f'日別実績（媒体別・{month[5:]}月 {upTo[8:10]}日まで）'; ws2['A1'].font=H
for i,h in enumerate(['日付','楽天','Amazon','自社','売上日計','実績累計'],1):
    c=ws2.cell(row=2,column=i,value=h); c.font=WHITE_B; c.fill=NAVY; c.border=B; c.alignment=Alignment(horizontal='center')
for d in range(1,N+1):
    iso=f'{month}-{d:02d}'; r=2+d
    v=days.get(iso,{'rakuten':0,'amazon':0,'own':0})
    ws2.cell(row=r,column=1,value=f'{int(month[5:7])}/{d}')
    ws2.cell(row=r,column=2,value=v['rakuten']); ws2.cell(row=r,column=3,value=v['amazon']); ws2.cell(row=r,column=4,value=v['own'])
    ws2.cell(row=r,column=5,value=f'=SUM(B{r}:D{r})')
    ws2.cell(row=r,column=6,value=f'=E{r}' if d==1 else f'=F{r-1}+E{r}')
    for cc in range(1,7):
        c=ws2.cell(row=r,column=cc); c.border=B; c.font=NORM
        if cc>=2: c.number_format=YEN
tot=2+N+1
ws2.cell(row=tot,column=1,value=f'計').font=BOLD
for col,letter in [(2,'B'),(3,'C'),(4,'D'),(5,'E')]:
    ws2.cell(row=tot,column=col,value=f'=SUM({letter}3:{letter}{tot-1})').font=BOLD
ws2.cell(row=tot,column=6,value=f'=F{tot-1}').font=BOLD
for cc in range(1,7):
    c=ws2.cell(row=tot,column=cc); c.border=B; c.fill=GRAY
    if cc>=2: c.number_format=YEN
for col,w in zip('ABCDEF',[9,12,12,12,13,14]): ws2.column_dimensions[col].width=w

# --- 目標乖離 ---
ws3 = wb.create_sheet('目標乖離')
ws3['A1']='日別目標との乖離（イベント日は目標を厚く配分・青字=変更可）'; ws3['A1'].font=H
ws3['A2']='重み: 通常1.0／楽天イベント1.7／タイムセール祭り2.0。イベント・重みを変えると全て自動再計算。'; ws3['A2'].font=SMALL
heads=['日付','イベント','重み','日次目標(1.1億)','累計目標(1.1億)','乖離(1.1億)','日次目標(1.2億)','累計目標(1.2億)','乖離(1.2億)','実績日計','累計実績']
for i,h in enumerate(heads,1):
    c=ws3.cell(row=3,column=i,value=h); c.font=WHITE_B; c.fill=NAVY; c.border=B; c.alignment=Alignment(horizontal='center',wrap_text=True)
WSUM=f'SUM($C$4:$C${3+DIM})'
for d in range(1,DIM+1):
    r=3+d
    ev=events.get(d,{}); label=ev.get('label',''); w=ev.get('weight',1.0)
    ws3.cell(row=r,column=1,value=f'{int(month[5:7])}/{d}')
    c=ws3.cell(row=r,column=2,value=label); c.font=BLUE
    c=ws3.cell(row=r,column=3,value=w); c.font=BLUE; c.number_format='0.0'
    ws3.cell(row=r,column=4,value=f'=ROUND(サマリー!$B$5*C{r}/{WSUM},0)')
    ws3.cell(row=r,column=5,value=f'=D{r}' if d==1 else f'=E{r-1}+D{r}')
    ws3.cell(row=r,column=7,value=f'=ROUND(サマリー!$C$5*C{r}/{WSUM},0)')
    ws3.cell(row=r,column=8,value=f'=G{r}' if d==1 else f'=H{r-1}+G{r}')
    if d<=N:
        ws3.cell(row=r,column=10,value=f'=日別実績!E{2+d}')
        ws3.cell(row=r,column=11,value=f'=J{r}' if d==1 else f'=K{r-1}+J{r}')
        ws3.cell(row=r,column=6,value=f'=K{r}-E{r}')
        ws3.cell(row=r,column=9,value=f'=K{r}-H{r}')
    for cc in range(1,12):
        c=ws3.cell(row=r,column=cc); c.border=B
        if cc not in (2,3) and (c.font is None or c.font.name!=AR): c.font=NORM
        if cc in (4,5,7,8,10,11): c.number_format=YEN
        if cc in (6,9): c.number_format=DIFF
    if label.startswith('楽天'):
        for cc in (1,2,3): ws3.cell(row=r,column=cc).fill=EVE
    if label.startswith('タイム'):
        for cc in (1,2,3): ws3.cell(row=r,column=cc).fill=TSF
r=3+DIM+1
ws3.cell(row=r,column=1,value='月計').font=BOLD
ws3.cell(row=r,column=3,value=f'={WSUM}').font=BOLD
ws3.cell(row=r,column=4,value=f'=SUM(D4:D{r-1})').font=BOLD
ws3.cell(row=r,column=7,value=f'=SUM(G4:G{r-1})').font=BOLD
for cc in range(1,12):
    c=ws3.cell(row=r,column=cc); c.border=B; c.fill=GRAY
    if cc in (4,7): c.number_format=YEN
for col,w in zip('ABCDEFGHIJK',[8,17,6,14,15,14,14,15,14,13,14]): ws3.column_dimensions[col].width=w
ws3.freeze_panes='A4'

# --- 残在庫 ---
ws4 = wb.create_sheet('残在庫')
stock = data.get('stock')
ws4['A1'] = '残在庫（在庫報告アプリ・CS出荷後の入力より）'
ws4['A1'].font = H
if stock and stock.get('rows'):
    rep = (stock.get('reportDate') or '').replace('-', '/')
    staff = stock.get('staff') or ''
    ws4['A2'] = f'報告日: {rep}' + (f'　記入者: {staff}' if staff else '')
    ws4['A2'].font = SMALL
    for i, h in enumerate(['商品', 'カラー/SKU', '残り在庫数', 'メモ'], 1):
        c = ws4.cell(row=4, column=i, value=h)
        c.font = WHITE_B; c.fill = NAVY; c.border = B
        c.alignment = Alignment(horizontal='center')
    r = 4
    for row in stock['rows']:
        r += 1
        ws4.cell(row=r, column=1, value=row.get('product', ''))
        ws4.cell(row=r, column=2, value=row.get('sku', ''))
        c = ws4.cell(row=r, column=3, value=row.get('qty', 0)); c.number_format = '#,##0'
        ws4.cell(row=r, column=4, value=row.get('memo', ''))
        for cc in range(1, 5):
            cell = ws4.cell(row=r, column=cc); cell.border = B; cell.font = NORM
    r += 1
    ws4.cell(row=r, column=1, value='合計').font = BOLD
    c = ws4.cell(row=r, column=3, value=f'=SUM(C5:C{r-1})')
    c.font = BOLD; c.number_format = '#,##0'
    for cc in range(1, 5):
        cell = ws4.cell(row=r, column=cc); cell.border = B; cell.fill = GRAY
    if stock.get('memo'):
        ws4.cell(row=r + 2, column=1, value=f'備考: {stock["memo"]}').font = SMALL
else:
    ws4['A3'] = '在庫報告アプリ（CS出荷後）にまだ入力がありません。'
    ws4['A4'] = 'CSチームが出荷後に入力すると、翌朝からこのシートに自動で載ります。'
    ws4['A3'].font = NORM; ws4['A4'].font = SMALL
for col, w in zip('ABCD', [22, 18, 12, 28]):
    ws4.column_dimensions[col].width = w

# --- 販促費 ---
ws5 = wb.create_sheet('販促費')
promo = data.get('promo')
ws5['A1'] = f'販促費（{int(month[5:7])}月・販促費管理アプリより）'
ws5['A1'].font = H
ws5['A2'] = 'Google広告・TikTok広告・案件依頼費・テレビ出演費用・PRタイムズなど。発生の都度アプリに入力すると翌朝ここに載ります。'
ws5['A2'].font = SMALL
rows5 = []
if promo and promo.get('rows'):
    # 朝礼シートはリベティ分のみ（ブランド未記入の行も含める）
    rows5 = [r for r in promo['rows'] if r.get('brand', '') in ('', 'リベティ')]
if rows5:
    # ① 費目別合計
    cats = []
    for row in rows5:
        c = row.get('category', 'その他')
        if c not in cats:
            cats.append(c)
    ws5['A4'] = '■ 費目別合計（今月）'; ws5['A4'].font = BOLD
    for i, h in enumerate(['費目', '金額', '件数'], 1):
        c = ws5.cell(row=5, column=i, value=h)
        c.font = WHITE_B; c.fill = NAVY; c.border = B; c.alignment = Alignment(horizontal='center')
    r = 5
    det_first = 5 + len(cats) + 5  # 明細の先頭データ行（下で使う）
    det_last = det_first + len(rows5) - 1
    for cat in cats:
        r += 1
        ws5.cell(row=r, column=1, value=cat)
        c = ws5.cell(row=r, column=2, value=f'=SUMIF(B{det_first}:B{det_last},A{r},C{det_first}:C{det_last})')
        c.number_format = YEN
        c2 = ws5.cell(row=r, column=3, value=f'=COUNTIF(B{det_first}:B{det_last},A{r})')
        c2.number_format = '#,##0'
        for cc in range(1, 4):
            cell = ws5.cell(row=r, column=cc); cell.border = B; cell.font = NORM
    r += 1
    ws5.cell(row=r, column=1, value='合計').font = BOLD
    c = ws5.cell(row=r, column=2, value=f'=SUM(B6:B{r-1})'); c.font = BOLD; c.number_format = YEN
    c2 = ws5.cell(row=r, column=3, value=f'=SUM(C6:C{r-1})'); c2.font = BOLD; c2.number_format = '#,##0'
    for cc in range(1, 4):
        cell = ws5.cell(row=r, column=cc); cell.border = B; cell.fill = GRAY
    # ② 明細
    hdr = det_first - 1
    ws5.cell(row=hdr - 1, column=1, value='■ 明細').font = BOLD
    for i, h in enumerate(['日付', '費目', '金額', '支払先・相手', '関連商品', '備考'], 1):
        c = ws5.cell(row=hdr, column=i, value=h)
        c.font = WHITE_B; c.fill = NAVY; c.border = B; c.alignment = Alignment(horizontal='center')
    r = hdr
    for row in rows5:
        r += 1
        d5 = (row.get('date') or '')
        ws5.cell(row=r, column=1, value=f'{int(d5[5:7])}/{int(d5[8:10])}' if len(d5) == 10 else d5)
        ws5.cell(row=r, column=2, value=row.get('category', ''))
        c = ws5.cell(row=r, column=3, value=row.get('amount', 0)); c.number_format = YEN
        ws5.cell(row=r, column=4, value=row.get('partner', ''))
        ws5.cell(row=r, column=5, value=row.get('product', ''))
        ws5.cell(row=r, column=6, value=row.get('memo', ''))
        for cc in range(1, 7):
            cell = ws5.cell(row=r, column=cc); cell.border = B; cell.font = NORM
    ws5.cell(row=r + 2, column=1, value='※ このシートは「広告費記載（全て）」アプリのリベティ分のみ。リベティのRPP/Amazon/Googleの確定実費は11時の広告費レポート（CSV自動集計）が正です。').font = SMALL
else:
    ws5['A4'] = '販促費管理アプリにまだ入力がありません。'
    ws5['A5'] = 'アプリに「計上日・費目・金額」を入力すると、翌朝からこのシートに自動で載ります。'
    ws5['A4'].font = NORM; ws5['A5'].font = SMALL
for col, w in zip('ABCDEF', [10, 16, 13, 20, 16, 30]):
    ws5.column_dimensions[col].width = w

# --- サマリー ---
ws = wb.create_sheet('サマリー',0)
ws['A1']=f'売上進捗サマリー（イベント加重・{upTo[5:].replace("-","/")}実績まで）'
ws['A1'].font=H
ws['A2']='売上=クーポン適用後。目標はイベント日加重で日割り。青字=変更可。'
ws['A2'].font=SMALL
for i,h in enumerate(['項目','目標1.1億','目標1.2億'],1):
    c=ws.cell(row=4,column=i,value=h); c.font=WHITE_B; c.fill=NAVY; c.border=B; c.alignment=Alignment(horizontal='center')
ws['B5']=MAIN; ws['C5']=STRETCH
ws['E5']='経過日数'; ws['F5']=N; ws['E6']='月日数'; ws['F6']=DIM
ws['E5'].font=SMALL; ws['E6'].font=SMALL; ws['F5'].font=BLUE; ws['F6'].font=BLUE
rows=[
    ('月間目標', None, None, YEN),
    ('本日までの目標（イベント加重）','=INDEX(目標乖離!E:E,3+$F$5)','=INDEX(目標乖離!H:H,3+$F$5)',YEN),
    ('実績累計',f'=日別実績!F{tot}',f'=日別実績!F{tot}',YEN),
    ('乖離（実績−加重目標）','=B7-B6','=C7-C6',DIFF),
    ('進捗率（対 加重目標）','=B7/B6','=C7/C6',PCT),
    ('進捗率（対 月間目標）','=B7/B5','=C7/C5',PCT),
    ('残り目標額','=B5-B7','=C5-C7',YEN),
    ('残りの重み合計',f'=目標乖離!C{3+DIM+1}-SUM(目標乖離!C4:C{3+N})','=B12','0.0'),
    ('残り期間の必要額：通常日','=B11/B12','=C11/C12',YEN),
    ('　同：楽天イベント日（×1.7）','=B13*1.7','=C13*1.7',YEN),
    ('　同：タイムセール祭り日（×2.0）','=B13*2','=C13*2',YEN),
]
for j,(nm,f1,f2,fmt) in enumerate(rows):
    r=5+j
    ws.cell(row=r,column=1,value=nm).font=NORM
    if f1 is not None: ws.cell(row=r,column=2,value=f1)
    if f2 is not None: ws.cell(row=r,column=3,value=f2)
    for cc in (1,2,3):
        c=ws.cell(row=r,column=cc); c.border=B
        if cc>1: c.number_format=fmt
    if nm=='月間目標': ws.cell(row=r,column=2).font=BLUE; ws.cell(row=r,column=3).font=BLUE
    if '乖離' in nm or '必要額' in nm: ws.cell(row=r,column=2).font=RED_B; ws.cell(row=r,column=3).font=RED_B
    if nm.startswith('進捗率'): ws.cell(row=r,column=2).font=BOLD; ws.cell(row=r,column=3).font=BOLD
for col,w in zip('ABC',[34,17,17]): ws.column_dimensions[col].width=w

out = os.path.join(ROOT,'out','売上進捗シート.xlsx')
wb.save(out)
print('saved', out)
