#!/usr/bin/env python3
# ============================================================
#  合算CPAシートの組み立て（合言葉「大怪獣ワチソンCPA」用）
# ------------------------------------------------------------
#  入力: cpa_inputs.json（キントーンから取得・復号済みのデータ。
#        リポジトリには含めない。実行環境のローカルにのみ置く）
#    {"month": "2026-08",
#     "units": {"1": {"amazon": 29, "rakuten": 49, "own": 16}, ...},
#     "ad": {"trav": {"1": 240512, ...}, "cat": {...}, "az": {...},
#            "rpp": {...}, "google": {...}},
#     "sales": {"1": 2895000, ...}}   ← スーツケース系の日別売上（税込）
#  出力: 合算CPA_<month>.xlsx
#    - サマリー（判定前提・月間平均CPA・広告比率・判定内訳）
#    - 日次（売上・広告費・広告比率・CPA・判定を先頭に、
#            媒体別広告費と媒体別個数を右側に）
#  判定基準: 客単価30,000円 × 目標広告比率15% / 許容20%
#            → 目標CPA 4,500円(🟢) / 許容CPA 6,000円(🟡) / 超過(🔴)
#            広告比率も同じ 15% / 20% 基準で色分け
#  実行: python3 scripts/buildCpaSheet.py cpa_inputs.json [出力パス]
# ============================================================
import json
import sys
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.utils import get_column_letter

UNIT_PRICE = 30000
TARGET_RATE = 0.15
ALLOW_RATE = 0.20

F = 'Yu Gothic'
NUM = '#,##0'
YEN = '¥#,##0;(¥#,##0);-'
PCT = '0.0%'


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else 'cpa_inputs.json'
    data = json.load(open(src, encoding='utf-8'))
    month = data['month']
    units = {int(k): v for k, v in data['units'].items()}
    ad = {m: {int(k): v for k, v in days.items()} for m, days in data['ad'].items()}
    sales = {int(k): v for k, v in data.get('sales', {}).items()}
    y, m = map(int, month.split('-'))
    ndays = calendar.monthrange(y, m)[1]
    last = max(units.keys()) if units else 0
    out = sys.argv[2] if len(sys.argv) > 2 else f'合算CPA_{month}.xlsx'

    hdrfill = PatternFill('solid', fgColor='188038')
    hdrfont = Font(name=F, bold=True, color='FFFFFF')
    bold = Font(name=F, bold=True)
    norm = Font(name=F)
    blue = Font(name=F, color='0000FF')
    gray = Font(name=F, size=9, color='808080')
    sumfill = PatternFill('solid', fgColor='E2EFDA')
    thin = Border(*[Side(style='thin', color='BFBFBF')] * 4)

    wb = Workbook()

    # ---------- 日次 ----------
    # 列: A日付 Bスーツケース売上 C合算広告費 D広告比率 E合算CPA F判定
    #     Gメタ HAmazon広告 IRPP JGoogle Kその他
    #     LAmazon個数 M楽天個数 N自社個数 O合計個数 P7日移動CPA
    ws = wb.create_sheet('日次')
    heads = ['日付', 'スーツケース売上', '合算広告費', '広告比率', '合算CPA', '判定',
             'メタ', 'Amazon広告', 'RPP', 'Google', 'その他',
             'Amazon個数', '楽天個数', '自社個数', '合計個数', '7日移動CPA']
    widths = [7, 15, 13, 9, 10, 6, 12, 12, 11, 10, 8, 11, 9, 9, 9, 12]
    for c, (h, w) in enumerate(zip(heads, widths), 1):
        cell = ws.cell(1, c, h)
        cell.font = hdrfont
        cell.fill = hdrfill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'B2'

    for day in range(1, ndays + 1):
        r = day + 1
        ws.cell(r, 1, f'{m}/{day}').font = norm
        meta = spend_d = qty = 0
        if day <= last:
            meta = ad.get('trav', {}).get(day, 0) + ad.get('cat', {}).get(day, 0)
            spend_d = (meta + (ad.get('az', {}).get(day) or 0)
                       + ad.get('rpp', {}).get(day, 0) + ad.get('google', {}).get(day, 0))
            qty = units[day]['amazon'] + units[day]['rakuten'] + units[day]['own']
            vals = {2: sales.get(day), 7: meta, 8: ad.get('az', {}).get(day),
                    9: ad.get('rpp', {}).get(day, 0), 10: ad.get('google', {}).get(day, 0),
                    12: units[day]['amazon'], 13: units[day]['rakuten'], 14: units[day]['own']}
            for c, v in vals.items():
                if v is None:
                    continue
                cell = ws.cell(r, c, v)
                cell.font = blue
                cell.number_format = YEN if c == 2 else NUM

        ws.cell(r, 3, f'=IF(COUNT(G{r}:K{r})=0,"",SUM(G{r}:K{r}))').number_format = YEN
        ws.cell(r, 4, f'=IFERROR(C{r}/B{r},"")').number_format = PCT
        ws.cell(r, 15, f'=IF(COUNT(L{r}:N{r})=0,"",SUM(L{r}:N{r}))').number_format = NUM
        ws.cell(r, 5, f'=IFERROR(C{r}/O{r},"")').number_format = YEN
        ws.cell(r, 6,
                f'=IF(E{r}="","",IF(E{r}<=サマリー!$B$7,"合格",IF(E{r}<=サマリー!$B$8,"注意","超過")))')
        ws.cell(r, 6).alignment = Alignment(horizontal='center')
        top = max(2, r - 6)
        ws.cell(r, 16, f'=IFERROR(SUM(C{top}:C{r})/SUM(O{top}:O{r}),"")').number_format = YEN

        # プレビュー環境でも色が出るよう、実績日には静的な色も焼き込む
        if day <= last:
            if qty:
                cpa = spend_d / qty
                chip = ('34A853' if cpa <= UNIT_PRICE * TARGET_RATE
                        else 'FBBC04' if cpa <= UNIT_PRICE * ALLOW_RATE else 'EA4335')
                ws.cell(r, 6).fill = PatternFill('solid', fgColor=chip)
                ws.cell(r, 6).font = Font(name=F, bold=True, color='FFFFFF')
                if chip == 'EA4335':
                    ws.cell(r, 5).font = Font(name=F, bold=True, color='C00000')
            if sales.get(day):
                ratio = spend_d / sales[day]
                rcol = ('34A853' if ratio <= TARGET_RATE
                        else 'BF8F00' if ratio <= ALLOW_RATE else 'C00000')
                ws.cell(r, 4).font = Font(name=F, bold=ratio > ALLOW_RATE, color=rcol)

        stripe = PatternFill('solid', fgColor='F3F6F4') if day % 2 == 0 else None
        for c in range(1, 17):
            cell = ws.cell(r, c)
            cell.border = thin
            if stripe and c != 6:
                cell.fill = stripe

    # 合計行
    r = ndays + 2
    ws.cell(r, 1, '合計').font = bold
    for c in [2, 3] + list(range(7, 16)):
        letter = get_column_letter(c)
        cell = ws.cell(r, c, f'=SUM({letter}2:{letter}{ndays + 1})')
        cell.font = bold
        cell.fill = sumfill
        cell.number_format = YEN if c in (2, 3) or c in range(7, 12) else NUM
        cell.border = thin
    for c, formula, fmt in ((4, f'=IFERROR(C{r}/B{r},"")', PCT),
                            (5, f'=IFERROR(C{r}/O{r},"")', YEN)):
        cell = ws.cell(r, c, formula)
        cell.font = bold
        cell.fill = sumfill
        cell.number_format = fmt
        cell.border = thin

    # 判定セル: スプレッドシート風の色チップ（濃色ベタ塗り＋白抜き太字）
    mrng = f'F2:F{ndays + 1}'
    ws.conditional_formatting.add(mrng, FormulaRule(
        formula=['$F2="合格"'], fill=PatternFill('solid', bgColor='34A853'),
        font=Font(name=F, bold=True, color='FFFFFF'), stopIfTrue=True))
    ws.conditional_formatting.add(mrng, FormulaRule(
        formula=['$F2="注意"'], fill=PatternFill('solid', bgColor='FBBC04'),
        font=Font(name=F, bold=True, color='FFFFFF'), stopIfTrue=True))
    ws.conditional_formatting.add(mrng, FormulaRule(
        formula=['$F2="超過"'], fill=PatternFill('solid', bgColor='EA4335'),
        font=Font(name=F, bold=True, color='FFFFFF'), stopIfTrue=True))
    # 合算CPA: データバーで大きさをひと目で
    ws.conditional_formatting.add(f'E2:E{ndays + 1}', DataBarRule(
        start_type='num', start_value=0, end_type='num', end_value=8000,
        color='F4B183', showValue=True))
    ws.conditional_formatting.add(f'E2:E{ndays + 1}', FormulaRule(
        formula=['$F2="超過"'], font=Font(name=F, bold=True, color='C00000'), stopIfTrue=False))
    # 広告比率: 目標15%以下=緑 / 許容20%以下=黄 / 超過=赤（文字色）
    prng = f'D2:D{ndays + 1}'
    ws.conditional_formatting.add(prng, FormulaRule(
        formula=['AND($D2<>"",$D2<=サマリー!$B$5)'],
        font=Font(name=F, color='34A853'), stopIfTrue=True))
    ws.conditional_formatting.add(prng, FormulaRule(
        formula=['AND($D2<>"",$D2<=サマリー!$B$6)'],
        font=Font(name=F, color='BF8F00'), stopIfTrue=True))
    ws.conditional_formatting.add(prng, FormulaRule(
        formula=['$D2<>""'],
        font=Font(name=F, bold=True, color='C00000'), stopIfTrue=True))

    ws.cell(r + 2, 1, '※スーツケース売上=売上明細のスーツケース系'
                      '（S/M/L・クラシックアルミ・多機能アルミ・アルミ型式未確認）の税込売上（全チャネル）。'
                      '広告比率=合算広告費÷スーツケース売上（目標15%・許容20%）。'
                      'メタ=トラベル+カタログ（リベティ分）。個数=同スーツケース系。'
                      'TDA・TikTok・ガジェティ等の月次手動値は未計上。').font = gray

    # ---------- サマリー ----------
    ws0 = wb.active
    ws0.title = 'サマリー'
    ws0.column_dimensions['A'].width = 32
    ws0.column_dimensions['B'].width = 16
    ws0.column_dimensions['C'].width = 46
    t = ws0.cell(1, 1, f'Libetee 合算CPA（{month}・{m}/1〜{m}/{last}実績）')
    t.font = Font(name=F, bold=True, size=14)
    ws0.cell(2, 1, '広告費（メタ+Amazon+RPP+Google）÷ スーツケース販売個数（全チャネル）').font = gray
    rows = [
        ('想定客単価（税込・全チャネル平均）', UNIT_PRICE, '手入力の前提値'),
        ('目標 広告比率', TARGET_RATE, ''),
        ('許容 広告比率', ALLOW_RATE, ''),
        ('→ 目標CPA（合格ライン）', '=B4*B5', ''),
        ('→ 許容CPA（注意ライン）', '=B4*B6', 'これを超えたら超過'),
        ('合算広告費 累計', f'=日次!C{ndays + 2}', ''),
        ('スーツケース販売個数 累計', f'=日次!O{ndays + 2}', ''),
        ('月間平均の合算CPA', '=B9/B10', ''),
        ('判定', '=IF(B11<=B7,"合格",IF(B11<=B8,"注意","超過"))', ''),
        ('合格（≤4,500円）の日数', f'=COUNTIF(日次!F2:F{ndays + 1},"合格")', ''),
        ('注意（〜6,000円）の日数', f'=COUNTIF(日次!F2:F{ndays + 1},"注意")', ''),
        ('超過（6,000円超）の日数', f'=COUNTIF(日次!F2:F{ndays + 1},"超過")', ''),
        ('スーツケース売上 累計', f'=日次!B{ndays + 2}', '税込・全チャネル'),
        ('月間 広告比率（広告費÷売上）', '=IFERROR(B9/B16,"")', '目標15%・許容20%'),
    ]
    for i, (label, val, note) in enumerate(rows):
        r = 4 + i
        ws0.cell(r, 1, label).font = bold if 'CPA' in label or '判定' in label or '比率' in label else norm
        cell = ws0.cell(r, 2, val)
        cell.font = blue if isinstance(val, (int, float)) else norm
        cell.number_format = PCT if '比率' in label else (NUM if '日数' in label or '個数' in label else YEN)
        ws0.cell(r, 3, note).font = gray
        for c in (1, 2, 3):
            ws0.cell(r, c).border = thin

    # 月間平均の判定・広告比率も静的に塗る（プレビュー対応）
    tot_spend = sum((ad.get('trav', {}).get(d, 0) + ad.get('cat', {}).get(d, 0)
                     + (ad.get('az', {}).get(d) or 0) + ad.get('rpp', {}).get(d, 0)
                     + ad.get('google', {}).get(d, 0)) for d in range(1, last + 1))
    tot_qty = sum(units[d]['amazon'] + units[d]['rakuten'] + units[d]['own']
                  for d in range(1, last + 1)) or 1
    avg = tot_spend / tot_qty
    chip = ('34A853' if avg <= UNIT_PRICE * TARGET_RATE
            else 'FBBC04' if avg <= UNIT_PRICE * ALLOW_RATE else 'EA4335')
    for rr in (11, 12):
        ws0.cell(rr, 2).fill = PatternFill('solid', fgColor=chip)
        ws0.cell(rr, 2).font = Font(name=F, bold=True, color='FFFFFF')
    tot_sales = sum(sales.get(d, 0) for d in range(1, last + 1))
    if tot_sales:
        mr = tot_spend / tot_sales
        rchip = ('34A853' if mr <= TARGET_RATE
                 else 'FBBC04' if mr <= ALLOW_RATE else 'EA4335')
        ws0.cell(17, 2).fill = PatternFill('solid', fgColor=rchip)
        ws0.cell(17, 2).font = Font(name=F, bold=True, color='FFFFFF')

    for rng in ('B11:B12',):
        ws0.conditional_formatting.add(rng, FormulaRule(
            formula=['$B$12="合格"'], fill=PatternFill('solid', bgColor='34A853'),
            font=Font(name=F, bold=True, color='FFFFFF'), stopIfTrue=True))
        ws0.conditional_formatting.add(rng, FormulaRule(
            formula=['$B$12="注意"'], fill=PatternFill('solid', bgColor='FBBC04'),
            font=Font(name=F, bold=True, color='FFFFFF'), stopIfTrue=True))
        ws0.conditional_formatting.add(rng, FormulaRule(
            formula=['$B$12="超過"'], fill=PatternFill('solid', bgColor='EA4335'),
            font=Font(name=F, bold=True, color='FFFFFF'), stopIfTrue=True))

    # 色の基準（凡例）
    lg = ws0.cell(19, 1, '■ 色の基準（判定）')
    lg.font = Font(name=F, bold=True, size=12)
    legend = [
        ('合格', '34A853', f'CPA ¥{UNIT_PRICE * TARGET_RATE:,.0f} 以下／広告比率15%以内（理想ペース）'),
        ('注意', 'FBBC04', f'CPA ¥{UNIT_PRICE * TARGET_RATE:,.0f}〜¥{UNIT_PRICE * ALLOW_RATE:,.0f}／比率20%以内（許容範囲）'),
        ('超過', 'EA4335', f'CPA ¥{UNIT_PRICE * ALLOW_RATE:,.0f} 超／比率20%超（広告費の使いすぎ）'),
    ]
    for i, (label, color, desc) in enumerate(legend):
        rr = 20 + i
        cell = ws0.cell(rr, 1, label)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(name=F, bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin
        ws0.cell(rr, 2, desc).font = norm
        ws0.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=3)
        ws0.cell(rr, 2).border = thin

    ws0.cell(24, 1, '※青字＝手入力の前提値・転記データ。それ以外は数式').font = gray

    wb.save(out)
    print(f'saved {out}')


if __name__ == '__main__':
    main()
