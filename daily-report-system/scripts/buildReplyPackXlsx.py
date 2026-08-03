#!/usr/bin/env python3
# ============================================================
#  返信パックのエクセルを作る（アルバイトさんのコピペ用）
# ------------------------------------------------------------
#  out/reply-pack.json → out/reply-pack.xlsx
#  列: No / 種別 / 商品 / 日付 / ★ / 投稿者 / レビュー / 返信文 / 要確認 / 対応チェック
# ============================================================
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

src = Path(sys.argv[1] if len(sys.argv) > 1 else 'out/reply-pack.json')
dst = Path(sys.argv[2] if len(sys.argv) > 2 else 'out/reply-pack.xlsx')
data = json.loads(src.read_text(encoding='utf-8'))
rows, meta = data['rows'], data['meta']

F = 'Yu Gothic'
def font(bold=False, size=11, color='000000'):
    return Font(name=F, bold=bold, size=size, color=color)

HDR = PatternFill('solid', fgColor='1F4E79')
WARN = PatternFill('solid', fgColor='FFE3E3')
OK = PatternFill('solid', fgColor='FFFFFF')
thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='top')

wb = Workbook()

# ---- シート1: 使い方 ----
ws0 = wb.active
ws0.title = 'はじめに'
guide = [
    ('📝 レビュー返信パック', 16, True),
    (f"対象期間: {meta['from']} 〜 {meta['to']}（未返信のみ・全{meta['counts']['total']}件）", 11, False),
    (f"　ショップレビュー {meta['counts']['shop']}件 ／ 商品レビュー {meta['counts']['item']}件 ／ ⚠要確認 {meta['counts']['needsHuman']}件", 11, False),
    ('', 11, False),
    ('【やり方（アルバイトさん向け）】', 12, True),
    ('1. 「返信リスト」シートを開き、上から順に対応します（低評価が先に並んでいます）', 11, False),
    ('2. H列の「返信文」をコピーします（セルを選んで Ctrl+C / ⌘+C）', 11, False),
    ('3. 楽天RMS →「レビューチェックツール」で該当レビューを探し、返信欄に貼り付けて投稿します', 11, False),
    ('   ※ ショップレビューと商品レビューはRMS上のタブが分かれています（B列の種別を確認）', 11, False),
    ('4. 投稿できたら J列「対応チェック」に「済」と入力します', 11, False),
    ('', 11, False),
    ('【⚠必ず守ること】', 12, True),
    ('・I列が「⚠要確認」の行（赤色）は、そのまま貼らず社員に確認してから投稿してください', 11, False),
    ('　（低評価・交換/不良などデリケートな内容です。文面はたたき台です）', 11, False),
    ('・お客様の名前や注文番号を返信に書かないでください', 11, False),
    ('・返金や無料交換など、この文面に無い約束を足さないでください', 11, False),
]
for i, (text, size, bold) in enumerate(guide, 1):
    c = ws0.cell(row=i, column=1, value=text)
    c.font = font(bold=bold, size=size)
ws0.column_dimensions['A'].width = 100

# ---- シート2: 返信リスト ----
ws = wb.create_sheet('返信リスト')
heads = ['No', '種別', '商品', '日付', '★', '投稿者', 'お客様のレビュー', '返信文（これをコピペ）', '要確認', '対応チェック']
for i, h in enumerate(heads, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = font(bold=True, color='FFFFFF')
    c.fill = HDR
    c.border = BORDER
for i, r in enumerate(rows, 1):
    rr = i + 1
    warn = r['needsHuman']
    vals = [
        i, r['kind'], r['product'] or '—', r['date'], '★' * int(r['star']),
        r['who'], r['body'], r['reply'],
        ('⚠要確認: ' + '・'.join(r['reasons'])) if warn else 'OK', '',
    ]
    for c0, v in enumerate(vals, 1):
        cell = ws.cell(row=rr, column=c0, value=v)
        cell.font = font(size=10)
        cell.border = BORDER
        if c0 in (7, 8):
            cell.alignment = WRAP
        if warn:
            cell.fill = WARN
    # 行の高さはレビューの長さに合わせてざっくり
    lines = max(len(r['body']) // 40, len(r['reply']) // 40, 3)
    ws.row_dimensions[rr].height = min(15 * lines, 220)

widths = [5, 13, 22, 11, 8, 14, 46, 46, 22, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:J{len(rows) + 1}'

wb.save(dst)
print(f'エクセル: {dst}（{len(rows)}件）')
