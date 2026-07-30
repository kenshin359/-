#!/usr/bin/env python3
# ============================================================
#  売上 簡易シート（Excel）の生成
# ------------------------------------------------------------
#  売上が記載されたときに添付する、直近14日だけの軽いシートです。
#  4シートある「売上管理シート」とは別に、その場で見る用に作ります。
#
#  実行:
#    python3 scripts/buildQuickSheet.py --date=2026-07-28
#
#  ★合計・前日比は Excel の数式で入れます。
#    数字を直したら自動で計算し直されます。
# ============================================================
import json
import os
import subprocess
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'out', '売上簡易シート.xlsx')
DAYS = 14

FONT = 'Arial'
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
HEAD_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TARGET_FILL = PatternFill('solid', fgColor='FFF2CC')
TOTAL_FILL = PatternFill('solid', fgColor='D9E2F3')
NOTE = Font(name=FONT, size=9, italic=True, color='808080')
THIN = Side(style='thin', color='BFBFBF')

YEN = '¥#,##0;(¥#,##0);-'
NUM = '#,##0;(#,##0);-'
PCT = '0.0%;(0.0%);-'
PCT2 = '0.00%;(0.00%);-'
WEEKDAY = ['月', '火', '水', '木', '金', '土', '日']


def arg(name, default=None):
    for a in sys.argv[1:]:
        if a.startswith(f'--{name}='):
            return a[len(name) + 3:]
    return default


def fetch_rows():
    """Kintone からの取得は JS 側にあるので、そちらを呼んでJSONで受け取る"""
    code = (
        "import('./lib/kintoneSalesDaily.js').then(async m => {"
        "  const rows = m.extractDailyRows(await m.fetchSalesApp());"
        "  console.log(JSON.stringify(rows));"
        "});"
    )
    r = subprocess.run(['node', '-e', code], cwd=ROOT, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f'Kintone の取得に失敗しました: {r.stderr.strip()[-300:]}')
    return json.loads(r.stdout.strip().splitlines()[-1])


def weekday_of(iso):
    y, m, d = (int(x) for x in iso.split('-'))
    return WEEKDAY[date(y, m, d).weekday()]


def main():
    target = arg('date')
    rows = fetch_rows()
    if not rows:
        print('売上データがありません')
        sys.exit(1)

    if target:
        idx = next((i for i, r in enumerate(rows) if r['date'] == target), len(rows) - 1)
    else:
        idx = len(rows) - 1
    window = rows[max(0, idx - DAYS + 1): idx + 1]

    wb = Workbook()
    ws = wb.active
    ws.title = '直近14日'

    cols = ['日付', '曜日', '楽天', 'Amazon', '自社サイト', '合計', '前日比',
            '楽天アクセス', '楽天転換率', 'お気に入り', '販売個数']
    widths = [11, 5, 13, 13, 13, 14, 9, 12, 11, 10, 10]
    for i, (c, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    r = 2
    for d in window:
        is_target = d['date'] == (target or window[-1]['date'])
        ws.cell(row=r, column=1, value=d['date']).font = BOLD if is_target else BODY
        ws.cell(row=r, column=2, value=weekday_of(d['date'])).font = BODY

        for i, key in enumerate(['rakuten', 'amazon', 'own']):
            c = ws.cell(row=r, column=3 + i, value=d['sales'].get(key) or 0)
            c.font = BODY
            c.number_format = YEN

        c = ws.cell(row=r, column=6, value=f'=SUM(C{r}:E{r})')
        c.font = BOLD
        c.number_format = YEN

        if r > 2:
            c = ws.cell(row=r, column=7, value=f'=IF(F{r-1}=0,"",F{r}/F{r-1}-1)')
            c.number_format = PCT
            c.font = BODY

        rk = (d.get('metrics') or {}).get('rakuten') or {}
        for i, (v, fmt) in enumerate([
            (rk.get('access'), NUM),
            (rk.get('cvr') / 100 if rk.get('cvr') is not None else None, PCT2),
            (rk.get('fav'), NUM),
        ]):
            c = ws.cell(row=r, column=8 + i, value=v)
            c.number_format = fmt
            c.font = BODY

        units = sum(sum(ch.values()) for ch in (d.get('units') or {}).values())
        c = ws.cell(row=r, column=11, value=units or None)
        c.number_format = NUM
        c.font = BODY

        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.border = Border(bottom=THIN)
            if is_target:
                cell.fill = TARGET_FILL
        r += 1

    last = r - 1

    # ── 合計と平均 ──
    ws.cell(row=r, column=1, value='合計').font = BOLD
    for col, letter in [(3, 'C'), (4, 'D'), (5, 'E'), (6, 'F')]:
        c = ws.cell(row=r, column=col, value=f'=SUM({letter}2:{letter}{last})')
        c.font = BOLD
        c.number_format = YEN
        c.fill = TOTAL_FILL
    c = ws.cell(row=r, column=11, value=f'=SUM(K2:K{last})')
    c.font = BOLD
    c.number_format = NUM
    c.fill = TOTAL_FILL
    for col in [1, 2, 7, 8, 9, 10]:
        ws.cell(row=r, column=col).fill = TOTAL_FILL

    r += 1
    ws.cell(row=r, column=1, value='平均').font = BOLD
    for col, letter, fmt in [(3, 'C', YEN), (4, 'D', YEN), (5, 'E', YEN), (6, 'F', YEN),
                             (8, 'H', NUM), (9, 'I', PCT2), (10, 'J', NUM), (11, 'K', NUM)]:
        c = ws.cell(row=r, column=col, value=f'=IFERROR(AVERAGE({letter}2:{letter}{last}),0)')
        c.font = BOLD
        c.number_format = fmt

    r += 2
    ws.cell(row=r, column=1, value='※ 黄色の行が、今回記載された日です。').font = NOTE
    r += 1
    ws.cell(row=r, column=1, value='※ 合計・前日比・平均は数式です。数字を直すと自動で計算し直されます。').font = NOTE
    r += 1
    ws.cell(row=r, column=1, value='※ 出典: Kintone 売上・転換率報告アプリ（読み取りのみ）').font = NOTE

    # openpyxl は数式に計算結果を持たせられないため、開いた瞬間に再計算させる
    wb.calculation.fullCalcOnLoad = True

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f'✅ 作成しました: {OUT}（{len(window)}日分）')


if __name__ == '__main__':
    main()
