#!/usr/bin/env python3
# ============================================================
#  広告費デイリー報告（毎朝11時・Chatwork）
# ------------------------------------------------------------
#  毎朝KPIアプリ(30)の添付から当月の広告費を媒体別に自動集計し、
#  チーム報告と同じ形式で朝礼ルームへ投稿します。
#  自動で取れない項目（ガジェティ/TDA/TikTok/案件費）は
#  config/chorei/adcost-extra-*.json の手動値を使います。
#  ★キントーンは読むだけ。--dry-run で本文表示のみ。
# ============================================================
import base64, csv, io, json, os, re, sys, unicodedata, urllib.request
from datetime import datetime, timedelta, timezone

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JST = timezone(timedelta(hours=9))
DRY = '--dry-run' in sys.argv

BASE = os.environ['KINTONE_BASE_URL'].rstrip('/')
AUTH = base64.b64encode(f"{os.environ['KINTONE_USER']}:{os.environ['KINTONE_PASSWORD']}".encode()).decode()
KPI_APP = os.environ.get('KINTONE_KPI_APP_ID') or '30'

def kget(path):
    req = urllib.request.Request(BASE + path, headers={'X-Cybozu-Authorization': AUTH})
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.read()

def yen_num(s):
    s = str(s).replace('￥', '').replace('¥', '').replace(',', '').replace('"', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0

def z2h(s):
    return unicodedata.normalize('NFKC', s)

def day_from_name(name):
    m = re.search(r'8[.:：月\s]?\s*([0-3]?\d)', z2h(name))
    return int(m.group(1)) if m else None

def read_text(raw):
    for enc in ('utf-8-sig', 'cp932', 'utf-16'):
        try:
            return raw.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode('utf-8', 'replace')

def sum_col(text, colname):
    rows = list(csv.reader(io.StringIO(text)))
    hdr = next((r for r in rows if any(colname in c for c in r)), None)
    if not hdr:
        return None
    i = next(j for j, c in enumerate(hdr) if colname in c)
    # 行の有効判定は合計列のセルで行う（RPP商品別CSVは先頭列が常に空のため）
    return int(sum(yen_num(r[i]) for r in rows[rows.index(hdr) + 1:]
                   if len(r) > i and str(r[i]).strip()))

def parse_date(s):
    m = re.search(r'(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})', z2h(str(s)))
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None

def daily_col(text, colname):
    """日付列（レポート開始日/日付）がある表を日ごとに合計して {(年,月,日): 金額} を返す。
    ファイル名の日付が間違っていても中身の日付で正しく配分できる。日付列が無ければ None。"""
    rows = list(csv.reader(io.StringIO(text)))
    hdr = next((r for r in rows if any(colname in c for c in r)), None)
    if not hdr:
        return None
    i = next(j for j, c in enumerate(hdr) if colname in c)
    di = next((j for j, c in enumerate(hdr) if 'レポート開始日' in c or c.strip() == '日付'), None)
    if di is None:
        return None
    out = {}
    for r in rows[rows.index(hdr) + 1:]:
        if len(r) <= max(i, di) or not str(r[i]).strip():
            continue
        d = parse_date(r[di])
        if d:
            out[d] = out.get(d, 0) + yen_num(r[i])
    return {k: int(v) for k, v in out.items()} or None

def main():
    now = datetime.now(JST)
    month = now.strftime('%Y-%m')
    upto = (now - timedelta(days=1)).day  # 前日まで

    # 当月レコードを全件取得（1回100件ずつ・最大10ページ）
    records = []
    offset = 0
    while True:
        q = urllib.parse.quote(
            f'report_date >= "{month}-01" order by report_date desc limit 100 offset {offset}')
        chunk = json.loads(kget(f'/k/v1/records.json?app={KPI_APP}&query={q}')).get('records', [])
        records.extend(chunk)
        if len(chunk) < 100 or offset >= 900:
            break
        offset += 100

    # (media, day) -> 金額。新しいレコードを先に処理し、最初の値を採用（重複添付対策）
    vals = {}
    google_daily = {}
    for rec in records:
        for field in ('file_ads', 'file_sales', 'file_other'):
            for f in rec.get(field, {}).get('value', []):
                name = f.get('name', '')
                # Macからのアップロードは濁点が分解されたNFD形式のことがあるため
                # 媒体判定はNFKC正規化した名前で行う
                nk = z2h(name)
                low = nk.lower()
                try:
                    if name.endswith('.xlsx') and 'google' in low and '日別' in nk:
                        from openpyxl import load_workbook
                        wb = load_workbook(io.BytesIO(kget(f"/k/v1/file.json?fileKey={f['fileKey']}")), data_only=True)
                        for row in wb.active.iter_rows(values_only=True):
                            if isinstance(row[0], datetime) and isinstance(row[2], (int, float)):
                                google_daily.setdefault(row[0].day, int(row[2]))
                        continue
                    if not name.endswith('.csv'):
                        continue
                    day = day_from_name(name)
                    media = None
                    colname = None
                    if 'トラベル' in nk:
                        media, colname = 'trav', '消化金額'
                    elif 'カタログ' in nk:
                        media, colname = 'cat', '消化金額'
                    elif 'rpp' in low or 'r pp' in low:
                        media, colname = 'rpp', '実績額(合計)'
                    elif ('amazon' in low or 'アマゾン' in nk) and '広告' in nk:
                        media, colname = 'az', '合計費用 (換算済み)'
                    if not media:
                        continue
                    if day and (day > upto or (media, day) in vals):
                        continue
                    text = read_text(kget(f"/k/v1/file.json?fileKey={f['fileKey']}"))
                    # Meta/RPPはCSV内の日付列で日次配分（添付名の日付間違いに強い）
                    daily = daily_col(text, colname) if media != 'az' else None
                    if daily:
                        for (yy, mm, dd), amt in daily.items():
                            if f'{yy:04d}-{mm:02d}' == month and dd <= upto \
                                    and (media, dd) not in vals:
                                vals[(media, dd)] = amt
                        continue
                    if not day or day > upto:
                        continue
                    total = sum_col(text, colname)
                    if total is not None and total >= 0:
                        vals[(media, day)] = total
                except Exception as e:  # 1ファイル失敗しても続行
                    print(f'  ⚠ {name}: {e}')

    def msum(media):
        return sum(v for (m, d), v in vals.items() if m == media)
    def mdays(media):
        return sorted(d for (m, d) in vals if m == media)

    trav, cat = msum('trav'), msum('cat')
    rpp, az = msum('rpp'), msum('az')
    google = sum(v for d, v in google_daily.items() if d <= upto)

    extra = {}
    ep = os.path.join(ROOT, 'config', 'chorei', f'adcost-extra-{month}.json')
    if os.path.exists(ep):
        extra = json.load(open(ep, encoding='utf-8'))
    gadg = extra.get('ガジェティ', {}).get('amount', 0)
    tda = extra.get('TDA', {}).get('amount', 0)
    tiktok = extra.get('TikTok広告', {}).get('amount', 0)
    anken = extra.get('案件依頼費用', {}).get('amount', 0)
    anken_note = extra.get('案件依頼費用', {}).get('note', '')
    tv = extra.get('TV等の広告費', {}).get('amount', 0)

    meta_total = trav + cat + gadg
    total = google + az + meta_total + rpp + tda + anken + tiktok + tv

    def y(n):
        return f'¥{n:,.0f}'

    # 欠測チェック（添付が無い日）
    missing = []
    for media, label in (('trav', 'トラベル'), ('cat', 'カタログ'), ('rpp', 'RPP'), ('az', 'Amazon広告')):
        got = set(mdays(media))
        miss = [d for d in range(1, upto + 1) if d not in got]
        if miss:
            missing.append(f'{label}: {"、".join(f"8/{d}" for d in miss[:6])}{" ほか" if len(miss) > 6 else ""}')
    gmiss = [d for d in range(1, upto + 1) if d not in google_daily]
    if gmiss:
        missing.append(f'Google: {"、".join(f"8/{d}" for d in gmiss[:6])}{" ほか" if len(gmiss) > 6 else ""}')

    L = []
    L.append(f'[info][title]💰 広告費レポート【8/1-8/{upto}】（自動集計）[/title]')
    L.append(f'▪️Google広告▶️{y(google)}')
    L.append('')
    L.append(f'▪️Amazon広告▶️{y(az)}')
    L.append('▪️Meta広告')
    L.append(f'トラベル▶️{y(trav)}')
    L.append(f'カタログ▶️{y(cat)}')
    L.append(f'ガジェティ▶️{y(gadg)} ※手動値')
    L.append(f'【Meta合計：{y(meta_total)}】')
    L.append('')
    L.append(f'▪️RPP▶️{y(rpp)}')
    L.append('')
    L.append(f'▪️TDA▶️{y(tda)} ※手動値')
    L.append('')
    L.append(f'▪️案件依頼費用{"（" + anken_note + "）" if anken_note else ""}')
    L.append(f'【{y(anken)}】')
    L.append('')
    L.append(f'▪️TV等の広告費▶️{"無し" if tv == 0 else y(tv)}')
    L.append('')
    L.append(f'▪️TikTok広告▶️{y(tiktok)} ※手動値')
    L.append('')
    L.append(f'合計：{y(total)}')
    if missing:
        L.append('')
        L.append('⚠添付が見つからない日: ' + ' / '.join(missing))
    L.append('[/info]')
    body = '\n'.join(L)

    print(body)
    if DRY:
        print('（--dry-run のため送信しません）')
        return
    room = os.environ.get('CHATWORK_CHOREI_ROOM_ID') or '433161347'
    token = os.environ.get('CHATWORK_API_TOKEN')
    if not token:
        raise SystemExit('CHATWORK_API_TOKEN が未設定です')
    payload = urllib.parse.urlencode({'body': body, 'self_unread': '0'}).encode()
    req = urllib.request.Request(
        f'https://api.chatwork.com/v2/rooms/{room}/messages',
        data=payload, headers={'X-ChatWorkToken': token})
    with urllib.request.urlopen(req, timeout=30) as r:
        r.read()
    print(f'✅ ルーム {room} に送信しました')

import urllib.parse  # noqa: E402
main()
