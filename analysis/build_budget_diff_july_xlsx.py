import datetime as dt, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; YEN='¥#,##0'; NUM='#,##0'; PCTF='0.00"%"'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2")
C_DIFF=PatternFill("solid",fgColor="FFE2DD")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
J=json.load(open('july_seg.json'))
# (key, 表示名, 8月日数, Δアクセス, 色)
plan=[('kabu','かぶり日（8/5・10・25）',3,10000,C_KABU),
      ('f50','5と0単独（8/15・20・30）',3,10000,C_50),
      ('won','ワンダフルデー（8/1）',1,10000,C_WON),
      ('mar','マラソン通常（8/4・6・7・8・9・24・26）',7,10000,C_MAR),
      ('hei','平日（17日）',17,5000,C_HEI)]
wb=openpyxl.Workbook()

# S1 差額まとめ
ws=wb.active; ws.title="差額まとめ_7月ベース"
ws.merge_cells("A1:J1"); ws["A1"]="通常予算 vs 増額：売上の差額【7月実績ベース・転換率キープ】"; ws["A1"].font=TITLE
ws.merge_cells("A2:J2"); ws["A2"]="転換率・客単価・販売ペース＝2026年7月実績（過去最高月）。増額＝平日+5千/イベント+1万アクセス。1アクセス価値＝7月の転換率×客単価。"; ws["A2"].font=SUB
heads=[("セグメント",34),("日数",7),("7月転換率",10),("7月客単価",11),("1アクセス\n価値",10),("ペース 売上/日",14),("増額後 売上/日",14),("差額/日",13),("月間差額",14),("月間 増額後",15)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
r=5; tb=0; ta=0
for k,name,n,da,fill in plan:
    s=J[k]; add=round(da*s['vpa'])
    ws.cell(r,1,name); ws.cell(r,2,n)
    ws.cell(r,3,s['cvr']*100).number_format=PCTF
    ws.cell(r,4,s['aov']).number_format=YEN
    ws.cell(r,5,round(s['vpa'],1))
    ws.cell(r,6,s['avg']).number_format=YEN
    ws.cell(r,7,s['avg']+add).number_format=YEN
    ws.cell(r,8,add).number_format=YEN
    ws.cell(r,9,add*n).number_format=YEN
    ws.cell(r,10,(s['avg']+add)*n).number_format=YEN
    for c in range(1,11):
        cell=ws.cell(r,c); cell.border=B; cell.fill=fill; cell.alignment=LEFT if c==1 else CEN
        if c in(8,9): cell.font=BOLD
    tb+=s['avg']*n; ta+=add*n; r+=1
ws.cell(r,1,'合計（8月・31日）').font=BOLD
ws.cell(r,6,tb).number_format=YEN; ws.cell(r,7,tb+ta).number_format=YEN
ws.cell(r,9,ta).number_format=YEN; ws.cell(r,10,tb+ta).number_format=YEN
for c in range(1,11): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
r+=2
ws.cell(r,1,'8月 月商（7月ペース継続・通常予算）').font=BOLD; ws.cell(r,3,tb).number_format=YEN
r+=1
ws.cell(r,1,'8月 月商（増額後）').font=BOLD; ws.cell(r,3,tb+ta).number_format=YEN
r+=1
ws.cell(r,1,'差額').font=font(bold=True,size=13,color='C0392B')
c=ws.cell(r,3,ta); c.number_format=YEN; c.font=font(bold=True,size=13,color='C0392B'); c.fill=C_DIFF
ws.cell(r,5,f'+{ta/tb*100:.0f}%').font=font(bold=True,color='C0392B')
r+=2
notes=['前年8月(¥26.3M)比：通常+101% / 増額後+154%。7月は過去最高月のため強気シナリオ。',
 '7月は扇風機シーズンで客単価が低く(平日¥11,488)転換率が高い。8月後半のファン需要減衰に注意。スーツケース比率が上がれば客単価は上振れ。',
 '7月後半(7/20¥3.5M・7/25¥4.2M)は2回目マラソンが走っていた可能性大。7月の公式カレンダーを頂ければ精緻化します。',
 '追加アクセス22.5万×単価¥12≒追加広告費約¥270万 → 増分ROAS約5.1倍。絶対条件はスーツケース在庫。']
for t in notes: ws.cell(r,1,t).font=SUB; r+=1

# S2 グラフ
w2=wb.create_sheet("グラフ")
w2["A1"]="グラフ（7月ベース）"; w2["A1"].font=TITLE
short=['かぶり日','5と0単独','ワンダフル','マラソン通常','平日']
w2["H1"]="セグメント"; w2["I1"]="7月ペース 売上/日"; w2["J1"]="増額後 売上/日"; w2["K1"]="月間差額"
for i,(k,name,n,da,fill) in enumerate(plan):
    s=J[k]; add=round(da*s['vpa'])
    w2.cell(2+i,8,short[i]); w2.cell(2+i,9,s['avg']); w2.cell(2+i,10,s['avg']+add); w2.cell(2+i,11,add*n)
w2["H8"]="プラン"; w2["I8"]="月商"
w2["H9"]="2025年8月実績"; w2["I9"]=26325956
w2["H10"]="通常(7月ペース)"; w2["I10"]=tb; w2["H11"]="増額後"; w2["I11"]=tb+ta
ch1=BarChart(); ch1.type="col"; ch1.title="1日あたり売上：7月ペース vs 増額後（転換率キープ）"
ch1.y_axis.title='円/日'; ch1.height=9; ch1.width=22; ch1.gapWidth=60
ch1.add_data(Reference(w2,min_col=9,max_col=10,min_row=1,max_row=6),titles_from_data=True)
ch1.set_categories(Reference(w2,min_col=8,min_row=2,max_row=6))
w2.add_chart(ch1,"A3")
ch2=BarChart(); ch2.type="bar"; ch2.title="月間差額 +¥13.8M の内訳"
ch2.height=9; ch2.width=22; ch2.legend=None
ch2.add_data(Reference(w2,min_col=11,min_row=1,max_row=6),titles_from_data=True)
ch2.set_categories(Reference(w2,min_col=8,min_row=2,max_row=6))
w2.add_chart(ch2,"A22")
ch3=BarChart(); ch3.type="col"; ch3.title="8月 月商：前年実績 → 通常(7月ペース) → 増額後"
ch3.height=9; ch3.width=13; ch3.legend=None; ch3.gapWidth=40
ch3.add_data(Reference(w2,min_col=9,min_row=8,max_row=11),titles_from_data=True)
ch3.set_categories(Reference(w2,min_col=8,min_row=9,max_row=11))
w2.add_chart(ch3,"N3")

# S3 日別明細
w3=wb.create_sheet("日別明細")
w3.merge_cells("A1:H1"); w3["A1"]="8月 日別明細（7月ベース・公式カレンダー・転換率キープ）"; w3["A1"].font=TITLE
heads=[("日付",10),("曜日",6),("セグメント",22),("Δアクセス",10),("7月ペース売上",14),("増額後売上",14),("差額",12),("メモ",24)]
for j,(h,w) in enumerate(heads,1):
    c=w3.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w3.column_dimensions[get_column_letter(j)].width=w
def akey(day):
    if day==1: return 'won','ワンダフルデー',''
    if day in(5,10,25): return 'kabu','かぶり日','★マラソン×5と0'
    if day in(4,24): return 'mar','マラソン通常','本番20:00開始'
    if day in(6,7,8,9,26): return 'mar','マラソン通常',''
    if day in(15,20,30): return 'f50','5と0単独',''
    if day==11: return 'hei','平日','マラソン①終了01:59'
    if day==27: return 'hei','平日','マラソン②終了09:59'
    return 'hei','平日',''
fillmap={'kabu':C_KABU,'f50':C_50,'won':C_WON,'mar':C_MAR,'hei':C_HEI}
r=4; t1=t2=0; d=dt.date(2026,8,1)
while d.month==8:
    k,label,memo=akey(d.day); s=J[k]
    da=10000 if k!='hei' else 5000
    add=round(da*s['vpa'])
    w3.cell(r,1,d).number_format='m/d(aaa)'; w3.cell(r,2,'月火水木金土日'[d.weekday()])
    w3.cell(r,3,label); w3.cell(r,4,da).number_format=NUM
    w3.cell(r,5,s['avg']).number_format=YEN; w3.cell(r,6,s['avg']+add).number_format=YEN
    w3.cell(r,7,add).number_format=YEN; w3.cell(r,8,memo)
    for c in range(1,9):
        cell=w3.cell(r,c); cell.border=B; cell.fill=fillmap[k]; cell.alignment=LEFT if c in(3,8) else CEN
    t1+=s['avg']; t2+=s['avg']+add; r+=1; d+=dt.timedelta(days=1)
w3.cell(r,1,'合計').font=BOLD
w3.cell(r,5,t1).number_format=YEN; w3.cell(r,6,t2).number_format=YEN; w3.cell(r,7,t2-t1).number_format=YEN
for c in range(1,9): w3.cell(r,c).fill=FT; w3.cell(r,c).border=B; w3.cell(r,c).font=BOLD
w3.freeze_panes="A4"
out="Libetee_増額シミュレーション_7月ベース.xlsx"
wb.save(out)
print('saved',out)
print(f'検算: まとめ 通常{tb:,} 増額後{tb+ta:,} 差額{ta:,}')
print(f'検算: 日別  通常{t1:,} 増額後{t2:,} 差額{t2-t1:,}')
print('一致' if (tb==t1 and tb+ta==t2) else '不一致!')
