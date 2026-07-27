import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series
F="Arial"; YEN='¥#,##0'; NUM='#,##0'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2")
C_DIFF=PatternFill("solid",fgColor="FFE2DD")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")

# セグメント実測（2026・転換率キープ）
segs=[('かぶり日（8/5・10・25）',3,30000,40000,2234584,1339254,C_KABU),
      ('5と0単独（8/15・20・30）',3,30000,40000,1631874,925842,C_50),
      ('ワンダフルデー（8/1）',1,30000,40000,1081629,1072855,C_WON),
      ('マラソン通常（8/4・6・7・8・9・24・26）',7,30000,40000,824846,630136,C_MAR),
      ('平日（17日）',17,10000,15000,740103,251218,C_HEI)]
wb=openpyxl.Workbook()

# ===== S1: 差額まとめ =====
ws=wb.active; ws.title="差額まとめ"
ws.merge_cells("A1:I1"); ws["A1"]="通常予算 vs 増額：売上の差額（転換率キープ・2026年8月公式カレンダー）"; ws["A1"].font=TITLE
ws.merge_cells("A2:I2"); ws["A2"]="平日1万→1.5万（+50%）／イベント日3万→4万アクセス。メタ広告（スーツケース）で調整。1アクセス価値＝2026実績の転換率×客単価。"; ws["A2"].font=SUB
heads=[("セグメント",34),("日数",7),("現行\nアクセス",10),("増額\nアクセス",10),("現行 売上/日",13),("増額後 売上/日",14),("差額/日",13),("月間差額",14),("月間 増額後",15)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
r=5; tc=0; ta=0
for name,n,a0,a1,cur,add,fill in segs:
    ws.cell(r,1,name); ws.cell(r,2,n)
    ws.cell(r,3,a0).number_format=NUM; ws.cell(r,4,a1).number_format=NUM
    ws.cell(r,5,cur).number_format=YEN; ws.cell(r,6,cur+add).number_format=YEN
    ws.cell(r,7,add).number_format=YEN; ws.cell(r,8,add*n).number_format=YEN
    ws.cell(r,9,(cur+add)*n).number_format=YEN
    for c in range(1,10):
        cell=ws.cell(r,c); cell.border=B; cell.fill=fill; cell.alignment=LEFT if c==1 else CEN
        if c in(7,8): cell.font=BOLD
    tc+=cur*n; ta+=add*n; r+=1
ws.cell(r,1,'合計（8月・31日）').font=BOLD
ws.cell(r,5,tc).number_format=YEN; ws.cell(r,6,tc+ta).number_format=YEN
ws.cell(r,7,'').number_format=YEN; ws.cell(r,8,ta).number_format=YEN; ws.cell(r,9,tc+ta).number_format=YEN
for c in range(1,10): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
tr=r
r+=2
ws.cell(r,1,'月商：通常予算').font=BOLD; ws.cell(r,2,tc).number_format=YEN; ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
r+=1
ws.cell(r,1,'月商：増額後').font=BOLD; ws.cell(r,2,tc+ta).number_format=YEN; ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
r+=1
ws.cell(r,1,'差額').font=font(bold=True,size=13,color='C0392B')
c=ws.cell(r,2,ta); c.number_format=YEN; c.font=font(bold=True,size=13,color='C0392B'); c.fill=C_DIFF
ws.cell(r,5,f'+{ta/tc*100:.0f}%').font=font(bold=True,color='C0392B')
r+=2
for t in ['前提：転換率キープ（理論値上限）。追加アクセス22.5万/月 × 単価¥12 ≒ 追加広告費 約¥270万 → 増分ROAS 約6.1倍。',
          '絶対条件：スーツケース在庫（追加注文 約+700件/月）。10月の在庫切れ▲44%を繰り返さないこと。',
          'かぶり日は+¥134万/日と価値最大。増額幅に上限がある場合は「かぶり日3日を死守、マラソン通常日と平日を削る」。']:
    ws.cell(r,1,t).font=SUB; r+=1

# ===== S2: グラフ =====
w2=wb.create_sheet("グラフ")
w2["A1"]="グラフ（Excel/Googleスプレッドシートで表示）"; w2["A1"].font=TITLE
# グラフ用データ（非表示気味に右側へ）
w2["H1"]="セグメント"; w2["I1"]="現行 売上/日"; w2["J1"]="増額後 売上/日"; w2["K1"]="月間差額"
short=['かぶり日','5と0単独','ワンダフル','マラソン通常','平日']
for i,(name,n,a0,a1,cur,add,fill) in enumerate(segs):
    w2.cell(2+i,8,short[i]); w2.cell(2+i,9,cur); w2.cell(2+i,10,cur+add); w2.cell(2+i,11,add*n)
w2["H8"]="プラン"; w2["I8"]="月商"
w2["H9"]="通常予算"; w2["I9"]=tc; w2["H10"]="増額後"; w2["I10"]=tc+ta
# チャート1：1日あたり 現行vs増額
ch1=BarChart(); ch1.type="col"; ch1.title="1日あたり売上：通常予算 vs 増額後（転換率キープ）"
ch1.y_axis.title='円/日'; ch1.height=9; ch1.width=22; ch1.gapWidth=60
data=Reference(w2,min_col=9,max_col=10,min_row=1,max_row=6)
cats=Reference(w2,min_col=8,min_row=2,max_row=6)
ch1.add_data(data,titles_from_data=True); ch1.set_categories(cats)
w2.add_chart(ch1,"A3")
# チャート2：月間差額の内訳
ch2=BarChart(); ch2.type="bar"; ch2.title="月間差額 +¥16.5M の内訳（セグメント別）"
ch2.x_axis.title=''; ch2.height=9; ch2.width=22; ch2.legend=None
d2=Reference(w2,min_col=11,min_row=1,max_row=6)
ch2.add_data(d2,titles_from_data=True); ch2.set_categories(cats)
w2.add_chart(ch2,"A22")
# チャート3：月商比較
ch3=BarChart(); ch3.type="col"; ch3.title="8月 月商：通常予算 → 増額後"
ch3.height=9; ch3.width=11; ch3.legend=None; ch3.gapWidth=40
d3=Reference(w2,min_col=9,min_row=8,max_row=10)
c3=Reference(w2,min_col=8,min_row=9,max_row=10)
ch3.add_data(d3,titles_from_data=True); ch3.set_categories(c3)
w2.add_chart(ch3,"N3")

# ===== S3: 日別明細 =====
w3=wb.create_sheet("日別明細")
w3.merge_cells("A1:H1"); w3["A1"]="8月 日別明細（通常予算 vs 増額・転換率キープ）"; w3["A1"].font=TITLE
heads=[("日付",10),("曜日",6),("セグメント",22),("現行アクセス",11),("増額アクセス",11),("現行売上",13),("増額後売上",13),("差額",12)]
for j,(h,w) in enumerate(heads,1):
    c=w3.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w3.column_dimensions[get_column_letter(j)].width=w
segmap={}
for name,n,a0,a1,cur,add,fill in segs: segmap[name.split('（')[0]]=(a0,a1,cur,add,fill)
def key(day):
    if day==1: return 'ワンダフルデー'
    if day in(5,10,25): return 'かぶり日'
    if day in(4,6,7,8,9,24,26): return 'マラソン通常'
    if day in(15,20,30): return '5と0単独'
    return '平日'
r=4; d=dt.date(2026,8,1); t1=t2=0
while d.month==8:
    k=key(d.day); a0,a1,cur,add,fill=segmap[k]
    w3.cell(r,1,d).number_format='m/d(aaa)'; w3.cell(r,2,'月火水木金土日'[d.weekday()])
    w3.cell(r,3,k); w3.cell(r,4,a0).number_format=NUM; w3.cell(r,5,a1).number_format=NUM
    w3.cell(r,6,cur).number_format=YEN; w3.cell(r,7,cur+add).number_format=YEN; w3.cell(r,8,add).number_format=YEN
    for c in range(1,9):
        cell=w3.cell(r,c); cell.border=B; cell.fill=fill; cell.alignment=LEFT if c==3 else CEN
    t1+=cur; t2+=cur+add; r+=1; d+=dt.timedelta(days=1)
w3.cell(r,1,'合計').font=BOLD
w3.cell(r,6,t1).number_format=YEN; w3.cell(r,7,t2).number_format=YEN; w3.cell(r,8,t2-t1).number_format=YEN
for c in range(1,9): w3.cell(r,c).fill=FT; w3.cell(r,c).border=B; w3.cell(r,c).font=BOLD
w3.freeze_panes="A4"

out="Libetee_増額シミュレーション.xlsx"
wb.save(out)
print('saved',out)
print(f'検算: 通常{tc:,} 増額後{tc+ta:,} 差額{ta:,} / 日別合計 通常{t1:,} 増額後{t2:,} 差額{t2-t1:,}')
print('一致' if (tc==t1 and tc+ta==t2) else '不一致!')
