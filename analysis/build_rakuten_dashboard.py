import json, datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

days=json.load(open('rakuten_clean.json'))
for d in days: d['d']=dt.date.fromisoformat(d['date'])
days.sort(key=lambda x:x['d'])

def primary(dt_):
    day=dt_.day; mo=dt_.month
    if mo in (3,6,9,12) and 4<=day<=11: return 'スーパーSALE'
    if day%5==0: return '5と0のつく日'
    if day==1: return 'ワンダフルデー'
    return '平常'
for d in days: d['ev']=primary(d['d'])

FONT="Arial"; YEN='¥#,##0'; PCT='0.00"%"'; NUM='#,##0'; XR='0.00"x"'; DATEF='yyyy/mm/dd'
BLUE=Font(name=FONT,color="0000FF"); BLACK=Font(name=FONT,color="000000"); GREEN=Font(name=FONT,color="008000")
HEADW=Font(name=FONT,color="FFFFFF",bold=True,size=10); TITLE=Font(name=FONT,bold=True,size=14)
SUB=Font(name=FONT,italic=True,size=9,color="555555"); BOLD=Font(name=FONT,bold=True); SECW=Font(name=FONT,bold=True,color="FFFFFF")
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2"); FS=PatternFill("solid",fgColor="2E75B6")
FWARN=PatternFill("solid",fgColor="FFC7CE"); FGOOD=PatternFill("solid",fgColor="C6EFCE")
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
wb=openpyxl.Workbook()

# ===== シート1：日次データ =====
ws=wb.active; ws.title="日次データ"; MAIN="日次データ"
ws.merge_cells("A1:H1"); ws["A1"]="楽天 日次データ（2025年7〜11月・実績）"; ws["A1"].font=TITLE
heads=[("日付",12),("曜日",7),("イベント種別",14),("売上金額",13),("売上件数",9),("アクセス人数",12),("転換率",9),("客単価",11)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=BORDER; ws.column_dimensions[get_column_letter(j)].width=w
DS=4
for i,d in enumerate(days):
    r=DS+i
    ws.cell(r,1,d['d']).number_format=DATEF
    ws.cell(r,2,d['dow']); ws.cell(r,3,d['ev'])
    ws.cell(r,4,round(d['sales'])).number_format=YEN
    ws.cell(r,5,int(d['orders'])).number_format=NUM
    ws.cell(r,6,int(d['access'])).number_format=NUM
    ws.cell(r,7,d['cvr']).number_format=PCT
    ws.cell(r,8,round(d['aov'])).number_format=YEN
    for c in range(1,9): ws.cell(r,c).border=BORDER; ws.cell(r,c).font=BLACK
DE=DS+len(days)-1
ws.freeze_panes="A4"
def rng(col): return f"'{MAIN}'!${col}${DS}:${col}${DE}"

# ===== シート2：月次トレンド =====
w2=wb.create_sheet("月次トレンド")
w2.merge_cells("A1:H1"); w2["A1"]="月次トレンド（楽天・実績）"; w2["A1"].font=TITLE
h2=[("月",12),("月商",15),("日数",7),("日平均",13),("件数",8),("客単価",11),("総アクセス",12),("前月比",10)]
for j,(h,w) in enumerate(h2,1):
    c=w2.cell(2,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=BORDER; w2.column_dimensions[get_column_letter(j)].width=w
mons=['2025-07','2025-08','2025-09','2025-10','2025-11']
# 月の開始/終了日
def mstart(ym): y,m=map(int,ym.split('-')); return dt.date(y,m,1)
def mend(ym): y,m=map(int,ym.split('-')); return (dt.date(y+(m//12),(m%12)+1,1)-dt.timedelta(days=1))
for i,ym in enumerate(mons):
    r=3+i; ms=mstart(ym); me=mend(ym)
    w2.cell(r,1,ym).font=GREEN
    crit=f'{rng("A")},">="&DATE({ms.year},{ms.month},{ms.day}),{rng("A")},"<="&DATE({me.year},{me.month},{me.day})'
    w2.cell(r,2,f'=SUMIFS({rng("D")},{crit})').number_format=YEN
    w2.cell(r,3,f'=COUNTIFS({crit})').number_format=NUM
    w2.cell(r,4,f'=IFERROR(B{r}/C{r},0)').number_format=YEN
    w2.cell(r,5,f'=SUMIFS({rng("E")},{crit})').number_format=NUM
    w2.cell(r,6,f'=IFERROR(B{r}/E{r},0)').number_format=YEN
    w2.cell(r,7,f'=SUMIFS({rng("F")},{crit})').number_format=NUM
    w2.cell(r,8,(f'=IFERROR(B{r}/B{r-1}-1,0)' if i>0 else '="—"')).number_format='0.0%'
    for c in range(1,9): w2.cell(r,c).border=BORDER; 
    if i>0: w2.cell(r,c-0)
# 前月比の色（悪化=赤）
w2.conditional_formatting.add(f"H4:H7", CellIsRule(operator="lessThan",formula=["-0.15"],fill=FWARN))
# 合計行
tr=3+len(mons)
w2.cell(tr,1,"合計/平均").font=BOLD
w2.cell(tr,2,f'=SUM(B3:B{tr-1})').number_format=YEN
w2.cell(tr,3,f'=SUM(C3:C{tr-1})').number_format=NUM
w2.cell(tr,4,f'=IFERROR(B{tr}/C{tr},0)').number_format=YEN
w2.cell(tr,5,f'=SUM(E3:E{tr-1})').number_format=NUM
w2.cell(tr,6,f'=IFERROR(B{tr}/E{tr},0)').number_format=YEN
w2.cell(tr,7,f'=SUM(G3:G{tr-1})').number_format=NUM
for c in range(1,9): w2.cell(tr,c).fill=FT; w2.cell(tr,c).font=BOLD; w2.cell(tr,c).border=BORDER

# ===== シート3：イベント効果（実測倍率）=====
w3=wb.create_sheet("イベント効果")
w3.merge_cells("A1:E1"); w3["A1"]="イベント効果（実測倍率・平常日=1.00）"; w3["A1"].font=TITLE
w3["A2"]="この倍率が予想の心臓部。仮値ではなく実績から算出。"; w3["A2"].font=SUB
h3=[("イベント種別",16),("日平均売上",15),("日数",8),("リフト倍率",12),("平常日比+",12)]
for j,(h,w) in enumerate(h3,1):
    c=w3.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=BORDER; w3.column_dimensions[get_column_letter(j)].width=w
types=['平常','5と0のつく日','ワンダフルデー','スーパーSALE']
r0=4
for i,t in enumerate(types):
    r=r0+i
    w3.cell(r,1,t).font=GREEN
    w3.cell(r,2,f'=IFERROR(AVERAGEIFS({rng("D")},{rng("C")},$A{r}),0)').number_format=YEN
    w3.cell(r,3,f'=COUNTIFS({rng("C")},$A{r})').number_format=NUM
    w3.cell(r,4,f'=IFERROR(B{r}/$B$4,0)').number_format=XR
    w3.cell(r,5,f'=IFERROR(B{r}/$B$4-1,0)').number_format='+0%;-0%'
    for c in range(1,6): w3.cell(r,c).border=BORDER; w3.cell(r,c).font=BLACK if c!=1 else GREEN
w3.cell(r0,1,"平常").font=BOLD

# ===== シート4：曜日別 =====
w4=wb.create_sheet("曜日別")
w4.merge_cells("A1:C1"); w4["A1"]="曜日別 日平均売上（実績）"; w4["A1"].font=TITLE
for j,(h,w) in enumerate([("曜日",10),("日平均売上",15),("日数",8)],1):
    c=w4.cell(2,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=BORDER; w4.column_dimensions[get_column_letter(j)].width=w
for i,wd in enumerate(['月','火','水','木','金','土','日']):
    r=3+i; w4.cell(r,1,wd).font=GREEN
    w4.cell(r,2,f'=IFERROR(AVERAGEIFS({rng("D")},{rng("B")},$A{r}),0)').number_format=YEN
    w4.cell(r,3,f'=COUNTIFS({rng("B")},$A{r})').number_format=NUM
    for c in range(1,4): w4.cell(r,c).border=BORDER; w4.cell(r,c).font=BLACK if c!=1 else GREEN

# ===== シート5：予想（実測倍率ベース）=====
w5=wb.create_sheet("予想")
for col,w in zip("ABCD",[28,18,16,20]): w5.column_dimensions[col].width=w
w5.merge_cells("A1:D1"); w5["A1"]="次のイベント日 売上予想（実測ベース）"; w5["A1"].font=TITLE
w5["A3"]="平常日ベースライン（実測）"; w5["A3"].font=BOLD
w5.cell(3,2,f"='イベント効果'!B4").number_format=YEN; w5.cell(3,2).border=BORDER
w5.cell(4,1,"予想したいイベント種別を選択→").font=BLUE
from openpyxl.worksheet.datavalidation import DataValidation
dv=DataValidation(type="list",formula1='"5と0のつく日,ワンダフルデー,スーパーSALE"',allow_blank=True)
w5.add_data_validation(dv); sel=w5.cell(4,2,"スーパーSALE"); sel.font=BLUE; sel.fill=PatternFill("solid",fgColor="FFF7CC"); sel.border=BORDER; dv.add("B4")
w5.cell(6,1,"採用リフト倍率（実測）").font=BOLD
w5.cell(6,2,"=IFERROR(INDEX('イベント効果'!$D$4:$D$7,MATCH($B$4,'イベント効果'!$A$4:$A$7,0)),1)").number_format=XR; w5.cell(6,2).border=BORDER
w5.cell(7,1,"本命 予想売上（単日）").font=BOLD
w5.cell(7,2,"=B3*B6").number_format=YEN; w5.cell(7,2).border=BORDER; w5.cell(7,2).fill=FGOOD
w5.cell(8,1,"弱気(−20%)").font=BLACK; w5.cell(8,2,"=B7*0.8").number_format=YEN; w5.cell(8,2).border=BORDER
w5.cell(9,1,"強気(+20%)").font=BLACK; w5.cell(9,2,"=B7*1.2").number_format=YEN; w5.cell(9,2).border=BORDER
w5.cell(11,1,"※ 倍率は「イベント効果」シートの実測値に自動連動。データが増えれば精度も上がる。").font=SUB

# ===== シート6：読み方 =====
w6=wb.create_sheet("読み方"); w6.column_dimensions["A"].width=100
lines=[
 ("楽天分析ダッシュボード（2025年7〜11月・実績）の読み方",TITLE),("",None),
 ("【最重要】9月→10月で月商が44%落ちた。原因はアクセス（集客）の半減。",BOLD),
 ("  客単価・転換率は安定〜改善 → 商品や価格ではなく、集客（広告・イベント露出）の問題。",None),
 ("  打ち手：イベント日への広告集中と、平常日の集客底上げ。",None),("",None),
 ("【実測の倍率（予想の心臓部）】平常日=1.00に対し",BOLD),
 ("  スーパーSALE ×2.99 / 5と0のつく日 ×2.23 / ワンダフルデー ×1.93",None),
 ("  平常日ベースライン ¥547,647/日（実測112日平均）",None),("",None),
 ("【シート】",BOLD),
 ("  日次データ … 153日の実績（売上/件数/アクセス/転換率/客単価/イベント種別）",None),
 ("  月次トレンド … 月商推移と前月比（悪化は赤）",None),
 ("  イベント効果 … 実測の種別別リフト倍率",None),
 ("  曜日別 … 日曜が最強・金曜が最弱",None),
 ("  予想 … 種別を選ぶと 実測倍率 × ベースライン で単日予想",None),("",None),
 ("色：青字＝入力 / 黒字＝自動計算 / 緑字＝参照。数値はすべて楽天の実データ。",SUB),
]
for i,(t,f) in enumerate(lines,1):
    c=w6.cell(i,1,t); c.font=f if f else Font(name=FONT); c.alignment=Alignment(wrap_text=True,vertical="center")

out="/tmp/claude-0/-home-user--/c7179f16-7eb7-52e8-aa4a-c082e5dbce5f/scratchpad/Libetee_楽天分析_2025.xlsx"
wb.save(out); print("saved:",out,"| 日次",DS,"-",DE)
