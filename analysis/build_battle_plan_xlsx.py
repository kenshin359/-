import json, datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; NUM='#,##0'; YEN='¥#,##0'; P3='0.000%'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
H_B4=PatternFill("solid",fgColor="2E75B6"); H_AF=PatternFill("solid",fgColor="C55A11")
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2")
C_SC=PatternFill("solid",fgColor="D6E4F0"); C_FAN=PatternFill("solid",fgColor="DDEBD8"); C_DIFF=PatternFill("solid",fgColor="FFE2DD")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
J=json.load(open('july_seg_official.json'))
SC_AOV=29500; FAN_AOV=3500; DROP=0.8; CPC=12
SEG={}
for k,(a0,a1) in {'hei':(20000,30000),'mar':(30000,50000),'won':(30000,50000),'f50':(30000,50000),'kabu':(30000,50000)}.items():
    s=J[k]; od=s['avg']/s['aov']; q=max(0,min(1,(s['aov']-FAN_AOV)/(SC_AOV-FAN_AOV)))
    sc_o=od*q; cvr0=sc_o/a0; cvr1=cvr0*DROP
    b4=round(sc_o*SC_AOV); af=round(a1*cvr1*SC_AOV); fan=s['avg']-b4
    SEG[k]=dict(a0=a0,a1=a1,cvr0=cvr0,cvr1=cvr1,b4=b4,af=af,fan=fan)
wb=openpyxl.Workbook()

# ===== S1 7月実績 =====
ws=wb.active; ws.title="7月実績"
ws.merge_cells("A1:E1"); ws["A1"]="7月実績：スーツケース／ハンディファン（SKUデータ・7/1〜7/28・税込）"; ws["A1"].font=TITLE
for j,(h,w) in enumerate([("商品",26),("販売個数",11),("売上(税込)",15),("実効単価",12),("備考",34)],1):
    c=ws.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
data=[('スーツケース 計',1126,38059460,C_SC,'売上の73%。多機能PCが96%',True),
 ('  多機能PC',1079,35830200,C_SC,'最重要SKU＝マットブラックS(203個)。S:M:L=57:31:11',False),
 ('  ノーマルアルミ(クラシック)',32,1410260,C_SC,'',False),
 ('  多機能アルミ(フルアルミ)',14,779200,C_SC,'',False),
 ('  アウトドアSC',1,39800,C_SC,'',False),
 ('ハンディファン 計',2933,13289900,C_FAN,'売上の26%。数量はSCの2.6倍',True),
 ('  首振り(3Way冷却)',1577,7743440,C_FAN,'ファンの54%',False),
 ('  スケルトン',717,4124740,C_FAN,'',False),
 ('  クリップ',551,1256280,C_FAN,'',False),
 ('  ミニファン',88,165440,C_FAN,'ほぼ動きなし',False)]
r=4
for name,u,s,fill,memo,tot in data:
    ws.cell(r,1,name); ws.cell(r,2,u).number_format=NUM
    ws.cell(r,3,s).number_format=YEN; ws.cell(r,4,round(s/u)).number_format=YEN
    ws.cell(r,5,memo).font=SUB
    for c in range(1,6):
        cell=ws.cell(r,c); cell.border=B; cell.fill=fill
        if tot: cell.font=BOLD
    r+=1
r+=1
ws.cell(r,1,'7月店舗全体(税抜・7/1-26)：¥43,920,694（過去最高ペース）').font=SUB

# ===== S2 8月勝負日 =====
w2=wb.create_sheet("8月勝負日_そのままvs攻め")
w2.merge_cells("A1:M1"); w2["A1"]="8月イベント勝負日（日付順）：そのまま vs 攻めパターン"; w2["A1"].font=TITLE
w2.merge_cells("A2:M2"); w2["A2"]="スーツケース広告の設定。そのまま＝平日2万/イベント3万・転換率7月実績。攻め＝平日3万/イベント5万・転換率×0.8。広告費＝アクセス×¥12(メタ実績単価)。売上/日＝SC+FAN(FAN据え置き)。"; w2["A2"].font=SUB
w2.merge_cells("D4:G4"); w2["D4"]="そのままパターン"
w2.merge_cells("H4:K4"); w2["H4"]="攻めパターン（転換率×0.8）"
for rng,fill in [("D4:G4",H_B4),("H4:K4",H_AF)]:
    for row in w2[rng]:
        for c in row: c.fill=fill; c.font=HEADW; c.alignment=CEN; c.border=B
heads=[("日付",10),("イベント",22),("日数",6),
 ("アクセス",10),("SC転換率",10),("広告費/日",11),("売上/日",12),
 ("アクセス",10),("SC転換率",10),("広告費/日",11),("売上/日",12),
 ("差額/日",12),("期間差額",13)]
for j,(h,w) in enumerate(heads,1):
    c=w2.cell(5,j,h); c.font=HEADW; c.alignment=CEN; c.border=B
    c.fill=FH if j<=3 else (H_B4 if j<=7 else (H_AF if j<=11 else FH))
    w2.column_dimensions[get_column_letter(j)].width=w
battles=[('8/1(土)','ワンダフルデー','won',1,C_WON),
 ('8/4(火)','マラソン① 本番開始 20:00〜','mar',1,C_MAR),
 ('8/5(水)','★かぶり：マラソン①×5と0','kabu',1,C_KABU),
 ('8/6〜9','マラソン① 中盤','mar',4,C_MAR),
 ('8/10(月)','★かぶり：マラソン①×5と0','kabu',1,C_KABU),
 ('8/15(土)','5と0のつく日','f50',1,C_50),
 ('8/20(木)','5と0のつく日','f50',1,C_50),
 ('8/24(月)','マラソン② 本番開始 20:00〜','mar',1,C_MAR),
 ('8/25(火)','★かぶり：マラソン②×5と0','kabu',1,C_KABU),
 ('8/26(水)','マラソン② 中盤','mar',1,C_MAR),
 ('8/30(日)','5と0のつく日','f50',1,C_50),
 ('平日17日','(2,3,11〜14,16〜19,21〜23,27〜29,31)','hei',17,C_HEI)]
r=6; t_b4=t_af=t_ad0=t_ad1=0
for date,name,k,n,fill in battles:
    s=SEG[k]
    d_b4=s['b4']+s['fan']; d_af=s['af']+s['fan']
    ad0=s['a0']*CPC; ad1=s['a1']*CPC
    w2.cell(r,1,date); w2.cell(r,2,name); w2.cell(r,3,n)
    w2.cell(r,4,s['a0']).number_format=NUM
    w2.cell(r,5,s['cvr0']).number_format=P3
    w2.cell(r,6,ad0).number_format=YEN
    w2.cell(r,7,d_b4).number_format=YEN
    w2.cell(r,8,s['a1']).number_format=NUM
    w2.cell(r,9,s['cvr1']).number_format=P3
    w2.cell(r,10,ad1).number_format=YEN
    w2.cell(r,11,d_af).number_format=YEN
    w2.cell(r,12,d_af-d_b4).number_format=YEN
    w2.cell(r,13,(d_af-d_b4)*n).number_format=YEN
    for c in range(1,14):
        cell=w2.cell(r,c); cell.border=B; cell.fill=fill; cell.alignment=LEFT if c==2 else CEN
        if c in(12,13): cell.font=BOLD
    t_b4+=d_b4*n; t_af+=d_af*n; t_ad0+=ad0*n; t_ad1+=ad1*n
    r+=1
w2.cell(r,1,'月間合計').font=BOLD
w2.cell(r,6,t_ad0).number_format=YEN; w2.cell(r,7,t_b4).number_format=YEN
w2.cell(r,10,t_ad1).number_format=YEN; w2.cell(r,11,t_af).number_format=YEN
w2.cell(r,12,'').number_format=YEN; w2.cell(r,13,t_af-t_b4).number_format=YEN
for c in range(1,14): w2.cell(r,c).fill=FT; w2.cell(r,c).border=B; w2.cell(r,c).font=BOLD
r+=2
w2.cell(r,1,'差額まとめ').font=font(bold=True,size=12,color='C0392B')
r+=1
rows_sum=[('売上差額（攻め−そのまま）',t_af-t_b4,C_DIFF),
 ('広告費差額（攻め−そのまま）',t_ad1-t_ad0,None),
 ('純増（売上差額−広告費差額）',(t_af-t_b4)-(t_ad1-t_ad0),C_DIFF)]
for label,v,fill in rows_sum:
    w2.cell(r,1,label).font=BOLD
    c=w2.cell(r,4,v); c.number_format=YEN; c.font=BOLD
    if fill: c.fill=fill
    r+=1
w2.cell(r,1,f'増分ROAS＝売上差額÷広告費差額＝{(t_af-t_b4)/(t_ad1-t_ad0):.1f}倍（転換率×0.8の保守前提。×0.9なら{16717450/(t_ad1-t_ad0):.1f}倍）').font=SUB
r+=1
w2.cell(r,1,'★勝負日の頂点は8/5・8/10・8/25のかぶり3日：攻めで1日¥508万（+¥110万/日）。在庫最厚で臨むこと。').font=SUB
w2.freeze_panes="A6"

# ===== S3 グラフ =====
w3=wb.create_sheet("グラフ")
w3["A1"]="グラフ"; w3["A1"].font=TITLE
w3["H1"]="勝負日"; w3["I1"]="そのまま"; w3["J1"]="攻め"
gl=[('ワンダフル8/1','won'),('かぶり8/5','kabu'),('マラソン中盤','mar'),('かぶり8/10','kabu'),('5と0 8/15','f50'),('かぶり8/25','kabu'),('平日','hei')]
for i,(n,k) in enumerate(gl):
    s=SEG[k]; w3.cell(2+i,8,n); w3.cell(2+i,9,s['b4']+s['fan']); w3.cell(2+i,10,s['af']+s['fan'])
ch=BarChart(); ch.type="col"; ch.title="勝負日別 売上/日：そのまま vs 攻め（SC+FAN）"
ch.height=10; ch.width=24; ch.gapWidth=50; ch.y_axis.title='円/日'
ch.add_data(Reference(w3,min_col=9,max_col=10,min_row=1,max_row=8),titles_from_data=True)
ch.set_categories(Reference(w3,min_col=8,min_row=2,max_row=8))
w3.add_chart(ch,"A3")
w3["H12"]="項目"; w3["I12"]="そのまま"; w3["J12"]="攻め"
w3["H13"]="月商"; w3["I13"]=t_b4; w3["J13"]=t_af
w3["H14"]="広告費"; w3["I14"]=t_ad0; w3["J14"]=t_ad1
ch2=BarChart(); ch2.type="col"; ch2.title="月間：月商と広告費"
ch2.height=10; ch2.width=16; ch2.gapWidth=60
ch2.add_data(Reference(w3,min_col=9,max_col=10,min_row=12,max_row=14),titles_from_data=True)
ch2.set_categories(Reference(w3,min_col=8,min_row=13,max_row=14))
w3.add_chart(ch2,"N3")
out="Libetee_8月勝負日プラン.xlsx"
wb.save(out); print('saved',out)
print(f'月商: そのまま{t_b4:,} → 攻め{t_af:,} (差額+{t_af-t_b4:,})')
print(f'広告費: {t_ad0:,} → {t_ad1:,} (+{t_ad1-t_ad0:,})')
print(f'純増: {(t_af-t_b4)-(t_ad1-t_ad0):,} / 増分ROAS {(t_af-t_b4)/(t_ad1-t_ad0):.2f}')
print('検証: 月商一致' if (t_b4==47273319 and t_af==57911421) else f'NG {t_b4} {t_af}')
