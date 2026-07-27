import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; NUM='#,##0'; YEN='¥#,##0'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_SC=PatternFill("solid",fgColor="D6E4F0"); C_FAN=PatternFill("solid",fgColor="DDEBD8")
C_HOT=PatternFill("solid",fgColor="F4B183"); C_WARN=PatternFill("solid",fgColor="FFC7CE"); C_OTH=PatternFill("solid",fgColor="F2F2F2")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
G=json.load(open('sku_july.json'))
SC_UNIT=30728  # 実効単価 税抜(33,801/1.1)
# 8月シナリオ(税抜・SC増額プランより)
SCB_S=37996670; SCA_S=54714120
u_b4=round(SCB_S/SC_UNIT); u_af=round(SCA_S/SC_UNIT)
fan_aug=round(2933/28*31)
wb=openpyxl.Workbook()

# S1 7月実績
ws=wb.active; ws.title="7月SKU実績"
ws.merge_cells("A1:E1"); ws["A1"]="7月 SKU別売上 実績（2026/7/1〜7/28・税込）"; ws["A1"].font=TITLE
heads=[("商品",26),("販売個数",10),("売上(税込)",14),("実効単価",11),("備考",30)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
order=[('多機能PC',C_SC,'スーツケース売上の94%。主力中の主力'),
 ('ノーマルアルミ(クラシック)',C_SC,''),('多機能アルミ(フルアルミ)',C_SC,''),('アウトドアSC',C_SC,''),
 ('首振り(3Way冷却)',C_FAN,'ファンの54%'),('スケルトン',C_FAN,''),('クリップ',C_FAN,''),('ミニファン',C_FAN,''),
 ('ツヤリス',C_OTH,''),('圧縮バッグ(セット)',C_OTH,'★SCとセットで1000円OFF：既に17セット/月'),
 ('圧縮バッグ',C_OTH,''),('洗顔ブラシ',C_OTH,''),('レンタル',C_OTH,''),('モバイルバッテリー',C_OTH,'')]
r=4
for name,fill,memo in order:
    u,s=G[name]
    ws.cell(r,1,name); ws.cell(r,2,round(u)).number_format=NUM
    ws.cell(r,3,round(s)).number_format=YEN
    ws.cell(r,4,round(s/u) if u else 0).number_format=YEN
    ws.cell(r,5,memo).font=SUB
    for c in range(1,6): ws.cell(r,c).border=B; ws.cell(r,c).fill=fill
    r+=1
SC=['多機能PC','ノーマルアルミ(クラシック)','多機能アルミ(フルアルミ)','アウトドアSC']
FAN=['首振り(3Way冷却)','スケルトン','クリップ','ミニファン']
scu=sum(G[k][0] for k in SC); scs=sum(G[k][1] for k in SC)
fu=sum(G[k][0] for k in FAN); fs=sum(G[k][1] for k in FAN)
for label,u,s,fill in [('スーツケース計',scu,scs,C_SC),('ハンディファン計',fu,fs,C_FAN)]:
    ws.cell(r,1,label).font=BOLD; ws.cell(r,2,round(u)).number_format=NUM
    ws.cell(r,3,round(s)).number_format=YEN; ws.cell(r,4,round(s/u)).number_format=YEN
    for c in range(1,6): ws.cell(r,c).fill=fill; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
    r+=1
r+=1
for t in ['税整合：SKU計(税込)¥51,793,660 ÷1.1 = ¥47.1M(税抜) ≒ 日次レポート(税抜)と一致。分析全体の裏付けが取れた。',
 '発見①：スーツケースは実は多機能PCがほぼ全て（個数の95.8%）。前回の広告費ベース仮定(74%)を訂正。',
 '発見②：ファンの実効単価は¥4,531と想定(¥3,500)より高い＝ファンの貢献も上方修正。',
 '発見③：圧縮バッグのSCセット(1000円OFF)が既に月17セット売れている→セット割戦略の実績あり。']:
    ws.cell(r,1,t).font=SUB; r+=1

# S2 多機能PC カラー・サイズ
w2=wb.create_sheet("多機能PC_カラーサイズ別")
w2.merge_cells("A1:D1"); w2["A1"]="多機能PC カラー・サイズ別実績（7月・上位）"; w2["A1"].font=TITLE
for j,(h,w) in enumerate([("SKU(カラー×サイズ)",26),("個数",9),("売上(税込)",14),("メモ",26)],1):
    c=w2.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w2.column_dimensions[get_column_letter(j)].width=w
tops=[('マットブラック × S',203,5919400,'★絶対戦略在庫。全SCの18%'),
 ('マットシルバー × S',114+10,3347200+295000,''),('エナメルシルバー × S',100+10,2933000+292000,''),
 ('マットブラック × M',88+9,3179400+326200,''),('マットシルバー × M',70+5,2538000+180000,''),
 ('マットホワイト × S',61+8,1792800+235400,''),('マットホワイト × M',48+3,1747400+109400,''),
 ('マットシルバー × L',33+6,1422400+258800,''),('エナメルカーキ × M',34+8,1230200+292400,''),
 ('マットグレー × L',29+3,1261200+130400,'')]
r=4
for name,u,s,memo in tops:
    w2.cell(r,1,name); w2.cell(r,2,u).number_format=NUM; w2.cell(r,3,s).number_format=YEN; w2.cell(r,4,memo).font=SUB
    for c in range(1,5): w2.cell(r,c).border=B
    if '★' in memo:
        for c in range(1,5): w2.cell(r,c).fill=C_HOT
    r+=1
r+=1
w2.cell(r,1,'サイズ構成（多機能PC）: S 57%（616個）／ M 31%（334個）／ L 11%（123個）').font=BOLD; r+=1
w2.cell(r,1,'カラーはマット系（黒・シルバー・白）で7割。エナメル系は2割強。').font=SUB

# S3 8月個数計画
w3=wb.create_sheet("8月個数計画")
w3.merge_cells("A1:F1"); w3["A1"]="8月 販売個数計画（実SKU構成比×増額プラン）"; w3["A1"].font=TITLE
w3.merge_cells("A2:F2"); w3["A2"]="SC個数＝シナリオ売上÷実効単価¥30,728(税抜)。構成比は7月実績。FANは7月ペース据え置き(31日換算)。"; w3["A2"].font=SUB
for j,(h,w) in enumerate([("商品",26),("構成比\n(7月実績)",11),("8月個数\n(増額前)",11),("8月個数\n(増額後)",11),("在庫目安",11),("メモ",28)],1):
    c=w3.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w3.column_dimensions[get_column_letter(j)].width=w
pcs=G['多機能PC'][0]/scu; na=G['ノーマルアルミ(クラシック)'][0]/scu; ma=G['多機能アルミ(フルアルミ)'][0]/scu
items=[('スーツケース計',1.0,u_b4,u_af,1.2,C_SC,'かぶり日3日だけで約480個(増額後)',True),
 ('  多機能PC',pcs,None,None,1.2,C_SC,'S:M:L≒57:31:11で発注',False),
 ('    うち マットブラックS',0.181,None,None,1.3,C_HOT,'★最重要SKU。安全率1.3',False),
 ('  ノーマルアルミ(クラシック)',na,None,None,1.2,C_SC,'',False),
 ('  多機能アルミ(フルアルミ)',ma,None,None,1.2,C_SC,'',False),
 ('ハンディファン計',1.0,fan_aug,fan_aug,1.1,C_FAN,'アクセス据え置き。8月後半の需要減に注意',True),
 ('  首振り(3Way冷却)',0.5377,None,None,1.1,C_FAN,'',False),
 ('  スケルトン',0.2445,None,None,1.1,C_FAN,'',False),
 ('  クリップ',0.1879,None,None,1.1,C_FAN,'',False),
 ('  ミニファン',0.03,None,None,1.1,C_FAN,'',False)]
r=5
for name,share,b4,af,safe,fill,memo,is_tot in items:
    if is_tot: base_b4,base_af=b4,af
    ub=b4 if b4 is not None else round(base_b4*share)
    ua=af if af is not None else round(base_af*share)
    w3.cell(r,1,name)
    w3.cell(r,2,share).number_format='0.0%'
    w3.cell(r,3,ub).number_format=NUM
    w3.cell(r,4,ua).number_format=NUM
    w3.cell(r,5,round(ua*safe)).number_format=NUM
    w3.cell(r,6,memo).font=SUB
    for c in range(1,7):
        cell=w3.cell(r,c); cell.border=B; cell.fill=fill
        if is_tot: cell.font=BOLD
    r+=1
r+=1
for t in [f'スーツケース 増額前{u_b4:,}個 → 増額後{u_af:,}個（+{u_af-u_b4:,}個）。実効単価が想定より高く、前回試算(1,855個)から個数は微減。',
 '★マットブラックSだけで8月約320個(増額後)。ここを切らすと山が崩れる——在庫目安420個。',
 'かぶり日(8/5・10・25)は1日あたりSC約160個出る計算。マラソン②(8/24〜)分の入庫は8月中旬まで必須。']:
    w3.cell(r,1,t).font=SUB; r+=1

# S4 グラフ
w4=wb.create_sheet("グラフ")
w4["A1"]="グラフ"; w4["A1"].font=TITLE
w4["H1"]="商品"; w4["I1"]="7月個数"
gl=[('多機能PC',G['多機能PC'][0]),('首振り',G['首振り(3Way冷却)'][0]),('スケルトン',G['スケルトン'][0]),
    ('クリップ',G['クリップ'][0]),('ミニファン',G['ミニファン'][0]),('ノーマルアルミ',G['ノーマルアルミ(クラシック)'][0]),('多機能アルミ',G['多機能アルミ(フルアルミ)'][0])]
for i,(n,u) in enumerate(gl):
    w4.cell(2+i,8,n); w4.cell(2+i,9,round(u))
ch=BarChart(); ch.type="bar"; ch.title="7月 商品別販売個数（実績）"
ch.height=10; ch.width=22; ch.legend=None
ch.add_data(Reference(w4,min_col=9,min_row=1,max_row=8),titles_from_data=True)
ch.set_categories(Reference(w4,min_col=8,min_row=2,max_row=8))
w4.add_chart(ch,"A3")
out="Libetee_SKU実績_8月個数計画.xlsx"
wb.save(out); print('saved',out)
print(f'SC 8月個数: 増額前{u_b4:,} → 増額後{u_af:,} / FAN {fan_aug:,}')
