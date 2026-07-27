import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
F="Arial"; YEN='¥#,##0'; XR='0.00"x"'; DATEF='m/d(aaa)'; NUM='#,##0'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")

# 2026年実績（セグメント日平均・CVR・客単価）
SEG={'won': dict(label='ワンダフルデー',sales=1081629,lift=1.46,cvr=0.0035,aov=30653,fill=C_WON),
     'kabu':dict(label='マラソン×5と0 かぶり',sales=2234584,lift=3.02,cvr=0.0054,aov=24801,fill=C_KABU),
     'mar': dict(label='お買い物マラソン',sales=824846,lift=1.11,cvr=0.0026,aov=24236,fill=C_MAR),
     'f50': dict(label='5と0のつく日',sales=1631874,lift=2.20,cvr=0.0046,aov=20127,fill=C_50),
     'hei': dict(label='平日',sales=740103,lift=1.00,cvr=0.0022,aov=22838,fill=C_HEI)}
# 公式カレンダー（画像より）
def day_info(day):
    memo=''
    if day==1: return 'won','ワンダフルデー',''
    if day==2: return 'hei','平日','マラソン①プレ 10:00〜'
    if day==4: return 'mar','マラソン①本番','20:00スタート（初日夜）'
    if day in(6,7,8,9): return 'mar','マラソン①本番',''
    if day in(5,10): return 'kabu','マラソン①×5と0','★最強かぶり日'
    if day==11: return 'hei','平日','マラソン①終了 01:59'
    if day in(15,20,30): return 'f50','5と0のつく日',''
    if day==18: return 'hei','平日','ご愛顧感謝デーP4倍(会員限定・データ無につき平日扱い)'
    if day==22: return 'hei','平日','マラソン②プレ 10:00〜'
    if day==24: return 'mar','マラソン②本番','20:00スタート（初日夜）'
    if day==25: return 'kabu','マラソン②×5と0','★最強かぶり日(公式で新発見)'
    if day==26: return 'mar','マラソン②本番',''
    if day==27: return 'hei','平日','マラソン②終了 09:59'
    return 'hei','平日',''

wb=openpyxl.Workbook()
ws=wb.active; ws.title="8月スケジュール_公式"
ws.merge_cells("A1:J1"); ws["A1"]="2026年8月 楽天スケジュール（★公式カレンダー確定版）＋アクセス増シミュレーション"; ws["A1"].font=TITLE
ws.merge_cells("A2:J2"); ws["A2"]="公式：マラソン①8/4 20:00〜8/11 01:59／マラソン②8/24 20:00〜8/27 09:59。現行=平日1万/イベント3万アクセス → 新プラン=平日1.5万/イベント4万。"; ws["A2"].font=SUB
heads=[("日付",10),("曜日",6),("公式イベント",20),("セグメント",20),("平日比",8),("現行予想売上",13),("現行\nアクセス",10),("新\nアクセス",10),("追加売上\n(CVR維持)",13),("メモ",30)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
r=4; tot=0; add_tot=0; counts={}
d=dt.date(2026,8,1)
while d.month==8:
    key,ev,memo=day_info(d.day); s=SEG[key]
    is_event = key!='hei'
    acc_now = 30000 if is_event else 10000
    acc_new = 40000 if is_event else 15000
    vpa = s['cvr']*s['aov']
    add = round((acc_new-acc_now)*vpa)
    ws.cell(r,1,d).number_format=DATEF; ws.cell(r,2,'月火水木金土日'[d.weekday()])
    ws.cell(r,3,ev); ws.cell(r,4,s['label']); ws.cell(r,5,s['lift']).number_format=XR
    ws.cell(r,6,s['sales']).number_format=YEN
    ws.cell(r,7,acc_now).number_format=NUM; ws.cell(r,8,acc_new).number_format=NUM
    ws.cell(r,9,add).number_format=YEN; ws.cell(r,10,memo)
    for c in range(1,11):
        cell=ws.cell(r,c); cell.border=B; cell.fill=s['fill']; cell.alignment=LEFT if c in(3,4,10) else CEN
    tot+=s['sales']; add_tot+=add; counts[key]=counts.get(key,0)+1
    r+=1; d+=dt.timedelta(days=1)
ws.cell(r,1,'合計').font=BOLD; ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
ws.cell(r,6,tot).number_format=YEN
ws.cell(r,9,add_tot).number_format=YEN
ws.cell(r,10,f'新プラン月商={tot+round(add_tot*0.6):,}〜{tot+add_tot:,}円')
for c in range(1,11): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
ws.freeze_panes="A4"

# シミュレーション要約シート
w2=wb.create_sheet("シミュレーション")
w2.column_dimensions['A'].width=44; w2.column_dimensions['B'].width=20; w2.column_dimensions['C'].width=42
w2.merge_cells("A1:C1"); w2["A1"]="アクセス増シミュレーション（平日+50%＝1.5万／イベント日4万・スーツケース調整）"; w2["A1"].font=TITLE
rows=[('前提',None,None),
 ('  現行アクセス',f'平日10,000 / イベント30,000','社長運用値'),
 ('  新プラン',f'平日15,000(+50%) / イベント40,000','メタ広告(スーツケース)で調整'),
 ('  1アクセスの価値(2026実績: 転換率×客単価)','平日¥50/かぶり¥134/5と0¥93/マラソン¥63/ワンダフル¥107','セグメント別に適用'),
 ('結果（8月・公式カレンダー）',None,None),
 ('  現行プラン 月商予測',f'¥{tot:,}','前年8月比 +18%'),
 ('  追加売上(上限：CVRが維持できた場合)',f'+¥{add_tot:,}','平日+¥25万/日、かぶり日+¥134万/日 等'),
 ('  追加売上(現実：追加流入のCVRは平均の6割想定)',f'+¥{round(add_tot*0.6):,}','メタ経由の新規はアプリ常連より転換低め'),
 ('  → 新プラン 月商予測',f'¥{tot+round(add_tot*0.6):,} 〜 ¥{tot+add_tot:,}','約¥41M〜¥47.6M'),
 ('コストとリターン',None,None),
 ('  追加アクセス数','225,000/月','平日5千×17日＋イベント1万×14日'),
 ('  追加メタ広告費(アクセス単価¥12)','約¥270万/月','7/25実績の単価¥10.5〜13.4より'),
 ('  増分ROAS','3.7〜6.1倍','追加売上÷追加広告費'),
 ('最重要の前提＝在庫',None,None),
 ('  追加注文数(推定)','+420〜700件/月','スーツケース中心。10月の在庫切れを繰り返さないこと'),
 ('  かぶり日(8/5・10・25)は特に','在庫と広告を最厚に','ここが月商の山')]
r=3
for a,b,c in rows:
    if b is None:
        w2.cell(r,1,a).font=font(bold=True,color='FFFFFF'); w2.cell(r,1).fill=FH
        for cc in(2,3): w2.cell(r,cc).fill=FH
    else:
        w2.cell(r,1,a); w2.cell(r,2,b).font=BOLD; w2.cell(r,3,c).font=SUB
    for cc in(1,2,3): w2.cell(r,cc).border=B
    r+=1
out="Libetee_8月スケジュール_公式.xlsx"
wb.save(out)
print(f'saved {out}')
print(f'日数内訳: {counts}')
print(f'現行予測 ¥{tot:,} (前年比{tot/26325956-1:+.0%})')
print(f'追加売上 上限+¥{add_tot:,} / 現実(60%)+¥{round(add_tot*0.6):,}')
print(f'新プラン月商 ¥{tot+round(add_tot*0.6):,} 〜 ¥{tot+add_tot:,}')
