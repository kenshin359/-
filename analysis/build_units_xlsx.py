import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; NUM='#,##0'; NUM1='#,##0.0'; YEN='¥#,##0'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2")
C_SC=PatternFill("solid",fgColor="D6E4F0"); C_FAN=PatternFill("solid",fgColor="DDEBD8"); C_IN=PatternFill("solid",fgColor="FFF2CC")
C_WARN=PatternFill("solid",fgColor="FFC7CE")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
J=json.load(open('july_seg_official.json'))
SC_AOV=29500; FAN_AOV=3500; DROP=0.9
plan=[('hei','平日',17,20000,30000,C_HEI),
      ('mar','イベント(5,0抜き)：マラソン通常',7,30000,50000,C_MAR),
      ('won','イベント(5,0抜き)：ワンダフルデー',1,30000,50000,C_WON),
      ('f50','5と0のつく日(単独)',3,30000,50000,C_50),
      ('kabu','イベント＋5,0(かぶり)',3,30000,50000,C_KABU)]
rows=[]
for k,name,n,a0,a1,fill in plan:
    s=J[k]; od=s['avg']/s['aov']; q=max(0,min(1,(s['aov']-FAN_AOV)/(SC_AOV-FAN_AOV)))
    u0=od*q; u1=a1*(u0/a0)*DROP; f=(s['avg']-u0*SC_AOV)/FAN_AOV
    rows.append(dict(name=name,n=n,fill=fill,u0=u0,u1=u1,f=f))
SCB=sum(r['u0']*r['n'] for r in rows); SCA=sum(r['u1']*r['n'] for r in rows)
FANU=sum(r['f']*r['n'] for r in rows)
wb=openpyxl.Workbook()

# S1 販売個数まとめ
ws=wb.active; ws.title="販売個数まとめ"
ws.merge_cells("A1:I1"); ws["A1"]="商品別 販売個数（8月予測・7月公式ベース）"; ws["A1"].font=TITLE
ws.merge_cells("A2:I2"); ws["A2"]="個数＝売上÷客単価（スーツケース¥29,500／ファン¥3,500）。増額後＝平日3万/イベント5万アクセス・転換率−10%。ファンは据え置き。"; ws["A2"].font=SUB
heads=[("区分",30),("日数",6),("SC個/日\n(増額前)",10),("SC個/日\n(増額後)",10),("SC月間\n(増額前)",10),("SC月間\n(増額後)",10),("増加個数",10),("FAN個/日",10),("FAN月間",10)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
r=5
for x in rows:
    ws.cell(r,1,x['name']); ws.cell(r,2,x['n'])
    ws.cell(r,3,round(x['u0'],1)).number_format=NUM1
    ws.cell(r,4,round(x['u1'],1)).number_format=NUM1
    ws.cell(r,5,round(x['u0']*x['n'])).number_format=NUM
    ws.cell(r,6,round(x['u1']*x['n'])).number_format=NUM
    ws.cell(r,7,round((x['u1']-x['u0'])*x['n'])).number_format=NUM
    ws.cell(r,8,round(x['f'],1)).number_format=NUM1
    ws.cell(r,9,round(x['f']*x['n'])).number_format=NUM
    for c in range(1,10):
        cell=ws.cell(r,c); cell.border=B; cell.fill=x['fill']; cell.alignment=LEFT if c==1 else CEN
        if c==7: cell.font=BOLD
    r+=1
ws.cell(r,1,'合計').font=BOLD
ws.cell(r,5,round(SCB)).number_format=NUM; ws.cell(r,6,round(SCA)).number_format=NUM
ws.cell(r,7,round(SCA-SCB)).number_format=NUM; ws.cell(r,9,round(FANU)).number_format=NUM
for c in range(1,10): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
r+=2
for t in [f'スーツケース：増額前 約{round(SCB):,}個/月 → 増額後 約{round(SCA):,}個/月（+{round(SCA-SCB):,}個）',
 f'ハンディファン：約{round(FANU):,}個/月（据え置き）',
 f'検算：SC個数×¥29,500＝¥{round(SCB*SC_AOV):,} ＝ SC売上と整合。',
 '※個数は注文件数ベース。ファンはまとめ買い（1注文複数個）でやや過小の可能性。セット割導入時はSC個数が上振れ。']:
    ws.cell(r,1,t).font=SUB; r+=1

# S2 SKU別と在庫目安
w2=wb.create_sheet("SKU別個数と在庫目安")
w2.merge_cells("A1:G1"); w2["A1"]="SKU別 販売個数と8月の在庫目安"; w2["A1"].font=TITLE
w2.merge_cells("A2:G2"); w2["A2"]="SKU構成比＝7/25広告費配分(自社+Amazon+メタ)からの仮定（黄色セル＝要調整）。在庫目安＝増額後×安全率(SC1.2/FAN1.1)。"; w2["A2"].font=SUB
heads=[("商品",22),("構成比\n(仮定)",10),("月間個数\n(増額前)",11),("月間個数\n(増額後)",11),("在庫目安\n(8月)",11),("うち かぶり日3日\n(8/5・10・25)",16),("メモ",26)]
for j,(h,w) in enumerate(heads,1):
    c=w2.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w2.column_dimensions[get_column_letter(j)].width=w
kabu_after=next(r for r in rows if 'かぶり' in r['name'])
sc_kabu=round(kabu_after['u1']*3)
skus=[('スーツケース',None,None,None,C_SC,''),
 ('  多機能PC',0.74,1.2,C_SC,None,'主力。セット割対象'),
 ('  ノーマルアルミ',0.15,1.2,C_SC,None,'セット割対象'),
 ('  多機能アルミ',0.11,1.2,C_SC,None,'セット割対象'),
 ('ハンディファン',None,None,None,C_FAN,''),
 ('  首振り',0.39,1.1,C_FAN,None,''),
 ('  スケルトン',0.39,1.1,C_FAN,None,''),
 ('  ミニファン',0.11,1.1,C_FAN,None,''),
 ('  クリップ',0.11,1.1,C_FAN,None,'')]
r=5
fan_kabu=round(kabu_after['f']*3)
for item in skus:
    name=item[0]
    if item[1] is None:
        is_sc = 'スーツ' in name
        tot0=SCB if is_sc else FANU; tot1=SCA if is_sc else FANU
        w2.cell(r,1,name).font=BOLD
        w2.cell(r,3,round(tot0)).number_format=NUM
        w2.cell(r,4,round(tot1)).number_format=NUM
        w2.cell(r,5,round(tot1*(1.2 if is_sc else 1.1))).number_format=NUM
        w2.cell(r,6,sc_kabu if is_sc else fan_kabu).number_format=NUM
        for c in range(1,8): w2.cell(r,c).fill=item[4]; w2.cell(r,c).border=B; w2.cell(r,c).font=BOLD
    else:
        share,safe,fill=item[1],item[2],item[3]
        is_sc='PC' in name or 'アルミ' in name
        tot0=SCB if is_sc else FANU; tot1=SCA if is_sc else FANU
        kab=sc_kabu if is_sc else fan_kabu
        w2.cell(r,1,name)
        c2=w2.cell(r,2,share); c2.number_format='0%'; c2.fill=C_IN
        w2.cell(r,3,round(tot0*share)).number_format=NUM
        w2.cell(r,4,round(tot1*share)).number_format=NUM
        w2.cell(r,5,round(tot1*share*safe)).number_format=NUM
        w2.cell(r,6,round(kab*share)).number_format=NUM
        w2.cell(r,7,item[5]).font=SUB
        for c in (1,3,4,5,6,7): w2.cell(r,c).fill=fill; w2.cell(r,c).border=B
        w2.cell(r,2).border=B
    r+=1
r+=1
for t in ['★かぶり日3日だけでスーツケース約504個(增額後)。ここで欠品すると+¥1,652万の山が消える——10月の教訓。',
 'Amazon在庫55個(多機能アルミ・7/25時点)のような水準ではかぶり日1日も持たない。8月中旬までにマラソン②分の入庫を。',
 '構成比(黄色)は広告費配分による仮定。RMS「商品別売上CSV」を頂ければ実数に置き換えて確定します。']:
    w2.cell(r,1,t).font=SUB; r+=1

# S3 グラフ
w3=wb.create_sheet("グラフ")
w3["A1"]="グラフ"; w3["A1"].font=TITLE
w3["H1"]="区分"; w3["I1"]="増額前"; w3["J1"]="増額後"
labels=['平日','マラソン通常','ワンダフル','5と0単独','かぶり日']
for i,x in enumerate(rows):
    w3.cell(2+i,8,labels[i]); w3.cell(2+i,9,round(x['u0']*x['n'])); w3.cell(2+i,10,round(x['u1']*x['n']))
ch=BarChart(); ch.type="col"; ch.title="スーツケース月間販売個数：増額前 vs 増額後（セグメント別）"
ch.height=10; ch.width=24; ch.gapWidth=60; ch.y_axis.title='個'
ch.add_data(Reference(w3,min_col=9,max_col=10,min_row=1,max_row=6),titles_from_data=True)
ch.set_categories(Reference(w3,min_col=8,min_row=2,max_row=6))
w3.add_chart(ch,"A3")

out="Libetee_販売個数と在庫目安.xlsx"
wb.save(out); print('saved',out)
print(f'SC {round(SCB):,}→{round(SCA):,}個 / FAN {round(FANU):,}個 / かぶり3日SC {sc_kabu}個')
