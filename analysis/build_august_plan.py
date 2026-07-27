import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
F="Arial"; YEN='¥#,##0'; PCT='0.00"%"'; XR='0.00"x"'; DATEF='m/d(aaa)'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
# セグメント色
C_KABU=PatternFill("solid",fgColor="F4B183")   # かぶり=オレンジ濃
C_KABU2=PatternFill("solid",fgColor="C0392B")
C_50=PatternFill("solid",fgColor="C6EFCE")      # 5と0=緑
C_WON=PatternFill("solid",fgColor="BDD7EE")     # ワンダフル=青
C_MAR=PatternFill("solid",fgColor="E2EFDA")     # マラソン普通=薄緑
C_HEI=PatternFill("solid",fgColor="F2F2F2")     # 平日=グレー
C_IN=PatternFill("solid",fgColor="FFF2CC")      # 入力=黄
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center"); RIGHT=Alignment(horizontal="right",vertical="center")
wb=openpyxl.Workbook()

# ===== シート1：8月 広告配分カレンダー =====
ws=wb.active; ws.title="8月_広告配分カレンダー"
ws.merge_cells("A1:H1"); ws["A1"]="2026年8月 楽天 広告配分カレンダー（色付き）"; ws["A1"].font=TITLE
ws.merge_cells("A2:H2"); ws["A2"]="色＝セグメント。予想売上＝2026年平日¥740,103に各セグメント倍率を適用。マラソン日程は【推定】→公式で要置換。"; ws["A2"].font=SUB
heads=[("日付",11),("曜日",6),("セグメント",22),("5と0",7),("平日比",8),("予想売上",13),("広告優先",10),("メモ",26)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
BASE=740103
# 8月の各日を分類
def seg_for(d):
    day=d.day
    mar = 4<=day<=11              # 推定マラソン
    is50 = day in(5,10,15,20,25,30)
    if day==1: return('ワンダフルデー',1.68,'A',C_WON,'月初ポイント。上乗せ')
    if mar and is50: return('マラソン×5と0 かぶり',3.01,'S 最優先',C_KABU,'★最強。広告を最も厚く')
    if mar: return('お買い物マラソン(推定)',1.18,'C 盛らない',C_MAR,'普通日は平日並み。厚くしない')
    if day in(15,20,25,30): return('5と0のつく日',2.06,'A',C_50,'転換率が上がる。厚めに')
    return('平日',1.00,'底維持',C_HEI,'切らない。薄い一定額')
r=4; tot=0
d=dt.date(2026,8,1)
while d.month==8:
    name,lift,pri,fill,memo=seg_for(d)
    sales=round(BASE*lift)
    ws.cell(r,1,d).number_format=DATEF
    ws.cell(r,2,'月火水木金土日'[d.weekday()])
    ws.cell(r,3,name); ws.cell(r,4,'●' if d.day in(5,10,15,20,25,30) else '')
    ws.cell(r,5,lift).number_format=XR
    ws.cell(r,6,sales).number_format=YEN
    ws.cell(r,7,pri); ws.cell(r,8,memo)
    for c in range(1,9):
        cell=ws.cell(r,c); cell.border=B; cell.fill=fill
        cell.alignment=LEFT if c in(3,8) else CEN
    tot+=sales; r+=1; d+=dt.timedelta(days=1)
# 合計
ws.cell(r,1,'月商予測').font=BOLD; ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
ws.cell(r,6,tot).number_format=YEN; ws.cell(r,6).font=BOLD
ws.cell(r,7,'前年比').font=BOLD; ws.cell(r,8,f'2025/8=¥26.3M → {tot/26325956-1:+.0%}')
for c in range(1,9): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B
ws.freeze_panes="A4"

# ===== シート2：セグメント別統計（数値ベタ焼き）=====
w2=wb.create_sheet("セグメント別統計")
w2.merge_cells("A1:F1"); w2["A1"]="楽天 セグメント別 売上・転換率（13か月実績・平日=1.00）"; w2["A1"].font=TITLE
h2=[("セグメント",26),("日数",7),("日平均売上",14),("平日比",9),("転換率",9),("客単価",12)]
for j,(h,w) in enumerate(h2,1):
    c=w2.cell(2,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w2.column_dimensions[get_column_letter(j)].width=w
rows=[('スーパーSALE × 5と0 かぶり',8,2942893,4.71,0.74,29651,C_KABU2,True),
      ('マラソン × 5と0 かぶり',18,1882548,3.01,0.73,22075,C_KABU,True),
      ('5と0のつく日（単独）',50,1286398,2.06,0.48,21277,C_50,False),
      ('ワンダフルデー',13,1049789,1.68,0.61,21324,C_WON,False),
      ('スーパーSALE（5と0以外）',24,1018230,1.63,0.32,29337,C_MAR,False),
      ('お買い物マラソン（5と0以外）',54,736866,1.18,0.40,19103,C_MAR,False),
      ('平日',224,624430,1.00,0.26,22093,C_HEI,False)]
r=3
for name,n,s,lift,cvr,aov,fill,strong in rows:
    w2.cell(r,1,name); w2.cell(r,2,n); w2.cell(r,3,s).number_format=YEN
    w2.cell(r,4,lift).number_format=XR; w2.cell(r,5,cvr).number_format=PCT; w2.cell(r,6,aov).number_format=YEN
    for c in range(1,7):
        cell=w2.cell(r,c); cell.border=B; cell.fill=fill; cell.alignment=LEFT if c==1 else CEN
        if strong: cell.font=BOLD
    r+=1
w2.cell(r+1,1,'結論：売上・転換率は「イベント × 5と0のつく日」の重なりに集中。マラソン普通日は×1.18で平日並み。').font=SUB
w2.cell(r+2,1,'広告配分：山(5と0＋かぶり)55% ／ SALE・マラソン平常日20% ／ 平日の底25%。平日はゼロにしない。').font=SUB
w2.cell(r+3,1,'※お気に入り率は本データ(売上・アクセス)に無し。RMS R-Karte「お気に入り分析」で追加可能。').font=SUB

# ===== シート3：イベント日マスタ（楽天公式）=====
w3=wb.create_sheet("イベント日マスタ_公式")
w3.merge_cells("A1:E1"); w3["A1"]="楽天 イベント日マスタ（★公式データで管理）"; w3["A1"].font=TITLE
w3.merge_cells("A2:E2"); w3["A2"]="公式日程を入力（黄色セル）。区分に『公式』と入れると確定。現状の推定は必ず公式へ置き換える。"; w3["A2"].font=SUB
h3=[("開始日",13),("終了日",13),("イベント名",22),("区分",12),("備考",30)]
for j,(h,w) in enumerate(h3,1):
    c=w3.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w3.column_dimensions[get_column_letter(j)].width=w
sample=[('2026-08-01','2026-08-01','ワンダフルデー','公式(定例)','毎月1日'),
        ('2026-08-04','2026-08-11','お買い物マラソン','推定→要公式','公式カレンダーで開始/終了を確定'),
        ('2026-08-05','2026-08-05','5と0のつく日','公式(定例)','毎月5/10/15/20/25/30'),
        ('2026-09-04','2026-09-11','スーパーSALE','推定→要公式','9月開催。日程は公式で確定'),]
r=4
for s in sample:
    est='推定' in str(s[3])
    for j,v in enumerate(s,1):
        cell=w3.cell(r,j,v); cell.border=B; cell.alignment=LEFT
        if j==4 and est: cell.fill=PatternFill("solid",fgColor="FFC7CE")
        elif j in(1,2,4): cell.fill=C_IN
    r+=1
for rr in range(r,r+8):
    for j in range(1,6):
        cell=w3.cell(rr,j); cell.border=B
        if j in(1,2,3,4,5): cell.fill=C_IN
w3.cell(r+9,1,'【公式データの取得元】RMS → 店舗運営Navi「イベント・キャンペーンカレンダー」／楽天市場 公式イベント告知ページ。').font=SUB
w3.cell(r+10,1,'ここに公式日程を入れれば、以降の分析・予測はすべて公式ベースに切替（推定は使わない）。').font=SUB

# ===== シート4：凡例・読み方 =====
w4=wb.create_sheet("凡例・読み方"); w4.column_dimensions["A"].width=95
def line(r,t,f=None,fill=None):
    c=w4.cell(r,1,t); c.font=f or font(); c.alignment=Alignment(wrap_text=True,vertical="center")
    if fill: c.fill=fill
L=[('Libetee 楽天 8月作戦＆イベント管理（色付き）',TITLE,None),('',None,None),
 ('【色の意味（セグメント）】',BOLD,None),
 ('  かぶり日（イベント×5と0）＝最強。広告最優先',None,C_KABU),
 ('  5と0のつく日（単独）＝厚めに',None,C_50),
 ('  ワンダフルデー＝上乗せ',None,C_WON),
 ('  お買い物マラソン 普通日＝平日並み。盛らない',None,C_MAR),
 ('  平日＝底を維持（切らない）',None,C_HEI),
 ('  黄色＝入力セル（公式日程などを記入）',None,C_IN),('',None,None),
 ('【シート】',BOLD,None),
 ('  8月_広告配分カレンダー … 8/1〜8/31を色分け＋予想売上＋広告優先度',None,None),
 ('  セグメント別統計 … 平日比・転換率・客単価（13か月実績）',None,None),
 ('  イベント日マスタ_公式 … ★楽天公式の開催日を入力して管理',None,None),('',None,None),
 ('【重要】イベント日は必ず楽天公式データで管理。現状のマラソンは推定表示。公式で置き換え次第、全分析を公式ベースに更新。',BOLD,None),
 ('【未取得】お気に入り率（R-Karte「お気に入り分析」データがあれば追加）。',None,None)]
for i,(t,f,fl) in enumerate(L,1): line(i,t,f,fl)

out="/tmp/claude-0/-home-user--/c7179f16-7eb7-52e8-aa4a-c082e5dbce5f/scratchpad/Libetee_楽天_8月作戦.xlsx"
wb.save(out); print("saved",out)
# 検算表示
import openpyxl as ox
w=ox.load_workbook(out)['8月_広告配分カレンダー']
print("カレンダー例:")
for rr in [4,8,13,18]:
    print(' ',[w.cell(rr,c).value for c in range(1,7)])
