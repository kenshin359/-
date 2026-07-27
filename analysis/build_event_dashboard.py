import json, datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
days=json.load(open('rakuten_all.json'))
for x in days: x['d']=dt.date.fromisoformat(x['date'])
days.sort(key=lambda x:x['d'])
FONT="Arial"; YEN='¥#,##0'; PCT='0.00"%"'; NUM='#,##0'; XR='0.00"x"'; DATEF='yyyy/mm/dd'
BLUE=Font(name=FONT,color="0000FF"); BLACK=Font(name=FONT,color="000000"); GREEN=Font(name=FONT,color="008000")
HEADW=Font(name=FONT,color="FFFFFF",bold=True,size=10); TITLE=Font(name=FONT,bold=True,size=14)
SUB=Font(name=FONT,italic=True,size=9,color="555555"); BOLD=Font(name=FONT,bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
FHI=PatternFill("solid",fgColor="C6EFCE")
thin=Side(style="thin",color="BFBFBF"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
wb=openpyxl.Workbook()

# 日次データ
ws=wb.active; ws.title="日次データ"; MAIN="日次データ"
ws.merge_cells("A1:H1"); ws["A1"]="楽天 日次データ（2025/07〜2026/07・391日・実績）"; ws["A1"].font=TITLE
heads=[("日付",12),("曜日",6),("イベント",20),("売上金額",13),("売上件数",9),("アクセス",11),("転換率",8),("客単価",11)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=BORDER; ws.column_dimensions[get_column_letter(j)].width=w
DS=4
for i,x in enumerate(days):
    r=DS+i
    ws.cell(r,1,x['d']).number_format=DATEF
    ws.cell(r,2,x['dow']); ws.cell(r,3,x['ev'])
    ws.cell(r,4,round(x['sales'])).number_format=YEN
    ws.cell(r,5,int(x['orders'])).number_format=NUM
    ws.cell(r,6,int(x['access'])).number_format=NUM
    ws.cell(r,7,x['cvr']).number_format=PCT
    ws.cell(r,8,round(x['aov'])).number_format=YEN
    for c in range(1,9): ws.cell(r,c).border=BORDER; ws.cell(r,c).font=BLACK
DE=DS+len(days)-1; ws.freeze_panes="A4"
def R(c): return f"'{MAIN}'!${c}${DS}:${c}${DE}"

# イベント別統計（リフト降順）
w2=wb.create_sheet("イベント別統計")
w2.merge_cells("A1:H1"); w2["A1"]="イベント別 売上統計（平常日=1.00・13か月実績）"; w2["A1"].font=TITLE
w2["A2"]="「どのイベントで売上が高くなるか」の答え。倍率・構成比とも実測。"; w2["A2"].font=SUB
h2=[("イベント",20),("日数",7),("日平均売上",14),("リフト倍率",11),("平均客単価",12),("平均転換率",11),("平均アクセス",12),("売上構成比",11)]
for j,(h,w) in enumerate(h2,1):
    c=w2.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=BORDER; w2.column_dimensions[get_column_letter(j)].width=w
evs=['スーパーSALE','5と0のつく日','ワンダフルデー','お買い物マラソン(推定)','平常日']
r0=4
totcell=f"SUM({R('D')})"
for i,ev in enumerate(evs):
    r=r0+i
    w2.cell(r,1,ev).font=GREEN
    w2.cell(r,2,f'=COUNTIF({R("C")},$A{r})').number_format=NUM
    w2.cell(r,3,f'=IFERROR(AVERAGEIFS({R("D")},{R("C")},$A{r}),0)').number_format=YEN
    w2.cell(r,4,f'=IFERROR(C{r}/$C${r0+4},0)').number_format=XR   # 平常日=最終行r0+4
    w2.cell(r,5,f'=IFERROR(AVERAGEIFS({R("H")},{R("C")},$A{r}),0)').number_format=YEN
    w2.cell(r,6,f'=IFERROR(AVERAGEIFS({R("G")},{R("C")},$A{r}),0)').number_format=PCT
    w2.cell(r,7,f'=IFERROR(AVERAGEIFS({R("F")},{R("C")},$A{r}),0)').number_format=NUM
    w2.cell(r,8,f'=IFERROR(SUMIFS({R("D")},{R("C")},$A{r})/{totcell},0)').number_format='0.0%'
    for c in range(1,9):
        w2.cell(r,c).border=BORDER
        if c!=1: w2.cell(r,c).font=BLACK
        if ev!='平常日' and c in (3,4): w2.cell(r,c).fill=FHI
# メモ
w2.cell(r0+6,1,"※ お買い物マラソンは開催日程が毎回変わるため、非SALE月の4〜11日を『推定マラソン期間』として集計。").font=SUB
w2.cell(r0+7,1,"　 正確なマラソン日程をいただければ、その期間で再集計して精度を上げます。").font=SUB
w2.cell(r0+8,1,"※ イベント日は全391日中167日(43%)だが、売上の58.8%を生む。イベント日の最大化が月商の鍵。").font=SUB

# 月商推移
w3=wb.create_sheet("月商推移")
w3.merge_cells("A1:D1"); w3["A1"]="月商推移（13か月）"; w3["A1"].font=TITLE
for j,(h,w) in enumerate([("月",12),("月商",16),("日数",8),("日平均",14)],1):
    c=w3.cell(2,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=BORDER; w3.column_dimensions[get_column_letter(j)].width=w
mons=sorted({x['d'].strftime('%Y-%m') for x in days})
for i,ym in enumerate(mons):
    r=3+i; y,m=map(int,ym.split('-'))
    me=(dt.date(y+(m//12),(m%12)+1,1)-dt.timedelta(days=1))
    crit=f'{R("A")},">="&DATE({y},{m},1),{R("A")},"<="&DATE({me.year},{me.month},{me.day})'
    w3.cell(r,1,ym).font=GREEN
    w3.cell(r,2,f'=SUMIFS({R("D")},{crit})').number_format=YEN
    w3.cell(r,3,f'=COUNTIFS({crit})').number_format=NUM
    w3.cell(r,4,f'=IFERROR(B{r}/C{r},0)').number_format=YEN
    for c in range(1,5): w3.cell(r,c).border=BORDER; 
    if c!=1: w3.cell(r,c).font=BLACK
w3.cell(3+len(mons),1,"※2026/07は26日時点（月途中）").font=SUB

# TOP20日
w4=wb.create_sheet("売上TOP20日")
w4.merge_cells("A1:E1"); w4["A1"]="売上TOP20日"; w4["A1"].font=TITLE
for j,(h,w) in enumerate([("順位",6),("日付",12),("イベント",20),("売上",13),("客単価",11)],1):
    c=w4.cell(2,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=BORDER; w4.column_dimensions[get_column_letter(j)].width=w
for i,x in enumerate(sorted(days,key=lambda k:-k['sales'])[:20]):
    r=3+i
    w4.cell(r,1,i+1); w4.cell(r,2,x['d']).number_format=DATEF
    w4.cell(r,3,x['ev']); w4.cell(r,4,round(x['sales'])).number_format=YEN
    w4.cell(r,5,round(x['aov'])).number_format=YEN
    for c in range(1,6): w4.cell(r,c).border=BORDER; w4.cell(r,c).font=BLACK

# 読み方
w5=wb.create_sheet("読み方"); w5.column_dimensions["A"].width=100
lines=[("楽天イベント別 売上統計（2025/07〜2026/07・13か月）",TITLE),("",None),
 ("【結論：どのイベントで売上が高くなるか】平常日=1.00に対し",BOLD),
 ("  1. スーパーSALE      ×2.40（年4回・3/6/9/12月の4〜11日）",None),
 ("  2. 5と0のつく日      ×2.06（毎月15/20/25/30日 ほか）",None),
 ("  3. ワンダフルデー     ×1.68（毎月1日）",None),
 ("  4. お買い物マラソン(推定)×1.64（非SALE月の4〜11日）",None),("",None),
 ("【売上構成】イベント日は日数では43%だが、売上の58.8%を生む。",BOLD),
 ("  → 平常日の底上げより、イベント日への広告集中の方が効率が高い。",None),("",None),
 ("【客単価の傾向】スーパーSALEは客単価も最高(¥28,943)。高単価商品はSALEに寄せると効く。",None),
 ("【月商推移】2025秋に一時1,400万台へ落ちたが、2026は完全復活。2026/07は26日で4,392万と過去最高ペース。",None),("",None),
 ("※ お買い物マラソンは日程が毎回変動。ここでは非SALE月の4〜11日で推定。正確な日程があれば再集計で精度向上。",SUB),
 ("※ 数値はすべて楽天RMSの実データ。イベント別統計シートは日次データからライブ集計。",SUB)]
for i,(t,f) in enumerate(lines,1):
    c=w5.cell(i,1,t); c.font=f if f else Font(name=FONT); c.alignment=Alignment(wrap_text=True,vertical="center")

out="/tmp/claude-0/-home-user--/c7179f16-7eb7-52e8-aa4a-c082e5dbce5f/scratchpad/Libetee_楽天イベント統計.xlsx"
wb.save(out); print("saved:",out,"| 日次",DS,"-",DE)
