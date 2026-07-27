import json, datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

F="Arial"; YEN='¥#,##0'; NUM='#,##0'; PCT='0.000%'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15)
SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA")
C_HEI=PatternFill("solid",fgColor="F2F2F2"); REDF=PatternFill("solid",fgColor="FFC7CE")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")

# ===== モデル（決定版6本柱：攻めアクセス × 転換率×0.9 × セット率5%） =====
J=json.load(open('july_seg_official.json'))
SC_AOV=29500; FAN_AOV=3500; DROP=0.9; SETF=1+0.05*24500/SC_AOV; CPC_SC=10
SEG={}
for k,(a0,a1) in {'hei':(20000,30000),'mar':(30000,50000),'won':(30000,50000),'f50':(30000,50000),'kabu':(30000,50000)}.items():
    s=J[k]; od=s['avg']/s['aov']; q=max(0,min(1,(s['aov']-FAN_AOV)/(SC_AOV-FAN_AOV)))
    sc_o=od*q; cvr0=sc_o/a0
    b4=round(sc_o*SC_AOV); fan=s['avg']-b4
    SEG[k]=dict(a1=a1,cvr=cvr0*DROP,sc=round(a1*cvr0*DROP*SC_AOV*SETF),fan=fan,
                units=a1*cvr0*DROP*SETF)

CAL={1:'won',5:'kabu',10:'kabu',25:'kabu',15:'f50',20:'f50',30:'f50'}
for d in (4,6,7,8,9,24,26): CAL[d]='mar'
for d in range(1,32): CAL.setdefault(d,'hei')
SEGNAME={'won':('ワンダフルデー',C_WON),'kabu':('★かぶり(マラソン×5と0)',C_KABU),
 'f50':('5と0のつく日',C_50),'mar':('お買い物マラソン',C_MAR),'hei':('平日',C_HEI)}
EVMEMO={4:'マラソン① 20:00開始',24:'マラソン② 20:00開始',26:'マラソン② 最終盤(〜27日9:59)',18:'ご愛顧感謝デー(会員限定・平日扱い)'}
TARGET=65_000_000

wb=openpyxl.Workbook()
ws=wb.active; ws.title="日別シミュレーション"
ws.merge_cells("A1:L1"); ws["A1"]="8月 ¥6,500万達成 日別シミュレーション（決定版6本柱）"; ws["A1"].font=TITLE
ws.merge_cells("A2:L2"); ws["A2"]="前提：SCアクセス 平日3万/イベント5万・転換率×0.9(コンテンツページで回復)・セット率5%・FANは据え置き。目標ペース＝¥6,500万÷31日の直線。貯金＝計画累計−目標ペース。"; ws["A2"].font=SUB
heads=[("日付",9),("セグメント",21),("SCアクセス",11),("SC転換率",10),("SC売上",13),("FAN売上",11),
       ("日合計",13),("累計",14),("目標ペース累計",14),("貯金(＋)/借金(−)",14),("SC広告費",11),("メモ",24)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B
    ws.column_dimensions[get_column_letter(j)].width=w
r=5; cum=0; tot=dict(sc=0,fan=0,ad=0,acc=0,units=0.0)
YB='月火水木金土日'
for d in range(1,32):
    k=CAL[d]; s=SEG[k]; name,fill=SEGNAME[k]
    date=dt.date(2026,8,d); wd=YB[date.weekday()]
    day=s['sc']+s['fan']; cum+=day
    pace=round(TARGET*d/31); ad=s['a1']*CPC_SC
    ws.cell(r,1,f"8/{d}({wd})")
    ws.cell(r,2,name)
    ws.cell(r,3,s['a1']).number_format=NUM
    ws.cell(r,4,s['cvr']).number_format=PCT
    ws.cell(r,5,s['sc']).number_format=YEN
    ws.cell(r,6,s['fan']).number_format=YEN
    ws.cell(r,7,day).number_format=YEN
    ws.cell(r,8,cum).number_format=YEN
    ws.cell(r,9,pace).number_format=YEN
    ws.cell(r,10,cum-pace).number_format='+¥#,##0;-¥#,##0'
    ws.cell(r,11,ad).number_format=YEN
    ws.cell(r,12,EVMEMO.get(d,''))
    for c in range(1,13):
        cell=ws.cell(r,c); cell.border=B; cell.fill=fill
        cell.alignment=LEFT if c in(2,12) else CEN
        if k=='kabu': cell.font=BOLD
    if cum-pace<0: ws.cell(r,10).fill=REDF
    tot['sc']+=s['sc']; tot['fan']+=s['fan']; tot['ad']+=ad; tot['acc']+=s['a1']; tot['units']+=s['units']
    r+=1
ws.cell(r,1,'月間').font=BOLD
ws.cell(r,3,tot['acc']).number_format=NUM
ws.cell(r,5,tot['sc']).number_format=YEN
ws.cell(r,6,tot['fan']).number_format=YEN
ws.cell(r,7,tot['sc']+tot['fan']).number_format=YEN
ws.cell(r,8,cum).number_format=YEN
ws.cell(r,9,TARGET).number_format=YEN
ws.cell(r,10,cum-TARGET).number_format='+¥#,##0;-¥#,##0'
ws.cell(r,11,tot['ad']).number_format=YEN
for c in range(1,13): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
r+=2
notes=[f"月間見込み ¥{cum:,}（目標¥65,000,000 に対し +¥{cum-TARGET:,} の貯金）。SC販売個数 約{round(tot['units']):,}個。",
 f"SC広告費 ¥{tot['ad']:,}（アクセス{tot['acc']:,}×¥10）。SC売上¥{tot['sc']:,} → SCのROAS {tot['sc']/tot['ad']:.1f}倍。FAN広告は据え置き(¥8/アクセス)のため本表に含めず。",
 "山は3つ：8/5(¥583万)・8/10(¥583万)・8/25(¥583万)のかぶり日。この3日で月の26%を売る。在庫を最厚に。",
 "8/2〜3と8/11〜14は借金(−)側に沈むのが正常。かぶり日で一気に回収する山谷型の計画。",
 "毎週の答え合わせ：実績の累計をH列の隣に書き、貯金列がマイナス継続なら広告・在庫・転換率のどれが欠けたかを検証チームで特定。"]
for t in notes:
    ws.cell(r,1,t).font=SUB; r+=1
ws.freeze_panes="A5"

# 累計グラフ（計画 vs 目標ペース）
ch=LineChart(); ch.title="累計売上：計画 vs ¥6,500万ペース"; ch.height=9; ch.width=22
data=Reference(ws,min_col=8,max_col=9,min_row=4,max_row=35)
cats=Reference(ws,min_col=1,min_row=5,max_row=35)
ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
ch.y_axis.numFmt='¥#,##0'; ch.y_axis.title='累計売上'; ch.x_axis.title='日付'
ws.add_chart(ch,"N4")

# ===== S2 前提条件・用語 =====
w2=wb.create_sheet("前提条件・用語")
w2.column_dimensions['A'].width=100
L=[('この日別シミュレーションの前提（決定版6本柱）',TITLE),('',None),
 ('1) アクセス：スーツケース(SC)を平日30,000／イベント日50,000にメタ広告で調整。ハンディファン(FAN)は据え置き。',None),
 ('2) 転換率：アクセス増で薄まる分を×0.8と見ていたが、SCコンテンツページ新設で×0.9まで回復させる（6本柱⑤）。',None),
 ('3) セット率5%：セット購入¥5,000引き（6本柱③）で、SC注文の5%が2個目(¥24,500)を追加購入する想定。',None),
 ('4) 単価：7月実績ベース。SC客単価¥29,500・FAN¥3,500。アクセス単価はSC¥10/アクセス・FAN¥8/アクセスで統一。',None),
 ('5) イベント日：楽天公式カレンダー（8月）。マラソン①8/4-11・②8/24-27、かぶり日=8/5・8/10・8/25、5と0単独=8/15・20・30、8/1ワンダフルデー。',None),
 ('6) 各セグメントの1日の売上・転換率は、7月の公式カレンダー実績から逆算（かぶり/5と0/ワンダフル/マラソン/平日の5区分）。',None),
 ('',None),('【用語のかんたん解説】',BOLD),
 ('・SC転換率＝スーツケースを見に来た100人のうち買う人の割合。0.123%なら1万アクセスで約12人が購入。',None),
 ('・目標ペース累計＝¥6,500万を31日で均等に割った「今日ここまで売れていれば合格」の線。',None),
 ('・貯金(＋)/借金(−)＝計画の累計が目標ペースより進んでいるか遅れているか。イベント前は借金、イベント後に貯金になるのが正常。',None),
 ('・かぶり日＝お買い物マラソンと「5と0のつく日」が重なる日。転換率が跳ね、1日で平日の4〜5倍売れる最重要日。',None),
 ('・セット率＝2個同時購入クーポン(¥5,000引き)を使ってくれる注文の割合。1組増えるごとに売上+¥24,500。',None),
 ('',None),('【使い方】8月が始まったら、毎日の実績をこの表の横に記入 → 貯金列のプラス/マイナスで「今月は勝てているか」が一目で分かる。',BOLD),
 ('週1回、検証チームが「予測との差」を分解（アクセス不足？転換率低下？在庫切れ？）して次週の広告調整に反映する。',None)]
for i,(t,f) in enumerate(L,1):
    c=w2.cell(i,1,t); c.font=f or font(); c.alignment=Alignment(wrap_text=True,vertical="center")

out="Libetee_8月6500万_日別シミュレーション.xlsx"
wb.save(out)
print('saved',out)
print('月間:',cum,'SC:',tot['sc'],'FAN:',tot['fan'],'広告費:',tot['ad'],'個数:',round(tot['units']))
