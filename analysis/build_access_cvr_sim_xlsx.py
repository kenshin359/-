import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; NUM='#,##0'; YEN='¥#,##0'; P3='0.000"%"'; P2='0.00"%"'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
H_REF=PatternFill("solid",fgColor="808B96")   # 7月実績ヘッダ=グレー
H_B4=PatternFill("solid",fgColor="2E75B6")    # 増額なし=青
H_AF=PatternFill("solid",fgColor="C55A11")    # 増額後=オレンジ
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2")
C_FAN=PatternFill("solid",fgColor="DDEBD8"); C_DIFF=PatternFill("solid",fgColor="FFE2DD")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
J=json.load(open('july_seg_official.json'))
SC_AOV=29500; FAN_AOV=3500; DROP=0.8
# 7月実績アクセス/日（公式分類・店舗全体）
JACC={'hei':28379,'mar':29286,'won':12728,'f50':26858,'kabu':41739}
plan=[('hei','平日',17,20000,30000,C_HEI),
      ('mar','イベント(5,0抜き)：マラソン通常',7,30000,50000,C_MAR),
      ('won','イベント(5,0抜き)：ワンダフルデー',1,30000,50000,C_WON),
      ('f50','5と0のつく日(単独)',3,30000,50000,C_50),
      ('kabu','イベント＋5,0(かぶり)',3,30000,50000,C_KABU)]
rows=[]
for k,name,n,a0,a1,fill in plan:
    s=J[k]; od=s['avg']/s['aov']; q=max(0,min(1,(s['aov']-FAN_AOV)/(SC_AOV-FAN_AOV)))
    sc_o=od*q; cvr0=sc_o/a0; cvr1=cvr0*DROP
    b4=round(sc_o*SC_AOV); af=round(a1*cvr1*SC_AOV); fan=s['avg']-b4
    rows.append(dict(k=k,name=name,n=n,a0=a0,a1=a1,fill=fill,cvr0=cvr0,cvr1=cvr1,b4=b4,af=af,fan=fan,
                     javg=s['avg'],jacc=JACC[k],jcvr=s['cvr']))
SCB=sum(r['b4']*r['n'] for r in rows); SCA=sum(r['af']*r['n'] for r in rows)
FANM=sum(r['fan']*r['n'] for r in rows)
wb=openpyxl.Workbook()

# メイン
ws=wb.active; ws.title="シミュレーション"
ws.merge_cells("A1:N1"); ws["A1"]="アクセス数 × 転換率 シミュレーション（スーツケース・転換率0.8掛け）"; ws["A1"].font=TITLE
ws.merge_cells("A2:N2"); ws["A2"]="増額なし＝平日2万/イベント3万・転換率据え置き(7月実績)。増額後＝平日3万/イベント5万・転換率×0.8。ハンディファンは据え置き。8月公式カレンダー。"; ws["A2"].font=SUB
# 2段ヘッダ
ws.merge_cells("C4:E4"); ws["C4"]="7月実績（参考・店舗全体）"; 
ws.merge_cells("F4:I4"); ws["F4"]="増額なし（スーツケース）"
ws.merge_cells("J4:N4"); ws["J4"]="増額後（スーツケース・転換率×0.8）"
for rng,fill in [("C4:E4",H_REF),("F4:I4",H_B4),("J4:N4",H_AF)]:
    for row in ws[rng]:
        for c in row: c.fill=fill; c.font=HEADW; c.alignment=CEN; c.border=B
heads=[("区分",28),("日数",6),
 ("売上/日",12),("アクセス/日",11),("転換率",9),
 ("アクセス",10),("転換率",10),("売上/日",12),("月間",13),
 ("アクセス",10),("転換率\n(×0.8)",10),("売上/日",12),("月間",13),("月間差額",13)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(5,j,h); c.font=HEADW; c.alignment=CEN; c.border=B
    c.fill = FH if j<=2 else (H_REF if j<=5 else (H_B4 if j<=9 else H_AF))
    ws.column_dimensions[get_column_letter(j)].width=w
r=6
for x in rows:
    ws.cell(r,1,x['name']); ws.cell(r,2,x['n'])
    ws.cell(r,3,x['javg']).number_format=YEN
    ws.cell(r,4,x['jacc']).number_format=NUM
    ws.cell(r,5,x['jcvr']*100).number_format=P2
    ws.cell(r,6,x['a0']).number_format=NUM
    ws.cell(r,7,x['cvr0']*100).number_format=P3
    ws.cell(r,8,x['b4']).number_format=YEN
    ws.cell(r,9,x['b4']*x['n']).number_format=YEN
    ws.cell(r,10,x['a1']).number_format=NUM
    ws.cell(r,11,x['cvr1']*100).number_format=P3
    ws.cell(r,12,x['af']).number_format=YEN
    ws.cell(r,13,x['af']*x['n']).number_format=YEN
    ws.cell(r,14,(x['af']-x['b4'])*x['n']).number_format=YEN
    for c in range(1,15):
        cell=ws.cell(r,c); cell.border=B; cell.fill=x['fill']; cell.alignment=LEFT if c==1 else CEN
        if c==14: cell.font=BOLD
    r+=1
ws.cell(r,1,'スーツケース 計').font=BOLD
ws.cell(r,9,SCB).number_format=YEN; ws.cell(r,13,SCA).number_format=YEN; ws.cell(r,14,SCA-SCB).number_format=YEN
for c in range(1,15): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
r+=1
ws.cell(r,1,'ハンディファン（アクセス据え置き）')
ws.cell(r,9,FANM).number_format=YEN; ws.cell(r,13,FANM).number_format=YEN; ws.cell(r,14,0).number_format=YEN
for c in range(1,15): ws.cell(r,c).fill=C_FAN; ws.cell(r,c).border=B
r+=1
ws.cell(r,1,'合算').font=font(bold=True,size=12)
ws.cell(r,9,SCB+FANM).number_format=YEN; ws.cell(r,13,SCA+FANM).number_format=YEN; ws.cell(r,14,SCA-SCB).number_format=YEN
for c in range(1,15): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=font(bold=True,size=11)
r+=2
ws.cell(r,1,'差額（増額後−増額なし）').font=font(bold=True,size=13,color='C0392B')
c=ws.cell(r,3,SCA-SCB); c.number_format=YEN; c.font=font(bold=True,size=13,color='C0392B'); c.fill=C_DIFF
r+=2
for t in ['転換率×0.8でも増額はプラス：アクセス+67%(イベント)/+50%(平日) × 転換率−20% ＝ 売上は約+28%。',
 '参考：転換率×0.9なら差額+¥16.7M(合算¥64.0M)。×0.8は保守ライン。実際はこの間に着地する公算。',
 '7月実績列は店舗全体(SC+FAN)の実測。増額なし/後の列はスーツケースのみ（転換率はSC注文÷指定アクセスで逆算）。',
 '個数の裏付け：SKU実績シート（多機能PC96%・マットブラックS最重要）参照。在庫が前提条件。']:
    ws.cell(r,1,t).font=SUB; r+=1
ws.freeze_panes="A6"

# グラフ
w2=wb.create_sheet("グラフ")
w2["A1"]="グラフ"; w2["A1"].font=TITLE
labels=['平日','マラソン通常','ワンダフル','5と0単独','かぶり日']
w2["H1"]="区分"; w2["I1"]="増額なし"; w2["J1"]="増額後(×0.8)"
for i,x in enumerate(rows):
    w2.cell(2+i,8,labels[i]); w2.cell(2+i,9,x['b4']); w2.cell(2+i,10,x['af'])
w2["H8"]="区分"; w2["I8"]="増額なし"; w2["J8"]="増額後"
w2["H9"]="スーツケース"; w2["I9"]=SCB; w2["J9"]=SCA
w2["H10"]="ハンディファン"; w2["I10"]=FANM; w2["J10"]=FANM
w2["H11"]="合算"; w2["I11"]=SCB+FANM; w2["J11"]=SCA+FANM
ch=BarChart(); ch.type="col"; ch.title="スーツケース 売上/日：増額なし vs 増額後（転換率×0.8）"
ch.height=10; ch.width=24; ch.gapWidth=60; ch.y_axis.title='円/日'
ch.add_data(Reference(w2,min_col=9,max_col=10,min_row=1,max_row=6),titles_from_data=True)
ch.set_categories(Reference(w2,min_col=8,min_row=2,max_row=6))
w2.add_chart(ch,"A3")
ch2=BarChart(); ch2.type="col"; ch2.title="月間：SC/FAN/合算"
ch2.height=10; ch2.width=16; ch2.gapWidth=60
ch2.add_data(Reference(w2,min_col=9,max_col=10,min_row=8,max_row=11),titles_from_data=True)
ch2.set_categories(Reference(w2,min_col=8,min_row=9,max_row=11))
w2.add_chart(ch2,"N3")
out="Libetee_アクセス転換率シミュレーション.xlsx"
wb.save(out); print('saved',out)
print(f'SC {SCB:,} → {SCA:,} (+{SCA-SCB:,}) / 合算 {SCB+FANM:,} → {SCA+FANM:,}')
