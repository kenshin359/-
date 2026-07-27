import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; YEN='¥#,##0'; MYEN='¥#,##0,,"M"'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
G_OK=PatternFill("solid",fgColor="C6EFCE"); Y_MID=PatternFill("solid",fgColor="FFF2CC"); GRY=PatternFill("solid",fgColor="F2F2F2")
C_NAVY=PatternFill("solid",fgColor="2E75B6"); C_GREEN=PatternFill("solid",fgColor="27AE60"); C_RED=PatternFill("solid",fgColor="C0392B")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center"); WRAP=Alignment(wrap_text=True,vertical="center")
FAN=9276649; SC10=48634772/0.8; AOV=29500; SETADD=24500
def total(r,s): return round(SC10*r*(1+s*SETADD/AOV)+FAN)
wb=openpyxl.Workbook()

# S1 サマリー
ws=wb.active; ws.title="サマリー"
ws.merge_cells("A1:F1"); ws["A1"]="月商¥6,500万への作戦（楽天・8月）"; ws["A1"].font=TITLE
ws.merge_cells("A2:F2"); ws["A2"]="土台＝攻めプラン（平日3万/イベント5万アクセス・SC単価¥10/FAN¥8・8月公式カレンダー・7月実績ベース）"; ws["A2"].font=SUB
for j,(h,w) in enumerate([("シナリオ",30),("月商",16),("目標65Mとの差",16),("判定",12)],1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
scen=[('転換率×0.8（保守見積のまま）',total(0.8,0)),('転換率×0.9',total(0.9,0)),
 ('転換率 据え置き(×1.0)',total(1.0,0)),('【推奨】保持88%＋セット率5%',total(0.88,0.05)),('保持88%＋セット率8%',total(0.88,0.08))]
r=5
for name,v in scen:
    ws.cell(r,1,name); ws.cell(r,2,v).number_format=YEN
    ws.cell(r,3,v-65000000).number_format='+¥#,##0;-¥#,##0'
    ok=v>=64900000
    c=ws.cell(r,4,'達成' if ok else '未達'); c.fill=G_OK if ok else GRY; c.alignment=CEN
    for cc in range(1,5): ws.cell(r,cc).border=B
    if '推奨' in name:
        for cc in range(1,5): ws.cell(r,cc).font=BOLD
    r+=1
r+=1
for t in ['読み方：転換率の下落を−8.3%以内に抑えれば(保持91.7%)、アクセス増だけで65M。',
 '推奨は併用：転換率防衛7策で「下落−12%以内」＋ セット販売でお客様の5%が2個購入 → ¥65.0M。',
 'セット1組の追加売上＝2個目¥29,500−クーポン¥5,000＝¥24,500。必要セット数：5%＝約91組/月。',
 '前提：スーツケース在庫（マットブラックS筆頭・かぶり日8/5・10・25の3日分）。']:
    ws.cell(r,1,t).font=SUB; r+=1

# S2 早見表（保持率×セット率）
w2=wb.create_sheet("達成早見表")
w2.merge_cells("A1:H1"); w2["A1"]="65M達成早見表：転換率保持率 × セット率 → 月商"; w2["A1"].font=TITLE
w2.merge_cells("A2:H2"); w2["A2"]="緑＝65M達成／黄＝60〜65M／グレー＝60M未満。行=転換率をどこまで守れたか、列=2個セット購入率。"; w2["A2"].font=SUB
rates=[1.00,0.95,0.92,0.90,0.88,0.85,0.80]
sets=[0,0.03,0.05,0.08,0.10]
w2.cell(4,1,'保持率＼セット率').font=HEADW; w2.cell(4,1).fill=FH; w2.cell(4,1).border=B; w2.cell(4,1).alignment=CEN
w2.column_dimensions['A'].width=16
for j,s in enumerate(sets,2):
    c=w2.cell(4,j,f'{s:.0%}'); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B
    w2.column_dimensions[get_column_letter(j)].width=14
for i,rr in enumerate(rates):
    r=5+i
    c=w2.cell(r,1,f'{rr:.0%}' + ('（据え置き）' if rr==1 else '')); c.font=BOLD; c.border=B
    for j,s in enumerate(sets,2):
        v=total(rr,s)
        cell=w2.cell(r,j,v); cell.number_format=YEN; cell.border=B; cell.alignment=CEN
        cell.fill = G_OK if v>=64900000 else (Y_MID if v>=60000000 else GRY)
w2.cell(5+len(rates)+1,1,'例：保持88%×セット5% ＝ ¥64,996,311 ≒ 65M（推奨ライン）。保持92%ならセット無しでも到達。').font=SUB

# S3 転換率を守る7策
w3=wb.create_sheet("転換率を守る7策")
w3.merge_cells("A1:D1"); w3["A1"]="転換率を「据え置き」に守る7策"; w3["A1"].font=TITLE
for j,(h,w) in enumerate([("#",5),("施策",26),("中身",62),("優先度",10)],1):
    c=w3.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w3.column_dimensions[get_column_letter(j)].width=w
P=[('楽天アプリへ流す導線','7月実測：アプリ転換率4〜8% vs スマホWeb0.03〜0.13%（約50倍）。メタ→アプリで開かせる（ディープリンク/アプリ限定クーポン訴求）。増えた流入の質が別物になる','S 最優先'),
 ('テレビ出演・放映実績の活用','ヒルナンデス実績を広告1枚目・LPファーストビューへ。新規TV露出が決まれば放映日に広告増額をぶつける','S'),
 ('LPに動きをつける（動画LP化）','ファーストビュー6秒動画：フロントオープン→ストッパー→静音キャスター実演。GIF/ショート動画は冷たい流入に強い','A'),
 ('楽天広告(RPP/CA)の精度','検索語の絞込・除外、売れ筋SKU(多機能PC)へ入札集中、イベント日だけ増額の山谷運用','A'),
 ('メタ広告の精度','購入者リストの類似オーディエンス、動画視聴→カート→購入のリタゲ階層、レビュー/UGC型クリエイティブ','A'),
 ('商品ページCRO','レビュー動画・サイズ比較表・Q&A先回り・クーポン視認性UP','B'),
 ('カゴ落ち回収','カート放棄への再訪クーポン(Rクーポン)。最も安い転換率改善','B')]
r=4
for i,(t,d,pr) in enumerate(P,1):
    w3.cell(r,1,i); w3.cell(r,2,t).font=BOLD; w3.cell(r,3,d).alignment=WRAP; w3.cell(r,4,pr).alignment=CEN
    for c in range(1,5): w3.cell(r,c).border=B
    if pr.startswith('S'):
        for c in range(1,5): w3.cell(r,c).fill=Y_MID
    w3.row_dimensions[r].height=32
    r+=1

# S4 客単価を上げる7策
w4=wb.create_sheet("客単価を上げる7策")
w4.merge_cells("A1:D1"); w4["A1"]="客単価を上げる7策"; w4["A1"].font=TITLE
for j,(h,w) in enumerate([("#",5),("施策",26),("中身",62),("優先度",10)],1):
    c=w4.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w4.column_dimensions[get_column_letter(j)].width=w
Q=[('2個セット¥5,000クーポン【社長案】','2個目の獲得コスト¥5,000＜広告の新規獲得(平日約¥6,600)。対象：多機能PC/ノーマルアルミ/多機能アルミ。目標セット率5%＝約91組/月','S 最優先'),
 ('S→Mアップセル','購入者の57%がSサイズ。「+¥6,900でMサイズ(3〜7泊)」比較表を目立たせる。値引きゼロで客単価UP','S'),
 ('圧縮バッグ同梱の拡販','SCセット1000円OFFは既に月17組の実績。全SCページに同梱枠を設置','A'),
 ('段階クーポン（3個¥9,000）','家族・帰省需要。8月は「家族分まとめて」が刺さる季節','A'),
 ('ペア割の見せ方','同じ2個割でも「夫婦で色違い」訴求（マットブラック×シルバー提案）','B'),
 ('夏旅セット（SC＋ハンディファン）','旅行文脈のクロスセル。ファン在庫消化にも効く','B'),
 ('お盆・帰省ギフト訴求','のし・ラッピング対応を前面に。8月特有の贈答・買い替え需要','B')]
r=4
for i,(t,d,pr) in enumerate(Q,1):
    w4.cell(r,1,i); w4.cell(r,2,t).font=BOLD; w4.cell(r,3,d).alignment=WRAP; w4.cell(r,4,pr).alignment=CEN
    for c in range(1,5): w4.cell(r,c).border=B
    if pr.startswith('S'):
        for c in range(1,5): w4.cell(r,c).fill=Y_MID
    w4.row_dimensions[r].height=32
    r+=1

# S5 グラフ
w5=wb.create_sheet("グラフ")
w5["A1"]="グラフ"; w5["A1"].font=TITLE
w5["H1"]="シナリオ"; w5["I1"]="月商"; w5["H7"]="目標"; 
data=[('×0.8',total(0.8,0)),('×0.9',total(0.9,0)),('据え置き',total(1.0,0)),('推奨:88%+セット5%',total(0.88,0.05))]
for i,(n,v) in enumerate(data):
    w5.cell(2+i,8,n); w5.cell(2+i,9,v)
w5.cell(6,8,'目標65M'); w5.cell(6,9,65000000)
ch=BarChart(); ch.type="col"; ch.title="シナリオ別 月商 vs 目標65M"
ch.height=10; ch.width=20; ch.legend=None
ch.add_data(Reference(w5,min_col=9,min_row=1,max_row=6),titles_from_data=True)
ch.set_categories(Reference(w5,min_col=8,min_row=2,max_row=6))
w5.add_chart(ch,"A3")
out="Libetee_6500万作戦.xlsx"
wb.save(out); print('saved',out)
print('検証: 推奨(88%,5%)=',f'{total(0.88,0.05):,}','/ 据え置き=',f'{total(1.0,0):,}','/ ×0.8=',f'{total(0.8,0):,}')
