import datetime as dt, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; YEN='¥#,##0'; NUM='#,##0'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_SC=PatternFill("solid",fgColor="D6E4F0")   # スーツケース=青系
C_FAN=PatternFill("solid",fgColor="DDEBD8")  # ファン=緑系
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2")
C_DIFF=PatternFill("solid",fgColor="FFE2DD")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
J=json.load(open('july_seg.json'))
SC_AOV=29500; FAN_AOV=3500
plan=[('kabu','かぶり日（8/5・10・25）',3,10000,C_KABU),
      ('f50','5と0単独（8/15・20・30）',3,10000,C_50),
      ('won','ワンダフルデー（8/1）',1,10000,C_WON),
      ('mar','マラソン通常（7日）',7,10000,C_MAR),
      ('hei','平日（17日）',17,5000,C_HEI)]
# 分解：SC比率、日次ベース/追加をSC・FANへ（FAN=全体-SCで合算を厳密一致させる）
rows=[]
for k,name,n,da,fill in plan:
    s=J[k]; aov=s['aov']
    q=max(0,min(1,(aov-FAN_AOV)/(SC_AOV-FAN_AOV))); share=q*SC_AOV/aov
    add=round(da*s['vpa'])
    sc_b=round(s['avg']*share); fan_b=s['avg']-sc_b
    sc_a=round(add*share); fan_a=add-sc_a
    rows.append(dict(k=k,name=name,n=n,da=da,fill=fill,share=share,avg=s['avg'],add=add,
                     sc_b=sc_b,fan_b=fan_b,sc_a=sc_a,fan_a=fan_a))
TB=sum(r['avg']*r['n'] for r in rows); TA=sum(r['add']*r['n'] for r in rows)
SCB=sum(r['sc_b']*r['n'] for r in rows); SCA=sum(r['sc_a']*r['n'] for r in rows)
FANB=TB-SCB; FANA=TA-SCA
wb=openpyxl.Workbook()

def sheet_product(title,pref,bkey,akey,fillc):
    w=wb.create_sheet(title)
    w.merge_cells("A1:G1"); w["A1"]=f"{pref}（7月ベース・転換率キープ）"; w["A1"].font=TITLE
    heads=[("セグメント",30),("日数",7),("売上比率",10),("ペース 売上/日",14),("増額の追加/日",14),("月間ベース",15),("月間 増額後",15)]
    for j,(h,wd) in enumerate(heads,1):
        c=w.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w.column_dimensions[get_column_letter(j)].width=wd
    r=4; t1=t2=0
    for x in rows:
        b=x[bkey]; a=x[akey]
        w.cell(r,1,x['name']); w.cell(r,2,x['n'])
        share = x['share'] if bkey=='sc_b' else 1-x['share']
        w.cell(r,3,share).number_format='0.0%'
        w.cell(r,4,b).number_format=YEN; w.cell(r,5,a).number_format=YEN
        w.cell(r,6,b*x['n']).number_format=YEN; w.cell(r,7,(b+a)*x['n']).number_format=YEN
        for c in range(1,8):
            cell=w.cell(r,c); cell.border=B; cell.fill=x['fill']; cell.alignment=LEFT if c==1 else CEN
        t1+=b*x['n']; t2+=(b+a)*x['n']; r+=1
    w.cell(r,1,'合計').font=BOLD
    w.cell(r,6,t1).number_format=YEN; w.cell(r,7,t2).number_format=YEN
    for c in range(1,8): w.cell(r,c).fill=FT; w.cell(r,c).border=B; w.cell(r,c).font=BOLD
    w.cell(r+2,1,f'月間ベース ¥{t1:,} → 増額後 ¥{t2:,}（差額 +¥{t2-t1:,}）').font=BOLD
    return t1,t2

# ===== S1 合算まとめ =====
ws=wb.active; ws.title="合算まとめ"
ws.merge_cells("A1:H1"); ws["A1"]="スーツケース × ハンディファン 分解 → 合算（7月ベース・転換率キープ・8月公式カレンダー）"; ws["A1"].font=TITLE
ws.merge_cells("A2:H2"); ws["A2"]="分解方法：冬(2026/1-2月=扇風機オフ季)の客単価¥29,500をスーツケース、ファンは価格帯¥3,500と置き、各セグメントの客単価から構成を逆算。"; ws["A2"].font=SUB
heads=[("区分",22),("ベース(通常予算)",17),("増額の追加",15),("増額後",17),("構成比(増額後)",13)]
for j,(h,wd) in enumerate(heads,1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=wd
data=[('スーツケース',SCB,SCA,C_SC),('ハンディファン',FANB,FANA,C_FAN)]
r=5
for name,b,a,fl in data:
    ws.cell(r,1,name).font=BOLD
    ws.cell(r,2,b).number_format=YEN; ws.cell(r,3,a).number_format=YEN
    ws.cell(r,4,b+a).number_format=YEN
    ws.cell(r,5,(b+a)/(TB+TA)).number_format='0.0%'
    for c in range(1,6): ws.cell(r,c).border=B; ws.cell(r,c).fill=fl; ws.cell(r,c).alignment=LEFT if c==1 else CEN
    r+=1
ws.cell(r,1,'合算').font=font(bold=True,size=12)
ws.cell(r,2,TB).number_format=YEN; ws.cell(r,3,TA).number_format=YEN
ws.cell(r,4,TB+TA).number_format=YEN; ws.cell(r,5,1).number_format='0.0%'
for c in range(1,6): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
r+=2
ws.cell(r,1,'差額（増額-通常）').font=font(bold=True,size=12,color='C0392B')
c=ws.cell(r,2,TA); c.number_format=YEN; c.font=font(bold=True,size=12,color='C0392B'); c.fill=C_DIFF
r+=2
for t in [f'スーツケース：ベース¥{SCB:,} → 増額後¥{SCB+SCA:,}（+¥{SCA:,}）＝売上の約{(SCB+SCA)/(TB+TA)*100:.0f}%を稼ぐ主力。',
 f'ハンディファン：ベース¥{FANB:,} → 増額後¥{FANB+FANA:,}（+¥{FANA:,}）。5と0・かぶり日に集中して売れる（数量型）。',
 '在庫の目安：スーツケースは高単価のため在庫1個の欠品損失が大。かぶり日3日はSC在庫を最優先で厚く。',
 '※商品別の正確な分解はRMS「商品別売上CSV」で確定可能。本表は客単価からの逆算モデル（前提と検算シート参照）。']:
    ws.cell(r,1,t).font=SUB; r+=1

t_sc=sheet_product("スーツケース","スーツケース 8月内訳",'sc_b','sc_a',C_SC)
t_fan=sheet_product("ハンディファン","ハンディファン 8月内訳",'fan_b','fan_a',C_FAN)

# ===== グラフ =====
w2=wb.create_sheet("グラフ")
w2["A1"]="グラフ"; w2["A1"].font=TITLE
w2["H1"]="区分"; w2["I1"]="ベース"; w2["J1"]="増額後"
w2["H2"]="スーツケース"; w2["I2"]=SCB; w2["J2"]=SCB+SCA
w2["H3"]="ハンディファン"; w2["I3"]=FANB; w2["J3"]=FANB+FANA
w2["H4"]="合算"; w2["I4"]=TB; w2["J4"]=TB+TA
ch=BarChart(); ch.type="col"; ch.title="8月 月商：ベース vs 増額後（スーツケース／ハンディファン／合算）"
ch.height=10; ch.width=24; ch.gapWidth=60; ch.y_axis.title='円'
ch.add_data(Reference(w2,min_col=9,max_col=10,min_row=1,max_row=4),titles_from_data=True)
ch.set_categories(Reference(w2,min_col=8,min_row=2,max_row=4))
w2.add_chart(ch,"A3")
# セグメント別SC/FAN積上げ（増額後/日）
w2["H8"]="セグメント"; w2["I8"]="スーツケース/日"; w2["J8"]="ハンディファン/日"
short=['かぶり日','5と0単独','ワンダフル','マラソン通常','平日']
for i,x in enumerate(rows):
    w2.cell(9+i,8,short[i]); w2.cell(9+i,9,x['sc_b']+x['sc_a']); w2.cell(9+i,10,x['fan_b']+x['fan_a'])
ch2=BarChart(); ch2.type="col"; ch2.grouping="stacked"; ch2.overlap=100
ch2.title="増額後の1日あたり売上（SC/FAN積み上げ・セグメント別)"
ch2.height=10; ch2.width=24
ch2.add_data(Reference(w2,min_col=9,max_col=10,min_row=8,max_row=13),titles_from_data=True)
ch2.set_categories(Reference(w2,min_col=8,min_row=9,max_row=13))
w2.add_chart(ch2,"A24")

# ===== 前提と検算 =====
w3=wb.create_sheet("前提と検算"); w3.column_dimensions['A'].width=95
L=[('分解モデルの前提',BOLD),
 ('  ・スーツケース客単価 = ¥29,500（2026年1-2月実績。扇風機が売れない季節＝スーツケースの素の客単価）',None),
 ('  ・ハンディファン客単価 = ¥3,500（首振り/スケルトン等の価格帯。仮定値）',None),
 ('  ・各セグメントの7月客単価から SC/FAN の注文構成を逆算 → 売上を分解',None),
 ('  ・増額分（平日+5千/イベント+1万）も同じ構成比で分解（メタの商品配分で意図的に傾けることは可能）',None),
 ('',None),('検算（合算＝7月ベース版と完全一致）',BOLD),
 (f'  ・ベース合算 ¥{TB:,} ＝ 7月ベース版の通常予算と一致',None),
 (f'  ・増額後合算 ¥{TB+TA:,} ＝ 7月ベース版の増額後と一致',None),
 (f'  ・SC+FAN = 合算（各セグメント・各行で厳密に一致するよう分解）',None),
 ('',None),('精度を上げるには',BOLD),
 ('  ・RMS「商品別売上CSV」（商品管理番号別の売上・件数）があれば、モデルではなく実数で分解できます。',None)]
for i,(t,f) in enumerate(L,1):
    w3.cell(i,1,t).font=f or font()

out="Libetee_増額シミュレーション_SC_FAN分解.xlsx"
wb.save(out)
print('saved',out)
print(f'SC: base {SCB:,} add {SCA:,} total {SCB+SCA:,}')
print(f'FAN: base {FANB:,} add {FANA:,} total {FANB+FANA:,}')
print(f'合算: base {TB:,} add {TA:,} total {TB+TA:,}')
print('検算(7月ベース版と一致):', 'OK' if (TB==53006535 and TB+TA==66771315) else 'NG!')
print('検算(SC+FAN=合算):', 'OK' if (SCB+FANB==TB and SCA+FANA==TA) else 'NG!')
print('検算(商品シート合計):', 'OK' if (t_sc[0]==SCB and t_fan[0]==FANB) else 'NG!')
