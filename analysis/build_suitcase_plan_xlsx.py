import datetime as dt, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; YEN='¥#,##0'; NUM='#,##0'; PCT3='0.000"%"'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2")
C_FAN=PatternFill("solid",fgColor="DDEBD8"); C_DIFF=PatternFill("solid",fgColor="FFE2DD"); C_IN=PatternFill("solid",fgColor="FFF2CC")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
J=json.load(open('july_seg_official.json'))
SC_AOV=29500; FAN_AOV=3500; DROP=0.9
plan=[('hei','平日',17,20000,30000,C_HEI),
      ('mar','イベント(5,0抜き)：マラソン通常',7,30000,50000,C_MAR),
      ('won','イベント(5,0抜き)：ワンダフルデー',1,30000,50000,C_WON),
      ('f50','5と0のつく日(単独 8/15・20・30)',3,30000,50000,C_50),
      ('kabu','イベント＋5,0(かぶり 8/5・10・25)',3,30000,50000,C_KABU)]
rows=[]
for k,name,n,a0,a1,fill in plan:
    s=J[k]; od=s['avg']/s['aov']; q=max(0,min(1,(s['aov']-FAN_AOV)/(SC_AOV-FAN_AOV)))
    sc_o=od*q; cvr0=sc_o/a0; cvr1=cvr0*DROP
    b4=round(sc_o*SC_AOV); af=round(a1*cvr1*SC_AOV); fan=s['avg']-b4
    rows.append(dict(k=k,name=name,n=n,a0=a0,a1=a1,fill=fill,cvr0=cvr0,cvr1=cvr1,b4=b4,af=af,fan=fan))
SCB=sum(r['b4']*r['n'] for r in rows); SCA=sum(r['af']*r['n'] for r in rows)
FANM=sum(r['fan']*r['n'] for r in rows)
wb=openpyxl.Workbook()

# ===== S1 まとめ =====
ws=wb.active; ws.title="まとめ_スーツケース増額"
ws.merge_cells("A1:L1"); ws["A1"]="スーツケース増額プラン（7月公式ベース・8月公式カレンダー）"; ws["A1"].font=TITLE
ws.merge_cells("A2:L2"); ws["A2"]="増額前：平日2万/イベント3万アクセス・転換率据え置き(7月実績)。増額後：平日3万/イベント5万・転換率は少し下げて−10%で仮定。ハンディファンはアクセスそのまま。"; ws["A2"].font=SUB
heads=[("区分",30),("日数",6),("増額前\nアクセス",10),("転換率\n(現在)",10),("売上/日\n(現在)",13),("増額後\nアクセス",10),("転換率\n(増額後)",10),("売上/日\n(増額後)",13),("差額/日",12),("月間 現在",14),("月間 増額後",14),("月間差額",13)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
r=5
for x in rows:
    ws.cell(r,1,x['name']); ws.cell(r,2,x['n'])
    ws.cell(r,3,x['a0']).number_format=NUM
    ws.cell(r,4,x['cvr0']*100).number_format=PCT3
    ws.cell(r,5,x['b4']).number_format=YEN
    ws.cell(r,6,x['a1']).number_format=NUM
    ws.cell(r,7,x['cvr1']*100).number_format=PCT3
    ws.cell(r,8,x['af']).number_format=YEN
    ws.cell(r,9,x['af']-x['b4']).number_format=YEN
    ws.cell(r,10,x['b4']*x['n']).number_format=YEN
    ws.cell(r,11,x['af']*x['n']).number_format=YEN
    ws.cell(r,12,(x['af']-x['b4'])*x['n']).number_format=YEN
    for c in range(1,13):
        cell=ws.cell(r,c); cell.border=B; cell.fill=x['fill']; cell.alignment=LEFT if c==1 else CEN
        if c in(9,12): cell.font=BOLD
    r+=1
ws.cell(r,1,'スーツケース 計').font=BOLD
ws.cell(r,10,SCB).number_format=YEN; ws.cell(r,11,SCA).number_format=YEN; ws.cell(r,12,SCA-SCB).number_format=YEN
for c in range(1,13): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
r+=1
ws.cell(r,1,'ハンディファン 4型（アクセスそのまま＝据え置き）')
ws.cell(r,10,FANM).number_format=YEN; ws.cell(r,11,FANM).number_format=YEN; ws.cell(r,12,0).number_format=YEN
for c in range(1,13): ws.cell(r,c).fill=C_FAN; ws.cell(r,c).border=B
r+=1
ws.cell(r,1,'合算').font=font(bold=True,size=12)
ws.cell(r,10,SCB+FANM).number_format=YEN; ws.cell(r,11,SCA+FANM).number_format=YEN; ws.cell(r,12,SCA-SCB).number_format=YEN
for c in range(1,13): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=font(bold=True,size=11)
r+=2
ws.cell(r,1,'差額（増額−現在）').font=font(bold=True,size=13,color='C0392B')
c=ws.cell(r,3,SCA-SCB); c.number_format=YEN; c.font=font(bold=True,size=13,color='C0392B'); c.fill=C_DIFF
r+=2
alt=round(sum((x['a1']*x['cvr0']*0.8*SC_AOV-x['b4'])*x['n'] for x in rows))
for t in [f'転換率を−20%まで下げた保守ケース：スーツケース差額 +¥{alt:,} → 合算 ¥{SCB+FANM+alt:,}。',
 '増額前のスーツケース合計はSC/FAN分解(公式確定版)と1円一致。転換率は「SC注文数÷指定アクセス」で7月実績から逆算。',
 '5と0単独は7月実績が1日(7/15)のみでサンプル極少 → 8月実績で要更新。']:
    ws.cell(r,1,t).font=SUB; r+=1

# ===== S2 客単価UP施策 =====
w2=wb.create_sheet("客単価UP施策")
w2.column_dimensions['A'].width=40; w2.column_dimensions['B'].width=26; w2.column_dimensions['C'].width=52
w2.merge_cells("A1:C1"); w2["A1"]="客単価を上げる施策（アクセスを増やさず売上を増やす第3の道）"; w2["A1"].font=TITLE
data=[('施策',None,None),
 ('① 2個セット割：セットで¥5,000オフ','対象：多機能アルミ・PC多機能・ノーマルアルミ限定','客単価 ¥29,500 → 約¥54,000（2個で−¥5,000）。※訴求は「1個あたり¥2,500お得」等、見せ方は要検討'),
 ('② 圧縮バッグのセット買い','スーツケース＋圧縮バッグ','旅行動線で自然なクロスセル。客単価+¥2,000〜3,000'),
 ('③ ハンディファンセット','スーツケース＋ハンディファン','夏旅セット。FAN在庫の消化にも効く。客単価+¥3,000前後'),
 ('なぜ割引が広告より安いのか（数字の根拠）',None,None),
 ('広告で新規1件を獲るコスト(CPA)','平日：約¥7,900/件','必要アクセス=1÷転換率0.152%≒660 × アクセス単価¥12'),
 ('','かぶり日：約¥3,200/件','1÷0.373%≒268アクセス × ¥12'),
 ('セット割で「2個目」を売るコスト','¥5,000/個（割引原資のみ）','既に来ている客に売るため広告費ゼロ。平日CPA(¥7,900)より確実に安い'),
 ('結論',None,None),
 ('セット割は「平日の新規獲得」より安く売上を積める','特に平日・マラソン通常日に有効','山(かぶり日)はアクセス増で、谷(平日)は客単価UPで稼ぐ二刀流が最強')]
r=3
for a,b,c in data:
    if b is None:
        w2.cell(r,1,a).font=font(bold=True,color='FFFFFF'); 
        for cc in(1,2,3): w2.cell(r,cc).fill=FH; w2.cell(r,cc).border=B
    else:
        w2.cell(r,1,a).font=BOLD; w2.cell(r,2,b); w2.cell(r,3,c).font=SUB
        for cc in(1,2,3): w2.cell(r,cc).border=B
    r+=1
r+=1
w2.cell(r,1,'試算：平日にセット率10%が付いた場合').font=BOLD; r+=1
sets_day=rows[0]['af']/SC_AOV*0.10  # 増額後平日のSC注文×10%
gain=round(sets_day*(SC_AOV-5000))
w2.cell(r,1,f'増額後の平日SC注文 約{rows[0]["af"]/SC_AOV:.0f}件/日 × 10% × (2個目¥29,500−割引¥5,000) ≒ +¥{gain:,}/日 → 月間(平日17日) +¥{gain*17:,}').font=font()

# ===== S3 グラフ =====
w3=wb.create_sheet("グラフ")
w3["A1"]="グラフ"; w3["A1"].font=TITLE
w3["H1"]="区分"; w3["I1"]="現在"; w3["J1"]="増額後"
labels=['平日','マラソン通常','ワンダフル','5と0単独','かぶり日']
for i,x in enumerate(rows):
    w3.cell(2+i,8,labels[i]); w3.cell(2+i,9,x['b4']); w3.cell(2+i,10,x['af'])
w3["H8"]="区分"; w3["I8"]="現在"; w3["J8"]="増額後"
w3["H9"]="スーツケース"; w3["I9"]=SCB; w3["J9"]=SCA
w3["H10"]="ハンディファン"; w3["I10"]=FANM; w3["J10"]=FANM
w3["H11"]="合算"; w3["I11"]=SCB+FANM; w3["J11"]=SCA+FANM
ch=BarChart(); ch.type="col"; ch.title="スーツケース 1日あたり売上：現在 vs 増額後（転換率−10%）"
ch.height=10; ch.width=24; ch.gapWidth=60
ch.add_data(Reference(w3,min_col=9,max_col=10,min_row=1,max_row=6),titles_from_data=True)
ch.set_categories(Reference(w3,min_col=8,min_row=2,max_row=6))
w3.add_chart(ch,"A3")
ch2=BarChart(); ch2.type="col"; ch2.title="月間：SC/FAN/合算（現在 vs 増額後）"
ch2.height=10; ch2.width=18; ch2.gapWidth=60
ch2.add_data(Reference(w3,min_col=9,max_col=10,min_row=8,max_row=11),titles_from_data=True)
ch2.set_categories(Reference(w3,min_col=8,min_row=9,max_row=11))
w3.add_chart(ch2,"A24")

# ===== S4 前提と検算 =====
w4=wb.create_sheet("前提と検算"); w4.column_dimensions['A'].width=98
L=[('前提',BOLD),
 ('  ・基礎数値＝2026年7月実績（公式カレンダー分類）。スーツケース客単価¥29,500（冬1-2月実績）。',None),
 ('  ・スーツケース転換率＝7月のSC注文数 ÷ 指定アクセス（平日2万/イベント3万）で逆算。',None),
 ('  ・増額後の転換率＝現在×0.90（「少し下げ」の仮定）。保守ケース×0.80も併記。',None),
 ('  ・ハンディファン＝アクセス・売上とも据え置き（月間¥9,276,649）。',None),
 ('',None),('検算',BOLD),
 (f'  ・SC増額前 月間¥{SCB:,} ＝ 公式確定版のSC分解と一致。',None),
 (f'  ・合算：現在¥{SCB+FANM:,} → 増額後¥{SCA+FANM:,}（差額+¥{SCA-SCB:,}）。',None),
 ('  ・各行：売上/日×日数＝月間、を全行で確認。',None)]
for i,(t,f) in enumerate(L,1): w4.cell(i,1,t).font=f or font()

out="Libetee_スーツケース増額プラン.xlsx"
wb.save(out); print('saved',out)
print(f'SC: {SCB:,} → {SCA:,} (+{SCA-SCB:,})')
print(f'FAN据置: {FANM:,}')
print(f'合算: {SCB+FANM:,} → {SCA+FANM:,}')
print('検算 SC増額前=公式確定版:', 'OK' if SCB==37996670 else f'差異 {SCB-37996670}')
