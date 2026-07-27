import json, math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F="Arial"; YEN='¥#,##0'; NUM='#,##0'; PM='+#,##0;-#,##0'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15)
SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
REDF=PatternFill("solid",fgColor="FFC7CE"); YEL=PatternFill("solid",fgColor="FFF2CC")
GRN=PatternFill("solid",fgColor="C6EFCE"); GRY=PatternFill("solid",fgColor="E7E6E6")
BLU=PatternFill("solid",fgColor="BDD7EE")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")

# ===== 需要モデル（決定版6本柱・日別モデルと同一） =====
J=json.load(open('july_seg_official.json'))
SC_AOV=29500; FAN_AOV=3500; DROP=0.9; SETF=1+0.05*24500/SC_AOV
UNITS={}
for k,(a0,a1) in {'hei':(20000,30000),'mar':(30000,50000),'won':(30000,50000),'f50':(30000,50000),'kabu':(30000,50000)}.items():
    s=J[k]; od=s['avg']/s['aov']; q=max(0,min(1,(s['aov']-FAN_AOV)/(SC_AOV-FAN_AOV)))
    cvr0=od*q/a0
    UNITS[k]=a1*cvr0*DROP*SETF
CAL={1:'won',5:'kabu',10:'kabu',25:'kabu',15:'f50',20:'f50',30:'f50'}
for d in (4,6,7,8,9,24,26): CAL[d]='mar'
for d in range(1,32): CAL.setdefault(d,'hei')
u_total=sum(UNITS[CAL[d]] for d in range(1,32))
u_pre=sum(UNITS[CAL[d]] for d in range(1,15))     # 8/1-14 入荷前
SH_PRE=u_pre/u_total
SCALE=1.7275

# (商品, 7月個数, 現在庫, 入荷8/15, 区分)
SKUS=[
 ('多機能PC マットブラックS',218,384,550,'PC'),('多機能PC マットシルバーS',124,277,25,'PC'),
 ('多機能PC エナメルシルバーS',110,131,6,'PC'),('多機能PC マットホワイトS',69,106,0,'PC'),
 ('多機能PC エナメルカーキS',31,62,17,'PC'),('多機能PC エナメルホワイトS',28,144,196,'PC'),
 ('多機能PC マットグレーS',21,3,286,'PC'),('多機能PC スカイブルーS',10,3,0,'PC'),
 ('多機能PC エナメルピンクS',5,0,0,'PC'),('多機能PC アルミカスタムS',6,11,0,'PC'),
 ('多機能PC マットブラックM',97,50,309,'PC'),('多機能PC マットシルバーM',75,373,6,'PC'),
 ('多機能PC エナメルシルバーM',60,98,213,'PC'),('多機能PC マットホワイトM',51,183,0,'PC'),
 ('多機能PC エナメルカーキM',42,93,3,'PC'),('多機能PC マットグレーM',9,11,194,'PC'),
 ('多機能PC マットシルバーL',39,5,0,'PC'),('多機能PC マットグレーL',32,116,0,'PC'),
 ('多機能PC エナメルシルバーL',23,334,0,'PC'),('多機能PC マットホワイトL',18,91,0,'PC'),
 ('多機能PC マットブラックL',11,6,0,'PC'),
 ('ノーマルアルミ(クラシック)',32,517,0,'ALU'),('多機能アルミ(フルアルミ)',14,11,0,'ALU'),
 ('首振り(3Way冷却)',1577,None,0,'FAN'),('スケルトン',717,None,0,'FAN'),
 ('クリップ(CF)',551,0,0,'FAN'),('ミニファン(MSF)',88,4209,0,'FAN')]

wb=openpyxl.Workbook()

# ===== S1 入荷サマリー =====
ws=wb.active; ws.title="入荷サマリー"
ws.merge_cells("A1:E1"); ws["A1"]="8/15頃 入荷予定（PO: LM260507・40GP×2本・計1,805個）"; ws["A1"].font=TITLE
ws.merge_cells("A2:E2"); ws["A2"]="パッキングリスト2枚より。コンテナ①=Sサイズ1,060個、コンテナ②=Mサイズ725個+Sサイズ20個。Lサイズ・ハンディファンは今回入荷なし。"; ws["A2"].font=SUB
for j,(h,w) in enumerate([("コンテナ",16),("サイズ",8),("色",18),("数量",9),("メモ",30)],1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
rows=[('TWIU4232923','S','マットブラック',550,'8月の主力。これで安心'),
 ('TWIU4232923','S','マットグレー',286,'現在庫3→一気に回復'),
 ('TWIU4232923','S','エナメルホワイト',196,''),
 ('TWIU4232923','S','マットシルバー',25,''),
 ('TWIU4232923','S','エナメルシルバー',3,'※必要数に対して少ない'),
 ('IAAU1022523','M','マットブラック',309,'最大の懸念だったM黒が回復'),
 ('IAAU1022523','M','エナメルシルバー',213,''),
 ('IAAU1022523','M','マットグレー',194,''),
 ('IAAU1022523','M','マットシルバー',6,''),
 ('IAAU1022523','M','エナメルカーキ',3,''),
 ('IAAU1022523','S','エナメルカーキ',17,''),
 ('IAAU1022523','S','エナメルシルバー',3,'')]
r=5
for cont,sz,col,qty,memo in rows:
    ws.cell(r,1,cont); ws.cell(r,2,sz); ws.cell(r,3,col)
    ws.cell(r,4,qty).number_format=NUM; ws.cell(r,5,memo).font=SUB
    for c in range(1,6): ws.cell(r,c).border=B; ws.cell(r,c).alignment=LEFT if c in(3,5) else CEN
    r+=1
ws.cell(r,1,'合計').font=BOLD; ws.cell(r,4,1805).number_format=NUM; ws.cell(r,4).font=BOLD
for c in range(1,6): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B
r+=2
for t in ['注意①：Lサイズ（マットシルバーL・マットブラックL等）は今回のコンテナに入っていない。',
 '注意②：ハンディファン（首振り・スケルトン・クリップ）も入荷なし。在庫確認は引き続き最優先。',
 '注意③：着日が8/15「頃」のため、遅延した場合は8/15・20の5と0に間に合わないリスクあり。着日確定を推奨。']:
    ws.cell(r,1,t).font=SUB; r+=1

# ===== S2 SKU別タイムライン =====
w2=wb.create_sheet("SKU別_入荷前後の過不足")
w2.merge_cells("A1:J1"); w2["A1"]="SKU別 在庫タイムライン：入荷前(8/1〜14)に持つか × 月末までに足りるか"; w2["A1"].font=TITLE
w2.merge_cells("A2:J2"); w2["A2"]=f"8月必要数＝7月実績×{SCALE}(決定版)。入荷前必要＝8月必要×{SH_PRE:.1%}（8/1〜14の販売ペース、かぶり8/5・10を含む）。判定：赤=不足/黄=ギリギリ/緑=OK/グレー=過剰。"; w2["A2"].font=SUB
heads=[("商品（色×サイズ）",26),("8月必要",9),("現在庫",9),("入荷前必要\n(8/1-14)",11),("入荷前\n過不足",10),("8/15入荷",9),("入荷後在庫",10),("月末残",9),("入荷前判定",13),("月末判定",13)]
for j,(h,w) in enumerate(heads,1):
    c=w2.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w2.column_dimensions[get_column_letter(j)].width=w
r=5; pre_alerts=[]; end_alerts=[]
for name,jul,stock,inc,kind in SKUS:
    need=math.ceil(jul*SCALE) if kind!='FAN' else math.ceil(jul*31/28)
    need_pre=math.ceil(need*SH_PRE) if kind!='FAN' else math.ceil(need*14/31)
    w2.cell(r,1,name); w2.cell(r,2,need).number_format=NUM
    if stock is None:
        w2.cell(r,3,'リスト外'); w2.cell(r,4,need_pre).number_format=NUM
        w2.cell(r,5,'—'); w2.cell(r,6,inc).number_format=NUM; w2.cell(r,7,'—'); w2.cell(r,8,'—')
        v1=v2='⚠在庫データ無し'; f1=f2=REDF
        pre_alerts.append(f'{name}: 在庫不明のまま（8月必要{need}個・今回入荷なし）')
    else:
        pre_gap=stock-need_pre
        after=stock+inc
        end=after-need
        w2.cell(r,3,stock).number_format=NUM; w2.cell(r,4,need_pre).number_format=NUM
        w2.cell(r,5,pre_gap).number_format=PM; w2.cell(r,6,inc).number_format=NUM
        w2.cell(r,7,after).number_format=NUM; w2.cell(r,8,end).number_format=PM
        if pre_gap<0: v1='不足'; f1=REDF; pre_alerts.append(f'{name}: 入荷前に{-pre_gap}個不足（在庫{stock}/8月14日まで必要{need_pre}）')
        elif pre_gap<need_pre*0.15: v1='ギリギリ'; f1=YEL
        else: v1='OK'; f1=GRN
        if end<0: v2='不足'; f2=REDF; end_alerts.append(f'{name}: 月末までに{-end}個不足（入荷後{after}/必要{need}）')
        elif need>0 and after>need*3: v2='過剰'; f2=GRY
        elif end<need*0.15: v2='ギリギリ'; f2=YEL
        else: v2='OK'; f2=GRN
    c=w2.cell(r,9,v1); c.fill=f1; c.alignment=CEN
    c=w2.cell(r,10,v2); c.fill=f2; c.alignment=CEN
    for cc in range(1,9): w2.cell(r,cc).border=B
    w2.cell(r,9).border=B; w2.cell(r,10).border=B
    r+=1
w2.freeze_panes="A5"

# ===== S3 残アラート =====
w3=wb.create_sheet("残アラートと打ち手")
w3.column_dimensions['A'].width=105
w3['A1']='入荷を反映しても残る問題と打ち手'; w3['A1'].font=TITLE
msgs=['【解決】マットブラックS(+550)・マットグレーS(+286)・マットブラックM(+309)・エナメルシルバーM(+213)・マットグレーM(+194)',
 '　→ 8/15入荷で月末まで問題なし。第2の山（8/21〜31マラソン②）は戦える。',
 '',
 '【残る問題①：入荷前(8/1〜14)のギャップ】※かぶり日8/5・8/10がこの期間',]
msgs+=['　・'+a for a in pre_alerts]
msgs+=['　→ 打ち手：(1)着日の前倒し交渉(8/10前着なら山に間に合う) (2)不足色は8/4〜14の広告・ページ露出を在庫色(S銀277/M銀373/M白183)へ寄せる',
 '　　(3)売り切れ表示ではなく「8/15入荷予約」表示にして注文を逃さない（楽天の予約販売設定）。',
 '',
 '【残る問題②：Lサイズは今回入荷ゼロ】']
msgs+=['　・'+a for a in end_alerts if 'L' in a.split(':')[0]]
msgs+=['　→ L需要(8月約110個)は在庫のあるエナメルシルバーL(334)・マットグレーL(116)・マットホワイトL(91)へ誘導。次回発注にL主力色を必ず入れる。',
 '',
 '【残る問題③：ハンディファンは入荷なし＆首振り・スケルトンの在庫が依然不明】',
 '　→ FAN売上¥9.3Mの78%がこの2機種。在庫数の確認が8月計画の最後のピース。',
 '',
 '結論：8/15入荷で「月末まで」はほぼ解決。勝負は8/1〜14の入荷前ギャップ（特にM黒・S色欠け）とL/FANの2点。']
for i,t in enumerate(msgs,3):
    c=w3.cell(i,1,t); c.font=font(bold=t.startswith('【') or t.startswith('結論'))
    c.alignment=Alignment(wrap_text=True,vertical="center")

out="Libetee_8月_入荷反映_在庫計画.xlsx"
wb.save(out)
print('saved',out)
print('SH_PRE=',round(SH_PRE,4))
print('入荷前不足:',len(pre_alerts)); [print(' -',a) for a in pre_alerts]
print('月末不足:',len(end_alerts)); [print(' -',a) for a in end_alerts]
