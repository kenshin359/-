import datetime as dt, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; YEN='¥#,##0'; NUM='#,##0'; PCTF='0.00"%"'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_SC=PatternFill("solid",fgColor="D6E4F0"); C_FAN=PatternFill("solid",fgColor="DDEBD8")
C_KABU=PatternFill("solid",fgColor="F4B183"); C_50=PatternFill("solid",fgColor="C6EFCE")
C_WON=PatternFill("solid",fgColor="BDD7EE"); C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2")
C_DIFF=PatternFill("solid",fgColor="FFE2DD")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
J=json.load(open('july_seg_official.json'))
SC_AOV=29500; FAN_AOV=3500
plan=[('kabu','かぶり日（8/5・10・25）',3,10000,C_KABU),
      ('f50','5と0単独（8/15・20・30）',3,10000,C_50),
      ('won','ワンダフルデー（8/1）',1,10000,C_WON),
      ('mar','マラソン通常（8/4・6・7・8・9・24・26）',7,10000,C_MAR),
      ('hei','平日（17日）',17,5000,C_HEI)]
rows=[]
for k,name,n,da,fill in plan:
    s=J[k]; aov=s['aov']
    q=max(0,min(1,(aov-FAN_AOV)/(SC_AOV-FAN_AOV))); share=q*SC_AOV/aov
    add=round(da*s['vpa'])
    sc_b=round(s['avg']*share); fan_b=s['avg']-sc_b
    sc_a=round(add*share); fan_a=add-sc_a
    rows.append(dict(k=k,name=name,n=n,da=da,fill=fill,share=share,avg=s['avg'],add=add,
                     cvr=s['cvr'],aov=aov,vpa=s['vpa'],sc_b=sc_b,fan_b=fan_b,sc_a=sc_a,fan_a=fan_a))
TB=sum(r['avg']*r['n'] for r in rows); TA=sum(r['add']*r['n'] for r in rows)
SCB=sum(r['sc_b']*r['n'] for r in rows); SCA=sum(r['sc_a']*r['n'] for r in rows)
FANB=TB-SCB; FANA=TA-SCA
wb=openpyxl.Workbook()

# S1 差額まとめ（公式）
ws=wb.active; ws.title="差額まとめ_公式7月ベース"
ws.merge_cells("A1:J1"); ws["A1"]="通常予算 vs 増額【7月公式カレンダー確定版・転換率キープ】"; ws["A1"].font=TITLE
ws.merge_cells("A2:J2"); ws["A2"]="公式判明：7月はマラソン2回（7/4-11・7/19-26）→7/20・7/25はかぶり日。旧推定版の分類を修正済み。増額＝平日+5千/イベント+1万アクセス。"; ws["A2"].font=SUB
heads=[("セグメント",34),("日数",7),("7月転換率",10),("7月客単価",11),("1アクセス\n価値",10),("ペース 売上/日",14),("差額/日",13),("月間ベース",15),("月間差額",14),("月間 増額後",15)]
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
r=5
for x in rows:
    ws.cell(r,1,x['name']); ws.cell(r,2,x['n'])
    ws.cell(r,3,x['cvr']*100).number_format=PCTF
    ws.cell(r,4,x['aov']).number_format=YEN
    ws.cell(r,5,round(x['vpa'],1))
    ws.cell(r,6,x['avg']).number_format=YEN
    ws.cell(r,7,x['add']).number_format=YEN
    ws.cell(r,8,x['avg']*x['n']).number_format=YEN
    ws.cell(r,9,x['add']*x['n']).number_format=YEN
    ws.cell(r,10,(x['avg']+x['add'])*x['n']).number_format=YEN
    for c in range(1,11):
        cell=ws.cell(r,c); cell.border=B; cell.fill=x['fill']; cell.alignment=LEFT if c==1 else CEN
        if c in(7,9): cell.font=BOLD
    r+=1
ws.cell(r,1,'合計（8月・31日）').font=BOLD
ws.cell(r,8,TB).number_format=YEN; ws.cell(r,9,TA).number_format=YEN; ws.cell(r,10,TB+TA).number_format=YEN
for c in range(1,11): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
r+=2
ws.cell(r,1,'8月 月商（通常予算・7月公式ペース）').font=BOLD; ws.cell(r,3,TB).number_format=YEN; r+=1
ws.cell(r,1,'8月 月商（増額後）').font=BOLD; ws.cell(r,3,TB+TA).number_format=YEN; r+=1
ws.cell(r,1,'差額').font=font(bold=True,size=13,color='C0392B')
c=ws.cell(r,3,TA); c.number_format=YEN; c.font=font(bold=True,size=13,color='C0392B'); c.fill=C_DIFF
ws.cell(r,5,f'+{TA/TB*100:.0f}%').font=font(bold=True,color='C0392B'); r+=2
for t in ['旧推定版との差：ベース¥53.0M→¥47.3M／増額後¥66.8M→¥59.6M（7/20・25をかぶりへ、7/19・21-24をマラソンへ正しく再分類したため）。',
 '前年8月(¥26.3M)比：通常+80% / 増額後+127%。',
 '5と0単独は7月に7/15の1日しかなくサンプン極少（n=1）→ 8/15・20・30の予測は確度低め。',
 '追加アクセス22.5万×¥12≒追加広告費約¥270万 → 増分ROAS約4.6倍。絶対条件はスーツケース在庫。']:
    ws.cell(r,1,t).font=SUB; r+=1

# S2 SC/FAN合算
w1=wb.create_sheet("SC_FAN_合算")
w1.merge_cells("A1:F1"); w1["A1"]="スーツケース × ハンディファン 分解 → 合算（公式7月ベース）"; w1["A1"].font=TITLE
w1.merge_cells("A2:F2"); w1["A2"]="分解：冬(1-2月)客単価¥29,500=スーツケース、ファン¥3,500と置き各セグメント客単価から逆算。"; w1["A2"].font=SUB
heads=[("区分",20),("ベース(通常)",16),("増額の追加",15),("増額後",16),("構成比",10)]
for j,(h,w) in enumerate(heads,1):
    c=w1.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w1.column_dimensions[get_column_letter(j)].width=w
for i,(name,b,a,fl) in enumerate([('スーツケース',SCB,SCA,C_SC),('ハンディファン',FANB,FANA,C_FAN)]):
    r=5+i
    w1.cell(r,1,name).font=BOLD
    w1.cell(r,2,b).number_format=YEN; w1.cell(r,3,a).number_format=YEN
    w1.cell(r,4,b+a).number_format=YEN; w1.cell(r,5,(b+a)/(TB+TA)).number_format='0.0%'
    for c in range(1,6): w1.cell(r,c).border=B; w1.cell(r,c).fill=fl
r=7
w1.cell(r,1,'合算').font=BOLD
w1.cell(r,2,TB).number_format=YEN; w1.cell(r,3,TA).number_format=YEN
w1.cell(r,4,TB+TA).number_format=YEN; w1.cell(r,5,1).number_format='0.0%'
for c in range(1,6): w1.cell(r,c).fill=FT; w1.cell(r,c).border=B; w1.cell(r,c).font=BOLD
# セグメント別内訳
r=10
w1.cell(r,1,'セグメント別内訳（増額後・月間）').font=BOLD; r+=1
for j,(h,w) in enumerate([("セグメント",34),("スーツケース",16),("ハンディファン",16),("合計",16)],1):
    c=w1.cell(r,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B
r+=1
for x in rows:
    w1.cell(r,1,x['name'])
    w1.cell(r,2,(x['sc_b']+x['sc_a'])*x['n']).number_format=YEN
    w1.cell(r,3,(x['fan_b']+x['fan_a'])*x['n']).number_format=YEN
    w1.cell(r,4,(x['avg']+x['add'])*x['n']).number_format=YEN
    for c in range(1,5): w1.cell(r,c).border=B; w1.cell(r,c).fill=x['fill']
    r+=1

# S3 グラフ
w2=wb.create_sheet("グラフ")
w2["A1"]="グラフ（公式7月ベース）"; w2["A1"].font=TITLE
w2["H1"]="区分"; w2["I1"]="ベース"; w2["J1"]="増額後"
w2["H2"]="スーツケース"; w2["I2"]=SCB; w2["J2"]=SCB+SCA
w2["H3"]="ハンディファン"; w2["I3"]=FANB; w2["J3"]=FANB+FANA
w2["H4"]="合算"; w2["I4"]=TB; w2["J4"]=TB+TA
ch=BarChart(); ch.type="col"; ch.title="8月 月商：ベース vs 増額後（SC/FAN/合算）"
ch.height=10; ch.width=24; ch.gapWidth=60
ch.add_data(Reference(w2,min_col=9,max_col=10,min_row=1,max_row=4),titles_from_data=True)
ch.set_categories(Reference(w2,min_col=8,min_row=2,max_row=4))
w2.add_chart(ch,"A3")
short=['かぶり日','5と0単独','ワンダフル','マラソン通常','平日']
w2["H8"]="セグメント"; w2["I8"]="ベース/日"; w2["J8"]="増額後/日"
for i,x in enumerate(rows):
    w2.cell(9+i,8,short[i]); w2.cell(9+i,9,x['avg']); w2.cell(9+i,10,x['avg']+x['add'])
ch2=BarChart(); ch2.type="col"; ch2.title="1日あたり売上：ベース vs 増額後（セグメント別）"
ch2.height=10; ch2.width=24; ch2.gapWidth=60
ch2.add_data(Reference(w2,min_col=9,max_col=10,min_row=8,max_row=13),titles_from_data=True)
ch2.set_categories(Reference(w2,min_col=8,min_row=9,max_row=13))
w2.add_chart(ch2,"A24")

# S4 日別明細
w3=wb.create_sheet("日別明細")
w3.merge_cells("A1:I1"); w3["A1"]="8月 日別明細（公式カレンダー・公式7月ベース・転換率キープ）"; w3["A1"].font=TITLE
heads=[("日付",10),("曜日",6),("セグメント",20),("SC売上",13),("FAN売上",12),("ベース計",13),("増額後",13),("差額",12),("メモ",22)]
for j,(h,w) in enumerate(heads,1):
    c=w3.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w3.column_dimensions[get_column_letter(j)].width=w
def akey(day):
    if day==1: return 'won','ワンダフルデー',''
    if day in(5,10,25): return 'kabu','かぶり日','★マラソン×5と0'
    if day in(4,24): return 'mar','マラソン通常','本番20:00開始'
    if day in(6,7,8,9,26): return 'mar','マラソン通常',''
    if day in(15,20,30): return 'f50','5と0単独',''
    if day==11: return 'hei','平日','マラソン①終了01:59'
    if day==27: return 'hei','平日','マラソン②終了09:59'
    return 'hei','平日',''
rmap={x['k']:x for x in rows}
r=4; t1=t2=0; d=dt.date(2026,8,1)
while d.month==8:
    k,label,memo=akey(d.day); x=rmap[k]
    w3.cell(r,1,d).number_format='m/d(aaa)'; w3.cell(r,2,'月火水木金土日'[d.weekday()])
    w3.cell(r,3,label)
    w3.cell(r,4,x['sc_b']).number_format=YEN; w3.cell(r,5,x['fan_b']).number_format=YEN
    w3.cell(r,6,x['avg']).number_format=YEN; w3.cell(r,7,x['avg']+x['add']).number_format=YEN
    w3.cell(r,8,x['add']).number_format=YEN; w3.cell(r,9,memo)
    for c in range(1,10):
        cell=w3.cell(r,c); cell.border=B; cell.fill=x['fill']; cell.alignment=LEFT if c in(3,9) else CEN
    t1+=x['avg']; t2+=x['avg']+x['add']; r+=1; d+=dt.timedelta(days=1)
w3.cell(r,1,'合計').font=BOLD
w3.cell(r,6,t1).number_format=YEN; w3.cell(r,7,t2).number_format=YEN; w3.cell(r,8,t2-t1).number_format=YEN
for c in range(1,10): w3.cell(r,c).fill=FT; w3.cell(r,c).border=B; w3.cell(r,c).font=BOLD
w3.freeze_panes="A4"

# S5 前提と検算
w4=wb.create_sheet("前提と検算"); w4.column_dimensions['A'].width=98
L=[('公式カレンダーによる修正内容',BOLD),
 ('  ・7月マラソン② 7/19(日)20:00〜7/26(日)01:59 が公式判明 → 7/20・7/25は「かぶり日」、7/19・21〜24は「マラソン通常」に再分類。',None),
 ('  ・旧推定版：ベース¥53.0M/増額後¥66.8M → 公式版：ベース¥47.3M/増額後¥59.6M（下方修正）。',None),
 ('  ・7月のかぶり日は4日(7/5,10,20,25)：日商平均¥3,979,792・転換0.73%。',None),
 ('',None),('前提',BOLD),
 ('  ・転換率・客単価・販売ペース＝2026年7月実績（公式カレンダー分類）。転換率キープの理論値。',None),
 ('  ・SC/FAN分解：スーツケース客単価¥29,500(冬実績)・ファン¥3,500(価格帯)から逆算。',None),
 ('  ・5と0単独は7月n=1(7/15のみ)のため確度低め。8月実績で更新推奨。',None),
 ('',None),('検算',BOLD),
 (f'  ・まとめ={TB:,}/{TB+TA:,}、日別明細合計と一致を確認。SC({SCB+SCA:,})+FAN({FANB+FANA:,})=合算({TB+TA:,})。',None)]
for i,(t,f) in enumerate(L,1): w4.cell(i,1,t).font=f or font()

out="Libetee_8月シミュレーション_公式確定版.xlsx"
wb.save(out); print('saved',out)
print(f'ベース {TB:,} 増額後 {TB+TA:,} 差額 {TA:,}')
print(f'SC {SCB:,}+{SCA:,}={SCB+SCA:,} / FAN {FANB:,}+{FANA:,}={FANB+FANA:,}')
print('検算 SC+FAN=合算:', 'OK' if (SCB+FANB==TB and SCA+FANA==TA) else 'NG')
print('検算 日別=まとめ:', 'OK' if (t1==TB and t2==TB+TA) else 'NG')
