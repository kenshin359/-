import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
F="Arial"; YEN='¥#,##0'; NUM='#,##0'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
REDF=PatternFill("solid",fgColor="FFC7CE"); YEL=PatternFill("solid",fgColor="FFF2CC"); GRN=PatternFill("solid",fgColor="C6EFCE"); GRY=PatternFill("solid",fgColor="E7E6E6")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
SC_SH=[0.421,0.238,0.341]; FAN_SH=[0.294,0.341,0.365]
SCALE=1.7275  # 8月決定版 ÷ 7月実績
# (商品, 7月個数, 現在庫, 区分)  区分: PC/ALU/FAN
SKUS=[
 ('多機能PC マットブラックS',218,384,'PC'),('多機能PC マットシルバーS',124,277,'PC'),
 ('多機能PC エナメルシルバーS',110,131,'PC'),('多機能PC マットホワイトS',69,106,'PC'),
 ('多機能PC エナメルカーキS',31,62,'PC'),('多機能PC エナメルホワイトS',28,144,'PC'),
 ('多機能PC マットグレーS',21,3,'PC'),('多機能PC スカイブルーS',10,3,'PC'),
 ('多機能PC エナメルピンクS',5,0,'PC'),('多機能PC アルミカスタムS',6,11,'PC'),
 ('多機能PC マットブラックM',97,50,'PC'),('多機能PC マットシルバーM',75,373,'PC'),
 ('多機能PC エナメルシルバーM',60,98,'PC'),('多機能PC マットホワイトM',51,183,'PC'),
 ('多機能PC エナメルカーキM',42,93,'PC'),('多機能PC マットグレーM',9,11,'PC'),
 ('多機能PC マットシルバーL',39,5,'PC'),('多機能PC マットグレーL',32,116,'PC'),
 ('多機能PC エナメルシルバーL',23,334,'PC'),('多機能PC マットホワイトL',18,91,'PC'),
 ('多機能PC マットブラックL',11,6,'PC'),
 ('ノーマルアルミ(クラシック)',32,517,'ALU'),('多機能アルミ(フルアルミ)',14,11,'ALU'),
 ('首振り(3Way冷却)',1577,None,'FAN'),('スケルトン',717,None,'FAN'),
 ('クリップ(CF)',551,0,'FAN'),('ミニファン(MSF)',88,4209,'FAN')]
wb=openpyxl.Workbook()

# S1 シミュレーション
ws=wb.active; ws.title="10日別シミュレーション"
ws.merge_cells("A1:G1"); ws["A1"]="8月 ¥6,500万 販売シミュレーション（決定版6本柱・10日ブロック）"; ws["A1"].font=TITLE
ws.merge_cells("A2:G2"); ws["A2"]="決定版＝攻めアクセス(平日3万/イベント5万)×転換率×0.9(コンテンツページで回復)＋セット率5%。"; ws["A2"].font=SUB
for j,(h,w) in enumerate([("期間",13),("主なイベント",30),("SC売上",14),("FAN売上",13),("期間合計",14),("累計",14),("SC販売個数",11)],1):
    c=ws.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
data=[('8/1-8/10','ワンダフル・マラソン①・★かぶり8/5,10',23981604,2731771,813),
 ('8/11-8/20','5と0(8/15,20)',13565462,3160206,460),
 ('8/21-8/31','マラソン②・★かぶり8/25・5と0(8/30)',19429300,3384672,659)]
r=5; cum=0
for period,ev,sc,fan,scu in data:
    cum+=sc+fan
    ws.cell(r,1,period).font=BOLD; ws.cell(r,2,ev)
    ws.cell(r,3,sc).number_format=YEN; ws.cell(r,4,fan).number_format=YEN
    ws.cell(r,5,sc+fan).number_format=YEN; ws.cell(r,6,cum).number_format=YEN
    ws.cell(r,7,scu).number_format=NUM
    for c in range(1,8): ws.cell(r,c).border=B; ws.cell(r,c).alignment=LEFT if c==2 else CEN
    r+=1
ws.cell(r,1,'月間').font=BOLD
ws.cell(r,3,56976366).number_format=YEN; ws.cell(r,4,9276649).number_format=YEN
ws.cell(r,5,66253015).number_format=YEN; ws.cell(r,6,66253015).number_format=YEN; ws.cell(r,7,1932).number_format=NUM
for c in range(1,8): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B; ws.cell(r,c).font=BOLD
r+=2
for t in ['8/1-10が最大の山（¥26.7M＝月の40%）。かぶり日8/5・8/10を含むため、8月頭の在庫が勝敗を決める。',
 '8/21-31の山（¥22.8M）はマラソン②。8/20までに補充が間に合うかが第2の勝負。',
 '※日次丸めにより決定版¥66,262,794と±1万円弱の差。']:
    ws.cell(r,1,t).font=SUB; r+=1

# S2 SKU別必要個数
w2=wb.create_sheet("SKU別必要個数_10日別")
w2.merge_cells("A1:I1"); w2["A1"]="SKU別 必要個数（10日おき）× 実在庫（7/28時点）"; w2["A1"].font=TITLE
w2.merge_cells("A2:I2"); w2["A2"]="8月必要数＝7月実績×1.73(決定版スケール)。ブロック配分 SC=42/24/34%・FAN=29/34/37%。判定：赤=不足/黄=ギリギリ/緑=OK/グレー=過剰。"; w2["A2"].font=SUB
for j,(h,w) in enumerate([("商品（色×サイズ）",26),("7月実績",9),("8月必要",9),("8/1-10",9),("8/11-20",9),("8/21-31",9),("現在庫",9),("過不足",9),("判定",16)],1):
    c=w2.cell(4,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w2.column_dimensions[get_column_letter(j)].width=w
r=5; alerts=[]
for name,jul,stock,kind in SKUS:
    need=math.ceil(jul*SCALE) if kind!='FAN' else math.ceil(jul*31/28)
    sh=SC_SH if kind!='FAN' else FAN_SH
    b1,b2=math.ceil(need*sh[0]),math.ceil(need*sh[1]); b3=need-b1-b2
    w2.cell(r,1,name); w2.cell(r,2,jul).number_format=NUM; w2.cell(r,3,need).number_format=NUM
    w2.cell(r,4,b1).number_format=NUM; w2.cell(r,5,b2).number_format=NUM; w2.cell(r,6,b3).number_format=NUM
    if stock is None:
        w2.cell(r,7,'リスト外'); w2.cell(r,8,'—'); verdict='⚠在庫データ無し→至急確認'; fill=REDF
        alerts.append(f'{name}: 在庫リストに存在しない（8月必要{need}個）')
    else:
        w2.cell(r,7,stock).number_format=NUM
        diff=stock-need
        w2.cell(r,8,diff).number_format='+#,##0;-#,##0'
        if stock<=0 or diff<0:
            verdict='不足→発注'; fill=REDF; alerts.append(f'{name}: 不足{-diff if diff<0 else need}個（在庫{stock}/必要{need}）')
        elif diff<need*0.15: verdict='ギリギリ'; fill=YEL
        elif need>0 and stock>need*3: verdict='過剰'; fill=GRY
        else: verdict='OK'; fill=GRN
    c=w2.cell(r,9,verdict); c.alignment=CEN
    for cc in range(1,10): w2.cell(r,cc).border=B; w2.cell(r,cc).fill=fill
    r+=1
w2.freeze_panes="A5"

# S3 緊急アラート
w3=wb.create_sheet("緊急アラート")
w3.column_dimensions['A'].width=100
w3['A1']='緊急アラート（在庫×8月計画の突き合わせ）'; w3['A1'].font=TITLE
msgs=['【最重要】首振り(F53K)・スケルトン(X05)が在庫リストに存在しない。ファン売上の78%を占める2枚看板。',
 '　→ 別倉庫/別管理なら在庫数を至急確認。本当に在庫ゼロなら8月FAN売上¥9.3Mの大半が消える。',
 '【重要】クリップ(CF001/002)：在庫1,280個あるが引当も1,280個＝出荷可能0の表示。引当の中身を確認。',
 '【不足SKU（スーツケース）】']+['　・'+a for a in alerts if '多機能' in a or 'アルミ' in a]+[
 '【構造問題】Sサイズ在庫は潤沢(1,110個)だがM/L主力色が薄い。8月はM/L比率が43%。',
 '　特にマットブラックM(在庫50/必要168)・マットシルバーL(5/68)・マットブラックL(6/19)は8/10までに枯渇見込み。',
 '【過剰在庫】ミニファンMSF系4,209個(需要の43か月分)＋F31系4,760個(7月売上ゼロ)。',
 '　→ 夏旅セット・おまけ・同梱プレゼントで消化を（保管費の垂れ流し防止）。',
 '',
 '結論：8月の勝敗は「マットブラックM・L系の補充」と「首振り/スケルトンの在庫確認」の2点に懸かっている。']
for i,t in enumerate(msgs,3):
    c=w3.cell(i,1,t); c.font=font(bold=t.startswith('【')) 
    c.alignment=Alignment(wrap_text=True,vertical="center")

out="Libetee_8月65M_10日別シミュレーション.xlsx"
wb.save(out); print('saved',out)
print('アラート数:',len(alerts))
