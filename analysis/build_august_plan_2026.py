import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
F="Arial"; YEN='¥#,##0'; PCT='0.00"%"'; XR='0.00"x"'; DATEF='m/d(aaa)'; YM='yyyy-mm'
def font(**k): return Font(name=F,**k)
HEADW=font(color="FFFFFF",bold=True,size=10); TITLE=font(bold=True,size=15); SUB=font(italic=True,size=9,color="666666"); BOLD=font(bold=True)
FH=PatternFill("solid",fgColor="1F4E78"); FT=PatternFill("solid",fgColor="D9E1F2")
C_KABU=PatternFill("solid",fgColor="F4B183"); C_KABU2=PatternFill("solid",fgColor="C0392B")
C_50=PatternFill("solid",fgColor="C6EFCE"); C_WON=PatternFill("solid",fgColor="BDD7EE")
C_MAR=PatternFill("solid",fgColor="E2EFDA"); C_HEI=PatternFill("solid",fgColor="F2F2F2"); C_IN=PatternFill("solid",fgColor="FFF2CC")
thin=Side(style="thin",color="BFBFBF"); B=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True); LEFT=Alignment(horizontal="left",vertical="center")
BASE=740103  # 2026平日ベースライン
wb=openpyxl.Workbook()

# ===== S1: 8月カレンダー（2026基準）=====
ws=wb.active; ws.title="8月_広告配分カレンダー"
ws.merge_cells("A1:H1"); ws["A1"]="2026年8月 楽天 広告配分カレンダー（今年データ基準・色付き）"; ws["A1"].font=TITLE
ws.merge_cells("A2:H2"); ws["A2"]="予想売上＝2026年平日¥740,103×各セグメント倍率(2026実績)。マラソン日程は【推定】→楽天公式で要置換。"; ws["A2"].font=SUB
for j,(h,w) in enumerate([("日付",11),("曜日",6),("セグメント",22),("5と0",7),("平日比",8),("予想売上",13),("広告優先",10),("メモ",26)],1):
    c=ws.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; ws.column_dimensions[get_column_letter(j)].width=w
def seg_for(d):
    day=d.day; mar=4<=day<=11; is50=day in(5,10,15,20,25,30)
    if day==1: return('ワンダフルデー',1.46,'A',C_WON,'月初ポイント。上乗せ')
    if mar and is50: return('マラソン×5と0 かぶり',3.02,'S 最優先',C_KABU,'★最強。広告を最も厚く')
    if mar: return('お買い物マラソン(推定)',1.11,'C 盛らない',C_MAR,'普通日は平日並み。厚くしない')
    if day in(15,20,25,30): return('5と0のつく日',2.20,'A',C_50,'転換率UP。厚めに')
    return('平日',1.00,'底維持',C_HEI,'切らない。薄い一定額')
r=4; tot=0; d=dt.date(2026,8,1)
while d.month==8:
    name,lift,pri,fill,memo=seg_for(d); s=round(BASE*lift)
    ws.cell(r,1,d).number_format=DATEF; ws.cell(r,2,'月火水木金土日'[d.weekday()])
    ws.cell(r,3,name); ws.cell(r,4,'●' if d.day in(5,10,15,20,25,30) else '')
    ws.cell(r,5,lift).number_format=XR; ws.cell(r,6,s).number_format=YEN; ws.cell(r,7,pri); ws.cell(r,8,memo)
    for c in range(1,9):
        cell=ws.cell(r,c); cell.border=B; cell.fill=fill; cell.alignment=LEFT if c in(3,8) else CEN
    tot+=s; r+=1; d+=dt.timedelta(days=1)
ws.cell(r,1,'月商予測').font=BOLD; ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
ws.cell(r,6,tot).number_format=YEN; ws.cell(r,6).font=BOLD
ws.cell(r,7,'前年比').font=BOLD; ws.cell(r,8,f'2025/8=¥26.3M → {tot/26325956-1:+.0%}')
for c in range(1,9): ws.cell(r,c).fill=FT; ws.cell(r,c).border=B
ws.freeze_panes="A4"

# ===== S2: セグメント別統計（2026）=====
w2=wb.create_sheet("セグメント別統計_2026")
w2.merge_cells("A1:F1"); w2["A1"]="楽天 セグメント別 売上・転換率（2026年1〜7月・平日=1.00）"; w2["A1"].font=TITLE
for j,(h,w) in enumerate([("セグメント",26),("日数",7),("日平均売上",14),("平日比",9),("転換率",9),("客単価",12)],1):
    c=w2.cell(2,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w2.column_dimensions[get_column_letter(j)].width=w
rows=[('スーパーSALE × 5と0 かぶり',4,3567832,4.82,0.66,30429,C_KABU2,True),
      ('マラソン × 5と0 かぶり',10,2234584,3.02,0.54,24801,C_KABU,True),
      ('5と0のつく日（単独）',26,1631874,2.20,0.46,20127,C_50,False),
      ('スーパーSALE（5と0以外）',12,1158456,1.57,0.26,30286,C_MAR,False),
      ('ワンダフルデー',7,1081629,1.46,0.35,30653,C_WON,False),
      ('お買い物マラソン（5と0以外）',30,824846,1.11,0.26,24236,C_MAR,False),
      ('平日',118,740103,1.00,0.22,22838,C_HEI,False)]
r=3
for name,n,s,lift,cvr,aov,fill,strong in rows:
    w2.cell(r,1,name); w2.cell(r,2,n); w2.cell(r,3,s).number_format=YEN
    w2.cell(r,4,lift).number_format=XR; w2.cell(r,5,cvr).number_format=PCT; w2.cell(r,6,aov).number_format=YEN
    for c in range(1,7):
        cell=w2.cell(r,c); cell.border=B; cell.fill=fill; cell.alignment=LEFT if c==1 else CEN
        if strong: cell.font=BOLD
    r+=1
for t in ['結論：今年も売上・転換率は「イベント×5と0の重なり」に集中。SALE×5と0=×4.82、マラソン普通日=×1.11(平日並み)。',
          '広告配分：山(5と0＋かぶり)55% ／ SALE・マラソン平常日20% ／ 平日の底25%。平日はゼロにしない。',
          '※お気に入り率は本データ(売上・アクセス)に無し。RMS R-Karte「お気に入り分析」で追加可能。']:
    w2.cell(r+1,1,t).font=SUB; r+=1

# ===== S3: 月商推移(2026) =====
w3=wb.create_sheet("月商推移_2026")
w3.merge_cells("A1:D1"); w3["A1"]="月商推移（2026年）"; w3["A1"].font=TITLE
for j,(h,w) in enumerate([("月",10),("月商",16),("日数",8),("日平均",14)],1):
    c=w3.cell(2,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w3.column_dimensions[get_column_letter(j)].width=w
md=[('2026-01',21775522,31),('2026-02',21626250,28),('2026-03',30688230,31),('2026-04',35044550,30),
    ('2026-05',26108970,31),('2026-06',33432028,30),('2026-07',43920694,26)]
r=3
for ym,s,n in md:
    w3.cell(r,1,ym); w3.cell(r,2,s).number_format=YEN; w3.cell(r,3,n); w3.cell(r,4,round(s/n)).number_format=YEN
    for c in range(1,5): w3.cell(r,c).border=B; w3.cell(r,c).alignment=CEN
    r+=1
tot=sum(s for _,s,_ in md)
w3.cell(r,1,'合計').font=BOLD; w3.cell(r,2,tot).number_format=YEN
for c in range(1,5): w3.cell(r,c).fill=FT; w3.cell(r,c).border=B; w3.cell(r,c).font=BOLD
w3.cell(r+2,1,'※2026/07は26日時点（月途中）。7月は過去最高ペース。').font=SUB

# ===== S4: イベント日マスタ(公式) =====
w4=wb.create_sheet("イベント日マスタ_公式")
w4.merge_cells("A1:E1"); w4["A1"]="楽天 イベント日マスタ（★公式データで管理）"; w4["A1"].font=TITLE
w4.merge_cells("A2:E2"); w4["A2"]="公式日程を黄色セルに入力。区分に『公式』で確定。推定は必ず公式へ置換。取得元：RMS 店舗運営Navi イベントカレンダー。"; w4["A2"].font=SUB
for j,(h,w) in enumerate([("開始日",13),("終了日",13),("イベント名",22),("区分",12),("備考",30)],1):
    c=w4.cell(3,j,h); c.font=HEADW; c.fill=FH; c.alignment=CEN; c.border=B; w4.column_dimensions[get_column_letter(j)].width=w
sample=[('2026-08-01','2026-08-01','ワンダフルデー','公式(定例)','毎月1日'),
        ('2026-08-04','2026-08-11','お買い物マラソン','推定→要公式','公式カレンダーで確定'),
        ('2026-08-05','2026-08-05','5と0のつく日','公式(定例)','毎月5/10/15/20/25/30'),
        ('2026-09-04','2026-09-11','スーパーSALE','推定→要公式','9月開催。公式で確定')]
r=4
for s in sample:
    est='推定' in s[3]
    for j,v in enumerate(s,1):
        cell=w4.cell(r,j,v); cell.border=B; cell.alignment=LEFT
        if j==4 and est: cell.fill=PatternFill("solid",fgColor="FFC7CE")
        elif j in(1,2,4): cell.fill=C_IN
    r+=1
for rr in range(r,r+8):
    for j in range(1,6): w4.cell(rr,j).border=B; w4.cell(rr,j).fill=C_IN

# ===== S5: 凡例 =====
w5=wb.create_sheet("凡例・読み方"); w5.column_dimensions["A"].width=95
L=[('Libetee 楽天 8月作戦（今年2026データ基準・色付き）',TITLE,None),('',None,None),
 ('【色の意味】',BOLD,None),
 ('  かぶり日（イベント×5と0）＝最強。広告最優先',None,C_KABU),
 ('  5と0のつく日（単独）＝厚めに',None,C_50),
 ('  ワンダフルデー＝上乗せ',None,C_WON),
 ('  お買い物マラソン普通日＝平日並み。盛らない',None,C_MAR),
 ('  平日＝底を維持（切らない）',None,C_HEI),
 ('  黄色＝入力セル（公式日程を記入）',None,C_IN),('',None,None),
 ('【重要】イベント日は必ず楽天公式データで管理。マラソンは現在推定。公式置換後に全分析を更新。',BOLD,None),
 ('【未取得】お気に入り率（R-Karte「お気に入り分析」データで追加可）。',None,None)]
for i,(t,f,fl) in enumerate(L,1):
    c=w5.cell(i,1,t); c.font=f or font(); c.alignment=Alignment(wrap_text=True,vertical="center")
    if fl: c.fill=fl

out="/tmp/claude-0/-home-user--/c7179f16-7eb7-52e8-aa4a-c082e5dbce5f/scratchpad/Libetee_楽天_8月作戦_2026.xlsx"
wb.save(out); print("saved",out,"| 8月予測 ¥{:,}".format(tot))
