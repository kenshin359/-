#!/usr/bin/env python3
# Libetee 広告費→予想シート（日次運用版）
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
YEN = '¥#,##0'; PCT = '0.00%'; NUM = '#,##0'; XR = '0.00"x"'
DATEF = 'yyyy/mm/dd'

BLUE = Font(name=FONT, color="0000FF")
BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000")
HEADW = Font(name=FONT, color="FFFFFF", bold=True, size=10)
TITLE = Font(name=FONT, bold=True, size=14)
SUB = Font(name=FONT, italic=True, size=9, color="555555")
BOLD = Font(name=FONT, bold=True)
SECW = Font(name=FONT, bold=True, color="FFFFFF")

FILL_HEAD = PatternFill("solid", fgColor="1F4E78")
FILL_INPUT = PatternFill("solid", fgColor="FFF7CC")
FILL_TOTAL = PatternFill("solid", fgColor="D9E1F2")
FILL_SEC = PatternFill("solid", fgColor="2E75B6")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

MEDIA = ["Amazon", "楽天", "自社サイト"]
PRODUCTS = [
    ("メタアカウント", "スーツケース（トラベル）"),
    ("メタアカウント", "スーツケース（カタログ）"),
    ("ハンディファンアカウント", "ガジェティ"),
    ("ツヤリスアカウント", "ツヤリス"),
]
ACCOUNTS = ["メタアカウント", "ハンディファンアカウント", "ツヤリスアカウント"]
PRODNAMES = [p for _, p in PRODUCTS]
# 商品→アカウント 逆引き
PROD2ACC = {p: a for a, p in PRODUCTS}

# 単価前提（商品×媒体）: CPA, アクセス単価, 転換率, 客単価
UNIT = {
    "スーツケース（トラベル）": {"Amazon": (6000, 60, 0.010, 15000), "楽天": (8000, 70, 0.006, 15000), "自社サイト": (7000, 55, 0.012, 16000)},
    "スーツケース（カタログ）": {"Amazon": (6500, 65, 0.009, 14000), "楽天": (8500, 75, 0.005, 14000), "自社サイト": (7200, 58, 0.011, 15000)},
    "ガジェティ": {"Amazon": (1500, 40, 0.020, 3200), "楽天": (1800, 45, 0.015, 3200), "自社サイト": (1600, 38, 0.022, 3500)},
    "ツヤリス": {"Amazon": (2200, 45, 0.018, 4200), "楽天": (2600, 50, 0.012, 4200), "自社サイト": (2400, 42, 0.020, 4500)},
}
# 日別のサンプル広告費（商品×媒体の基準額）
ADBASE = {
    "スーツケース（トラベル）": {"Amazon": 300000, "楽天": 250000, "自社サイト": 150000},
    "スーツケース（カタログ）": {"Amazon": 200000, "楽天": 180000, "自社サイト": 120000},
    "ガジェティ": {"Amazon": 180000, "楽天": 150000, "自社サイト": 90000},
    "ツヤリス": {"Amazon": 160000, "楽天": 140000, "自社サイト": 100000},
}
SAMPLE_DATES = [dt.date(2026, 7, 1), dt.date(2026, 7, 2), dt.date(2026, 7, 5)]
DAY_MULT = {dt.date(2026, 7, 1): 1.0, dt.date(2026, 7, 2): 0.9, dt.date(2026, 7, 5): 1.4}  # 5日=イベント寄せ

wb = openpyxl.Workbook()

# ============================================================
# シート1：日次入力
# ============================================================
ws = wb.active
ws.title = "日次入力"
MAIN = "日次入力"
headers = [
    ("日付", 12), ("アカウント", 20), ("商品/ブランド", 22), ("媒体", 12),
    ("広告費", 12), ("CPA\n(獲得単価)", 11), ("アクセス単価", 11), ("予想\n転換率", 9), ("客単価", 11),
    ("予想\nアクセス数", 12), ("予想獲得数\n(CPA基準)", 12), ("予想獲得数\n(ｱｸｾｽ×CVR)", 13),
    ("整合差異", 10), ("予想売上", 14), ("ROAS", 8),
]
ws.merge_cells("A1:O1")
ws["A1"] = "Libetee 日次入力（広告費 → 売上予想）"; ws["A1"].font = TITLE
ws.merge_cells("A2:O2")
ws["A2"] = "毎日、1行＝その日・その商品・その媒体。青字セル（黄色）だけ入力。上の3日分はサンプルです。"
ws["A2"].font = SUB

HROW = 3
for j, (h, w) in enumerate(headers, start=1):
    c = ws.cell(row=HROW, column=j, value=h)
    c.font = HEADW; c.fill = FILL_HEAD; c.alignment = CEN; c.border = BORDER
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[HROW].height = 30

DATA_START = 4
NROWS = 400          # 余裕をもって数式を敷いておく
DATA_END = DATA_START + NROWS - 1

# サンプル行を先頭に投入
sample_rows = []
for d in SAMPLE_DATES:
    for acc, prod in PRODUCTS:
        for m in MEDIA:
            cpa, cpc, cvr, aov = UNIT[prod][m]
            adv = round(ADBASE[prod][m] * DAY_MULT[d])
            sample_rows.append((d, acc, prod, m, adv, cpa, cpc, cvr, aov))

input_cols = {"E", "F", "G", "H", "I"}
for i in range(NROWS):
    r = DATA_START + i
    if i < len(sample_rows):
        d, acc, prod, m, adv, cpa, cpc, cvr, aov = sample_rows[i]
        ws.cell(row=r, column=1, value=d).number_format = DATEF
        ws.cell(row=r, column=2, value=acc)
        ws.cell(row=r, column=3, value=prod)
        ws.cell(row=r, column=4, value=m)
        ws.cell(row=r, column=5, value=adv)
        ws.cell(row=r, column=6, value=cpa)
        ws.cell(row=r, column=7, value=cpc)
        ws.cell(row=r, column=8, value=cvr)
        ws.cell(row=r, column=9, value=aov)
    # 計算列（全行に数式を敷く）
    ws.cell(row=r, column=10, value=f"=IFERROR(E{r}/G{r},0)")     # アクセス=広告費/アクセス単価
    ws.cell(row=r, column=11, value=f"=IFERROR(E{r}/F{r},0)")     # 獲得(CPA)=広告費/CPA
    ws.cell(row=r, column=12, value=f"=J{r}*H{r}")                # 獲得(アクセス×CVR)
    ws.cell(row=r, column=13, value=f"=K{r}-L{r}")                # 整合差異
    ws.cell(row=r, column=14, value=f"=K{r}*I{r}")               # 予想売上
    ws.cell(row=r, column=15, value=f"=IFERROR(N{r}/E{r},0)")   # ROAS
    # 書式
    for col in range(1, 16):
        cell = ws.cell(row=r, column=col); cell.border = BORDER
        L = get_column_letter(col)
        cell.font = BLUE if L in input_cols else BLACK
        if L in input_cols: cell.fill = FILL_INPUT
    ws.cell(row=r, column=1).number_format = DATEF
    ws.cell(row=r, column=2).alignment = LEFT
    ws.cell(row=r, column=3).alignment = LEFT
    for col in (5, 6, 7, 9): ws.cell(row=r, column=col).number_format = YEN
    ws.cell(row=r, column=8).number_format = PCT
    ws.cell(row=r, column=10).number_format = NUM
    ws.cell(row=r, column=11).number_format = '#,##0.0'
    ws.cell(row=r, column=12).number_format = '#,##0.0'
    ws.cell(row=r, column=13).number_format = '#,##0.0;[RED]-#,##0.0'
    ws.cell(row=r, column=14).number_format = YEN
    ws.cell(row=r, column=15).number_format = XR

# 入力補助：ドロップダウン
dv_acc = DataValidation(type="list", formula1='"%s"' % ",".join(ACCOUNTS), allow_blank=True)
dv_prod = DataValidation(type="list", formula1='"%s"' % ",".join(PRODNAMES), allow_blank=True)
dv_med = DataValidation(type="list", formula1='"%s"' % ",".join(MEDIA), allow_blank=True)
ws.add_data_validation(dv_acc); ws.add_data_validation(dv_prod); ws.add_data_validation(dv_med)
dv_acc.add(f"B{DATA_START}:B{DATA_END}")
dv_prod.add(f"C{DATA_START}:C{DATA_END}")
dv_med.add(f"D{DATA_START}:D{DATA_END}")
ws.freeze_panes = "E4"

def sum_by_date(sumcol, datecell):
    return f"=SUMIFS('{MAIN}'!${sumcol}${DATA_START}:${sumcol}${DATA_END},'{MAIN}'!$A${DATA_START}:$A${DATA_END},{datecell})"

def sumifs_month(sumcol, critcol, critcell, ms, me):
    base = f"'{MAIN}'!${sumcol}${DATA_START}:${sumcol}${DATA_END}"
    dr = f"'{MAIN}'!$A${DATA_START}:$A${DATA_END}"
    cr = f"'{MAIN}'!${critcol}${DATA_START}:${critcol}${DATA_END}"
    return f'=SUMIFS({base},{dr},">="&{ms},{dr},"<="&{me},{cr},{critcell})'

# ============================================================
# シート2：日別サマリー（月次トレンド）
# ============================================================
ws2 = wb.create_sheet("日別サマリー")
ws2.merge_cells("A1:G1"); ws2["A1"] = "日別サマリー（月次トレンド）"; ws2["A1"].font = TITLE
ws2["A2"] = "対象年月（1日を入力）"; ws2["A2"].font = BOLD
mcell = ws2.cell(row=2, column=2, value=dt.date(2026, 7, 1))
mcell.font = BLUE; mcell.fill = FILL_INPUT; mcell.number_format = 'yyyy"年"m"月"'; mcell.border = BORDER
ws2.cell(row=2, column=3, value="↑ この月の1日を入力（例 2026/7/1）").font = SUB

sh2 = [("日付", 12), ("広告費", 14), ("予想アクセス数", 13), ("予想獲得数", 12), ("予想売上", 15), ("ROAS", 9), ("月累計売上", 16)]
HR2 = 4
for j, (h, w) in enumerate(sh2, start=1):
    c = ws2.cell(row=HR2, column=j, value=h); c.font = HEADW; c.fill = FILL_HEAD; c.alignment = CEN; c.border = BORDER
    ws2.column_dimensions[get_column_letter(j)].width = w
d0 = HR2 + 1
for i in range(31):
    r = d0 + i
    dcell = f"$A{r}"
    ws2.cell(row=r, column=1, value=f"=$B$2+{i}").number_format = DATEF
    ws2.cell(row=r, column=2, value=sum_by_date("E", dcell))
    ws2.cell(row=r, column=3, value=sum_by_date("J", dcell))
    ws2.cell(row=r, column=4, value=sum_by_date("K", dcell))
    ws2.cell(row=r, column=5, value=sum_by_date("N", dcell))
    ws2.cell(row=r, column=6, value=f"=IFERROR(E{r}/B{r},0)")
    ws2.cell(row=r, column=7, value=f"=SUM($E${d0}:$E{r})")
    for col in range(1, 8):
        cell = ws2.cell(row=r, column=col); cell.border = BORDER; cell.font = BLACK
    ws2.cell(row=r, column=1).number_format = DATEF
    ws2.cell(row=r, column=2).number_format = YEN
    ws2.cell(row=r, column=3).number_format = NUM
    ws2.cell(row=r, column=4).number_format = '#,##0.0'
    ws2.cell(row=r, column=5).number_format = YEN
    ws2.cell(row=r, column=6).number_format = XR
    ws2.cell(row=r, column=7).number_format = YEN
d_end2 = d0 + 30
# 月合計行
tr = d_end2 + 1
ws2.cell(row=tr, column=1, value="月合計").font = BOLD
for col, L in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    ws2.cell(row=tr, column=col, value=f"=SUM({L}{d0}:{L}{d_end2})")
ws2.cell(row=tr, column=6, value=f"=IFERROR(E{tr}/B{tr},0)")
ws2.cell(row=tr, column=7, value=f"=E{tr}")
for col in range(1, 8):
    cell = ws2.cell(row=tr, column=col); cell.fill = FILL_TOTAL; cell.font = BOLD; cell.border = BORDER
ws2.cell(row=tr, column=2).number_format = YEN
ws2.cell(row=tr, column=3).number_format = NUM
ws2.cell(row=tr, column=4).number_format = '#,##0.0'
ws2.cell(row=tr, column=5).number_format = YEN
ws2.cell(row=tr, column=6).number_format = XR
ws2.cell(row=tr, column=7).number_format = YEN
ws2.freeze_panes = "A5"
MONTH_TOTAL_SALES = f"'日別サマリー'!$E${tr}"
MONTH_TOTAL_ADV = f"'日別サマリー'!$B${tr}"
MONTH_TOTAL_CV = f"'日別サマリー'!$D${tr}"
MSTART = "'日別サマリー'!$B$2"
MEND = "(DATE(YEAR('日別サマリー'!$B$2),MONTH('日別サマリー'!$B$2)+1,1)-1)"

# ============================================================
# シート3：商品別サマリー（当月）
# ============================================================
ws3 = wb.create_sheet("商品別サマリー")
ws3.merge_cells("A1:F1"); ws3["A1"] = "商品別サマリー（当月・日別サマリーの対象年月に連動）"; ws3["A1"].font = TITLE
sh3 = [("商品/ブランド", 24), ("広告費", 14), ("予想獲得数", 13), ("予想売上", 16), ("ブレンドCPA", 13), ("ROAS", 10)]
for j, (h, w) in enumerate(sh3, start=1):
    c = ws3.cell(row=2, column=j, value=h); c.font = HEADW; c.fill = FILL_HEAD; c.alignment = CEN; c.border = BORDER
    ws3.column_dimensions[get_column_letter(j)].width = w
r0 = 3
for i, p in enumerate(PRODNAMES):
    r = r0 + i
    ws3.cell(row=r, column=1, value=p).font = GREEN
    ws3.cell(row=r, column=2, value=sumifs_month("E", "C", f"$A{r}", MSTART, MEND))
    ws3.cell(row=r, column=4, value=sumifs_month("N", "C", f"$A{r}", MSTART, MEND))
    ws3.cell(row=r, column=3, value=sumifs_month("K", "C", f"$A{r}", MSTART, MEND))
    ws3.cell(row=r, column=5, value=f"=IFERROR(B{r}/C{r},0)")
    ws3.cell(row=r, column=6, value=f"=IFERROR(D{r}/B{r},0)")
rE = r0 + len(PRODNAMES) - 1
tr3 = rE + 1
ws3.cell(row=tr3, column=1, value="合計").font = BOLD
for col, L in [(2,"B"),(3,"C"),(4,"D")]:
    ws3.cell(row=tr3, column=col, value=f"=SUM({L}{r0}:{L}{rE})")
ws3.cell(row=tr3, column=5, value=f"=IFERROR(B{tr3}/C{tr3},0)")
ws3.cell(row=tr3, column=6, value=f"=IFERROR(D{tr3}/B{tr3},0)")
for r in range(r0, tr3 + 1):
    for col in range(1, 7):
        cell = ws3.cell(row=r, column=col); cell.border = BORDER
        if r == tr3: cell.fill = FILL_TOTAL; cell.font = BOLD
        elif col != 1: cell.font = BLACK
    ws3.cell(row=r, column=2).number_format = YEN
    ws3.cell(row=r, column=3).number_format = '#,##0.0'
    ws3.cell(row=r, column=4).number_format = YEN
    ws3.cell(row=r, column=5).number_format = YEN
    ws3.cell(row=r, column=6).number_format = XR

# ============================================================
# シート4：媒体別サマリー（当月）
# ============================================================
ws4 = wb.create_sheet("媒体別サマリー")
ws4.merge_cells("A1:F1"); ws4["A1"] = "媒体別サマリー（当月：Amazon / 楽天 / 自社サイト）"; ws4["A1"].font = TITLE
for j, (h, w) in enumerate(sh3, start=1):
    c = ws4.cell(row=2, column=j, value=h if h != "商品/ブランド" else "媒体"); c.font = HEADW; c.fill = FILL_HEAD; c.alignment = CEN; c.border = BORDER
    ws4.column_dimensions[get_column_letter(j)].width = w
r0 = 3
for i, m in enumerate(MEDIA):
    r = r0 + i
    ws4.cell(row=r, column=1, value=m).font = GREEN
    ws4.cell(row=r, column=2, value=sumifs_month("E", "D", f"$A{r}", MSTART, MEND))
    ws4.cell(row=r, column=3, value=sumifs_month("K", "D", f"$A{r}", MSTART, MEND))
    ws4.cell(row=r, column=4, value=sumifs_month("N", "D", f"$A{r}", MSTART, MEND))
    ws4.cell(row=r, column=5, value=f"=IFERROR(B{r}/C{r},0)")
    ws4.cell(row=r, column=6, value=f"=IFERROR(D{r}/B{r},0)")
rE = r0 + len(MEDIA) - 1
tr4 = rE + 1
ws4.cell(row=tr4, column=1, value="合計").font = BOLD
for col, L in [(2,"B"),(3,"C"),(4,"D")]:
    ws4.cell(row=tr4, column=col, value=f"=SUM({L}{r0}:{L}{rE})")
ws4.cell(row=tr4, column=5, value=f"=IFERROR(B{tr4}/C{tr4},0)")
ws4.cell(row=tr4, column=6, value=f"=IFERROR(D{tr4}/B{tr4},0)")
for r in range(r0, tr4 + 1):
    for col in range(1, 7):
        cell = ws4.cell(row=r, column=col); cell.border = BORDER
        if r == tr4: cell.fill = FILL_TOTAL; cell.font = BOLD
        elif col != 1: cell.font = BLACK
    ws4.cell(row=r, column=2).number_format = YEN
    ws4.cell(row=r, column=3).number_format = '#,##0.0'
    ws4.cell(row=r, column=4).number_format = YEN
    ws4.cell(row=r, column=5).number_format = YEN
    ws4.cell(row=r, column=6).number_format = XR

# ============================================================
# シート5：ダッシュボード
# ============================================================
ws5 = wb.create_sheet("ダッシュボード")
for col, w in zip("ABCD", [30, 20, 24, 20]): ws5.column_dimensions[col].width = w
ws5.merge_cells("A1:D1"); ws5["A1"] = "ダッシュボード ── 当月の着地と月商1億/10億"; ws5["A1"].font = TITLE

def kv(row, label, formula, fmt, note=""):
    ws5.cell(row=row, column=1, value=label).font = BOLD
    c = ws5.cell(row=row, column=2, value=formula); c.number_format = fmt; c.font = BLACK; c.border = BORDER
    ws5.cell(row=row, column=1).border = BORDER
    if note:
        ws5.cell(row=row, column=3, value=note).font = SUB

kv(3, "当月 予想売上（累計）", f"={MONTH_TOTAL_SALES}", YEN, "日別サマリーの対象年月に連動")
kv(4, "当月 広告費（累計）", f"={MONTH_TOTAL_ADV}", YEN)
kv(5, "当月 予想獲得数（累計）", f"={MONTH_TOTAL_CV}", '#,##0.0')
kv(6, "当月 ROAS", "=IFERROR(B3/B4,0)", XR, "＝売上 ÷ 広告費")
kv(7, "当月 ブレンドCPA", "=IFERROR(B4/B5,0)", YEN, "＝広告費 ÷ 獲得数")

ws5.cell(row=9, column=1, value="■ 着地見込み（日割りペース）").font = SECW
for col in (1,2,3,4): ws5.cell(row=9, column=col).fill = FILL_SEC
# 実績が入っている日数＝広告費>0 の日数
cntrng = f"'日別サマリー'!$B$5:$B$35"
kv(10, "データが入っている日数", f'=COUNTIF({cntrng},">0")', '#,##0"日"', "広告費が入力された日数")
kv(11, "当月の日数", "=DAY(DATE(YEAR('日別サマリー'!$B$2),MONTH('日別サマリー'!$B$2)+1,1)-1)", '#,##0"日"')
kv(12, "月末 着地見込み", "=IFERROR(B3/B10*B11,0)", YEN, "＝累計売上 ÷ 経過日数 × 当月日数")

ws5.cell(row=14, column=1, value="■ 月商目標への距離").font = SECW
for col in (1,2,3,4): ws5.cell(row=14, column=col).fill = FILL_SEC
kv(15, "月商1億 達成率（着地見込ベース）", "=IFERROR(B12/100000000,0)", '0.0%')
kv(16, "月商10億 達成率（着地見込ベース）", "=IFERROR(B12/1000000000,0)", '0.0%')
kv(17, "1億まであと（着地との差）", "=100000000-B12", YEN)

# ============================================================
# シート6：使い方
# ============================================================
ws6 = wb.create_sheet("使い方")
ws6.column_dimensions["A"].width = 100
lines = [
    ("Libetee 日次入力シート 使い方", TITLE), ("", None),
    ("【毎日やること】「日次入力」に、その日動いた分だけ行を足す（1行＝日付×商品×媒体）。", BOLD),
    ("  1. 日付を入れる", None),
    ("  2. アカウント・商品・媒体をドロップダウンから選ぶ（表記ゆれ防止＝集計が崩れない）", None),
    ("  3. 青字セル（黄色）を入力：広告費 / CPA / アクセス単価 / 予想転換率 / 客単価", None),
    ("  → 予想アクセス数・予想獲得数・予想売上・ROAS が自動で出る", None),
    ("", None),
    ("【計算式（社長指定）】", BOLD),
    ("  予想アクセス数 ＝ 広告費 ÷ アクセス単価", None),
    ("  予想獲得数(CPA基準) ＝ 広告費 ÷ CPA", None),
    ("  予想獲得数(アクセス基準) ＝ 予想アクセス数 × 予想転換率", None),
    ("  整合差異 ＝ 2つの獲得数の差（大きくズレたら CPA か 転換率の前提を見直す合図）", None),
    ("  予想売上 ＝ 予想獲得数(CPA基準) × 客単価 ／ ROAS ＝ 予想売上 ÷ 広告費", None),
    ("", None),
    ("【見るシート】", BOLD),
    ("  ・日別サマリー … 対象年月を入れると、1〜月末の日別トレンドと月累計が出る", None),
    ("  ・商品別サマリー / 媒体別サマリー … その月の合計を自動集計（対象年月に連動）", None),
    ("  ・ダッシュボード … 当月ROASと『日割りペースからの月末着地見込み』、月商1億/10億の達成率", None),
    ("", None),
    ("【ポイント】月の途中でも、ダッシュボードの『着地見込み』が今のペースでの月末売上を示します。", None),
    ("", None),
    ("色：青字＝入力 / 黒字＝自動計算 / 緑字＝他シート参照。数値はサンプルです。", SUB),
]
for i, (txt, fnt) in enumerate(lines, start=1):
    c = ws6.cell(row=i, column=1, value=txt); c.font = fnt if fnt else Font(name=FONT)
    c.alignment = Alignment(wrap_text=True, vertical="center")

out = "/tmp/claude-0/-home-user--/c7179f16-7eb7-52e8-aa4a-c082e5dbce5f/scratchpad/Libetee_広告予想シート_日次.xlsx"
wb.save(out)
print("saved:", out, "| data rows", DATA_START, "-", DATA_END, "| 日別合計行", tr)
