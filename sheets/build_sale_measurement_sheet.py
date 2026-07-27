#!/usr/bin/env python3
# Libetee 自社セール 効果測定・予想シート（媒体×アカウント×商品SKU）
import datetime as dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
YEN = '¥#,##0'; PCT = '0.00%'; NUM = '#,##0'; XR = '0.00"x"'; DATEF = 'yyyy/mm/dd'
BLUE = Font(name=FONT, color="0000FF"); BLACK = Font(name=FONT, color="000000")
GREEN = Font(name=FONT, color="008000"); HEADW = Font(name=FONT, color="FFFFFF", bold=True, size=10)
TITLE = Font(name=FONT, bold=True, size=14); SUB = Font(name=FONT, italic=True, size=9, color="555555")
BOLD = Font(name=FONT, bold=True); SECW = Font(name=FONT, bold=True, color="FFFFFF")
FILL_HEAD = PatternFill("solid", fgColor="1F4E78"); FILL_INPUT = PatternFill("solid", fgColor="FFF7CC")
FILL_TOTAL = PatternFill("solid", fgColor="D9E1F2"); FILL_SEC = PatternFill("solid", fgColor="2E75B6")
FILL_WARN = PatternFill("solid", fgColor="FFC7CE")
thin = Side(style="thin", color="BFBFBF"); BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True); LEFT = Alignment(horizontal="left", vertical="center")

DATE = dt.date(2026, 7, 25)
MEDIA = ["自社", "Amazon", "メタ広告"]
ACCTS = ["トラベル", "カタログ", "ガジェティ"]
GROUP = {"トラベル": "スーツケース", "カタログ": "スーツケース", "ガジェティ": "ハンディファン"}
SKU_SC = ["多機能PC", "ノーマルアルミ", "多機能アルミ", "セール訴求"]
SKU_FAN = ["首振り", "スケルトン", "ミニファン", "クリップ", "セール訴求"]
ALL_SKU = ["多機能PC", "ノーマルアルミ", "多機能アルミ", "首振り", "スケルトン", "ミニファン", "クリップ", "セール訴求"]

# (媒体, アカウント, 商品, 広告費[円])  ※いただいた7/25実データ（万円→円）
ROWS = [
    ("自社", "トラベル", "多機能PC", 20000), ("自社", "トラベル", "ノーマルアルミ", 10000), ("自社", "トラベル", "多機能アルミ", 5000),
    ("自社", "カタログ", "多機能PC", 10000), ("自社", "カタログ", "ノーマルアルミ", 5000), ("自社", "カタログ", "多機能アルミ", 5000),
    ("自社", "ガジェティ", "首振り", 6000), ("自社", "ガジェティ", "スケルトン", 6000), ("自社", "ガジェティ", "ミニファン", 3000), ("自社", "ガジェティ", "クリップ", 3000),
    ("Amazon", "トラベル", "多機能PC", 60000), ("Amazon", "トラベル", "多機能アルミ", 30000), ("Amazon", "トラベル", "ノーマルアルミ", 10000),
    ("Amazon", "カタログ", "多機能PC", 20000), ("Amazon", "カタログ", "ノーマルアルミ", 10000), ("Amazon", "カタログ", "多機能アルミ", 10000),
    ("Amazon", "ガジェティ", "首振り", 15000), ("Amazon", "ガジェティ", "スケルトン", 15000),
    ("メタ広告", "トラベル", "セール訴求", 50000), ("メタ広告", "トラベル", "多機能PC", 220000), ("メタ広告", "トラベル", "ノーマルアルミ", 30000), ("メタ広告", "トラベル", "多機能アルミ", 0),
    ("メタ広告", "カタログ", "多機能PC", 80000), ("メタ広告", "カタログ", "ノーマルアルミ", 20000), ("メタ広告", "カタログ", "多機能アルミ", 10000),
    ("メタ広告", "ガジェティ", "セール訴求", 30000), ("メタ広告", "ガジェティ", "首振り", 60000), ("メタ広告", "ガジェティ", "スケルトン", 60000), ("メタ広告", "ガジェティ", "ミニファン", 20000), ("メタ広告", "ガジェティ", "クリップ", 20000),
]
# ご申告のアカウント合計（差異チェック用・円）
STATED = {("自社","トラベル"):35000,("自社","カタログ"):30000,("自社","ガジェティ"):12000,
          ("Amazon","トラベル"):100000,("Amazon","カタログ"):40000,("Amazon","ガジェティ"):30000,
          ("メタ広告","トラベル"):320000,("メタ広告","カタログ"):110000,("メタ広告","ガジェティ"):200000}
# 推定アクセス数（媒体×商品群・下限/上限）※いただいた実データ
ACCESS = {("自社","スーツケース"):(5000,6000),("自社","ハンディファン"):(900,1200),
          ("Amazon","スーツケース"):(10000,15000),("Amazon","ハンディファン"):(3000,3000),
          ("メタ広告","スーツケース"):(32000,40000),("メタ広告","ハンディファン"):(15000,20000)}
# 転換率・客単価は【仮】。実績が分かり次第の差し替え前提。
CVR = {"スーツケース":0.010, "ハンディファン":0.020}
AOV = {"スーツケース":15000, "ハンディファン":3200}

wb = openpyxl.Workbook()

# ========== シート1：配信内訳（SKU別広告費） ==========
ws = wb.active; ws.title = "配信内訳(SKU)"
MAIN = "配信内訳(SKU)"
ws.merge_cells("A1:F1"); ws["A1"] = "配信内訳（媒体 × アカウント × 商品SKU）7/25 実データ"; ws["A1"].font = TITLE
ws.merge_cells("A2:F2"); ws["A2"] = "青字＝入力（黄色）。毎日ここに1行ずつ配信を記録。集計は各サマリーへ自動反映。"; ws["A2"].font = SUB
heads = [("日付",12),("媒体",12),("アカウント",14),("商品群",14),("商品(SKU)",18),("広告費",13)]
HR = 3
for j,(h,w) in enumerate(heads,1):
    c=ws.cell(HR,j,h); c.font=HEADW; c.fill=FILL_HEAD; c.alignment=CEN; c.border=BORDER
    ws.column_dimensions[get_column_letter(j)].width=w
DS=4; NR=300; DE=DS+NR-1
for i in range(NR):
    r=DS+i
    if i<len(ROWS):
        med,acc,sku,adv=ROWS[i]
        ws.cell(r,1,DATE).number_format=DATEF
        ws.cell(r,2,med); ws.cell(r,3,acc)
        ws.cell(r,4,f'=IFERROR(IF(C{r}="","",IF(C{r}="ガジェティ","ハンディファン","スーツケース")),"")')
        ws.cell(r,5,sku); ws.cell(r,6,adv)
    else:
        ws.cell(r,4,f'=IF(C{r}="","",IF(C{r}="ガジェティ","ハンディファン","スーツケース"))')
    for col in range(1,7):
        cell=ws.cell(r,col); cell.border=BORDER
        cell.font = BLUE if col in (1,2,3,5,6) else BLACK
        if col in (1,2,3,5,6): cell.fill=FILL_INPUT
    ws.cell(r,1).number_format=DATEF; ws.cell(r,6).number_format=YEN
    ws.cell(r,3).alignment=LEFT; ws.cell(r,5).alignment=LEFT
dv_med=DataValidation(type="list",formula1='"%s"'%",".join(MEDIA),allow_blank=True)
dv_acc=DataValidation(type="list",formula1='"%s"'%",".join(ACCTS),allow_blank=True)
dv_sku=DataValidation(type="list",formula1='"%s"'%",".join(ALL_SKU),allow_blank=True)
for dv,col in [(dv_med,"B"),(dv_acc,"C"),(dv_sku,"E")]:
    ws.add_data_validation(dv); dv.add(f"{col}{DS}:{col}{DE}")
ws.freeze_panes="A4"

def sif(sumcol, crits):
    base=f"'{MAIN}'!${sumcol}${DS}:${sumcol}${DE}"
    parts=[base]
    for cc,val in crits:
        parts.append(f"'{MAIN}'!${cc}${DS}:${cc}${DE}")
        parts.append(val)
    return "=SUMIFS("+",".join(parts)+")"

# ========== シート2：効果測定・予想（媒体×商品群） ==========
ws2=wb.create_sheet("効果測定・予想")
ws2.merge_cells("A1:O1"); ws2["A1"]="効果測定・予想（媒体 × 商品群）"; ws2["A1"].font=TITLE
ws2.merge_cells("A2:O2"); ws2["A2"]="広告費は配信内訳から自動集計。推定アクセスは実測入力。転換率・客単価は【仮】＝実績が分かり次第差し替え。"; ws2["A2"].font=SUB
h2=[("媒体",11),("商品群",13),("広告費",12),("推定ｱｸｾｽ\n下限",10),("推定ｱｸｾｽ\n上限",10),
    ("ｱｸｾｽ単価\n(最小)",10),("ｱｸｾｽ単価\n(最大)",10),("予想\n転換率",9),("客単価",11),
    ("予想獲得\n下限",10),("予想獲得\n上限",10),("予想売上\n(弱気)",13),("予想売上\n(本命)",13),("予想売上\n(強気)",13),("ROAS\n(本命)",9)]
HR2=3
for j,(h,w) in enumerate(h2,1):
    c=ws2.cell(HR2,j,h); c.font=HEADW; c.fill=FILL_HEAD; c.alignment=CEN; c.border=BORDER
    ws2.column_dimensions[get_column_letter(j)].width=w
ws2.row_dimensions[HR2].height=30
combos=[(m,g) for m in MEDIA for g in ["スーツケース","ハンディファン"]]
r0=HR2+1
for i,(m,g) in enumerate(combos):
    r=r0+i
    lo,hi=ACCESS[(m,g)]
    ws2.cell(r,1,m); ws2.cell(r,2,g).font=GREEN
    ws2.cell(r,3, sif("F",[("B",f'$A{r}'),("D",f'$B{r}')]))   # 広告費 SUMIFS(媒体,商品群)
    ws2.cell(r,4,lo); ws2.cell(r,5,hi)
    ws2.cell(r,6,f"=IFERROR(C{r}/E{r},0)")   # 単価(最小)=広告費/上限アクセス
    ws2.cell(r,7,f"=IFERROR(C{r}/D{r},0)")   # 単価(最大)=広告費/下限アクセス
    ws2.cell(r,8,CVR[g]); ws2.cell(r,9,AOV[g])
    ws2.cell(r,10,f"=D{r}*H{r}"); ws2.cell(r,11,f"=E{r}*H{r}")
    ws2.cell(r,12,f"=J{r}*I{r}"); ws2.cell(r,14,f"=K{r}*I{r}")
    ws2.cell(r,13,f"=(J{r}+K{r})/2*I{r}")
    ws2.cell(r,15,f"=IFERROR(M{r}/C{r},0)")
    for col in range(1,16):
        cell=ws2.cell(r,col); cell.border=BORDER
        cell.font = BLUE if col in (4,5,8,9) else BLACK
        if col in (4,5,8,9): cell.fill=FILL_INPUT
    ws2.cell(r,3).number_format=YEN
    for col in (4,5,10,11): ws2.cell(r,col).number_format=NUM
    for col in (6,7,9,12,13,14): ws2.cell(r,col).number_format=YEN
    ws2.cell(r,8).number_format=PCT; ws2.cell(r,15).number_format=XR
re=r0+len(combos)-1; tr=re+1
ws2.cell(tr,1,"合計").font=BOLD; ws2.merge_cells(start_row=tr,start_column=1,end_row=tr,end_column=2)
for col,L in [(3,"C"),(10,"J"),(11,"K"),(12,"L"),(13,"M"),(14,"N")]:
    ws2.cell(tr,col,f"=SUM({L}{r0}:{L}{re})")
ws2.cell(tr,15,f"=IFERROR(M{tr}/C{tr},0)")
for col in range(1,16):
    cell=ws2.cell(tr,col); cell.fill=FILL_TOTAL; cell.font=BOLD; cell.border=BORDER
ws2.cell(tr,3).number_format=YEN
for col in (10,11): ws2.cell(tr,col).number_format=NUM
for col in (12,13,14): ws2.cell(tr,col).number_format=YEN
ws2.cell(tr,15).number_format=XR
TOT2=tr

# ========== シート3：媒体・商品群サマリー ==========
ws3=wb.create_sheet("サマリー")
ws3.merge_cells("A1:E1"); ws3["A1"]="サマリー（媒体別・商品群別）"; ws3["A1"].font=TITLE
def block(startrow, title, keys, keycol):
    ws3.cell(startrow,1,title).font=SECW
    for c in range(1,6): ws3.cell(startrow,c).fill=FILL_SEC
    hh=[("区分",16),("広告費",14),("予想売上(本命)",16),("ROAS",10),("売上構成比",12)]
    for j,(h,w) in enumerate(hh,1):
        cc=ws3.cell(startrow+1,j,h); cc.font=HEADW; cc.fill=FILL_HEAD; cc.alignment=CEN; cc.border=BORDER
        ws3.column_dimensions[get_column_letter(j)].width=w
    rr=startrow+2
    for k in keys:
        ws3.cell(rr,1,k).font=GREEN
        ws3.cell(rr,2,f"=SUMIFS('効果測定・予想'!$C${r0}:$C${re},'効果測定・予想'!${keycol}${r0}:${keycol}${re},$A{rr})")
        ws3.cell(rr,3,f"=SUMIFS('効果測定・予想'!$M${r0}:$M${re},'効果測定・予想'!${keycol}${r0}:${keycol}${re},$A{rr})")
        ws3.cell(rr,4,f"=IFERROR(C{rr}/B{rr},0)")
        rr+=1
    te=rr-1; trow=rr
    ws3.cell(trow,1,"合計").font=BOLD
    ws3.cell(trow,2,f"=SUM(B{startrow+2}:B{te})"); ws3.cell(trow,3,f"=SUM(C{startrow+2}:C{te})")
    ws3.cell(trow,4,f"=IFERROR(C{trow}/B{trow},0)")
    for rr2 in range(startrow+2,trow+1):
        ws3.cell(rr2,5,f"=IFERROR(C{rr2}/C${trow},0)")
        for c in range(1,6):
            cell=ws3.cell(rr2,c); cell.border=BORDER
            if rr2==trow: cell.fill=FILL_TOTAL; cell.font=BOLD
            elif c!=1: cell.font=BLACK
        ws3.cell(rr2,2).number_format=YEN; ws3.cell(rr2,3).number_format=YEN
        ws3.cell(rr2,4).number_format=XR; ws3.cell(rr2,5).number_format='0.0%'
    return trow
end1=block(3,"■ 媒体別（自社 / Amazon / メタ広告）",MEDIA,"A")
block(end1+2,"■ 商品群別（スーツケース / ハンディファン）",["スーツケース","ハンディファン"],"B")

# ========== シート4：差異チェック（SKU内訳 vs ご申告） ==========
ws4=wb.create_sheet("差異チェック")
ws4.merge_cells("A1:E1"); ws4["A1"]="差異チェック（SKU内訳の合計 vs ご申告のアカウント合計）"; ws4["A1"].font=TITLE
ws4.merge_cells("A2:E2"); ws4["A2"]="差異が0以外＝内訳と申告合計が不一致。赤セルを潰す（SKU修正 or 未内訳行の追加）と整います。"; ws4["A2"].font=SUB
hh=[("媒体",12),("アカウント",14),("SKU内訳合計",16),("ご申告合計",16),("差異",14)]
for j,(h,w) in enumerate(hh,1):
    c=ws4.cell(3,j,h); c.font=HEADW; c.fill=FILL_HEAD; c.alignment=CEN; c.border=BORDER
    ws4.column_dimensions[get_column_letter(j)].width=w
rr=4
for (m,a),st in STATED.items():
    ws4.cell(rr,1,m); ws4.cell(rr,2,a)
    ws4.cell(rr,3,sif("F",[("B",f'$A{rr}'),("C",f'$B{rr}')]))
    ws4.cell(rr,4,st).font=BLUE; ws4.cell(rr,4).fill=FILL_INPUT
    ws4.cell(rr,5,f"=D{rr}-C{rr}")
    for c in range(1,6):
        cell=ws4.cell(rr,c); cell.border=BORDER
        if c==5: cell.font=BLACK
    ws4.cell(rr,3).number_format=YEN; ws4.cell(rr,4).number_format=YEN; ws4.cell(rr,5).number_format='¥#,##0;[RED]-¥#,##0'
    rr+=1
# 条件付き書式：差異≠0 を赤
from openpyxl.formatting.rule import CellIsRule
ws4.conditional_formatting.add(f"E4:E{rr-1}", CellIsRule(operator="notEqual", formula=["0"], fill=FILL_WARN))

# ========== シート5：使い方 ==========
ws5=wb.create_sheet("使い方"); ws5.column_dimensions["A"].width=100
lines=[
 ("Libetee 自社セール 効果測定・予想シート 使い方",TITLE),("",None),
 ("【構造】媒体（自社/Amazon/メタ広告）× アカウント（トラベル/カタログ/ガジェティ）× 商品SKU。",BOLD),
 ("  ・トラベル と カタログ ＝ スーツケース群 ／ ガジェティ ＝ ハンディファン群（商品群は自動判定）",None),
 ("",None),
 ("【毎日やること】「配信内訳(SKU)」に、その日の配信を1行ずつ記録（媒体・アカウント・SKUはドロップダウン、広告費を入力）。",BOLD),
 ("",None),
 ("【効果測定・予想シート】",BOLD),
 ("  ・広告費 … 配信内訳から自動集計",None),
 ("  ・推定アクセス数（下限/上限）… 実測を入力",None),
 ("  ・アクセス単価 ＝ 広告費 ÷ アクセス数（最小＝広告費÷上限、最大＝広告費÷下限）",None),
 ("  ・予想獲得数 ＝ アクセス数 × 予想転換率",None),
 ("  ・予想売上 ＝ 獲得数 × 客単価（弱気/本命/強気の3本）",None),
 ("  ※ 予想転換率・客単価は現在【仮】。実績が分かり次第、差し替えると予想が本物になります。",SUB),
 ("",None),
 ("【差異チェック】SKU内訳の合計と、ご申告のアカウント合計を突き合わせ。赤＝不一致。",BOLD),
 ("  現時点の不一致：自社カタログ・自社ガジェティ・メタトラベル・メタガジェティ の4か所。",None),
 ("",None),
 ("色：青字＝入力 / 黒字＝自動計算 / 緑字＝他シート参照。広告費は7/25の実データ。",SUB),
]
for i,(t,f) in enumerate(lines,1):
    c=ws5.cell(i,1,t); c.font=f if f else Font(name=FONT); c.alignment=Alignment(wrap_text=True,vertical="center")

out="/tmp/claude-0/-home-user--/c7179f16-7eb7-52e8-aa4a-c082e5dbce5f/scratchpad/Libetee_効果測定_予想シート.xlsx"
wb.save(out); print("saved:",out,"| SKU rows",DS,"-",DE,"| 予想合計行",TOT2)
