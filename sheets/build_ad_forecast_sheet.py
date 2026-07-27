#!/usr/bin/env python3
# Libetee 広告費→予想 業務効率化シート
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
YEN = '¥#,##0'
YEN2 = '¥#,##0.0'
PCT = '0.00%'
NUM = '#,##0'
XR = '0.00"x"'

# 色
BLUE = Font(name=FONT, color="0000FF")          # 入力
BLACK = Font(name=FONT, color="000000")         # 計算
GREEN = Font(name=FONT, color="008000")         # 他シート参照
HEADW = Font(name=FONT, color="FFFFFF", bold=True, size=10)
TITLE = Font(name=FONT, bold=True, size=14)
SUB = Font(name=FONT, italic=True, size=9, color="555555")
BOLD = Font(name=FONT, bold=True)

FILL_HEAD = PatternFill("solid", fgColor="1F4E78")
FILL_INPUT = PatternFill("solid", fgColor="FFF7CC")   # 入力セル=薄い黄
FILL_TOTAL = PatternFill("solid", fgColor="D9E1F2")
FILL_ACC = PatternFill("solid", fgColor="EDEDED")
FILL_SEC = PatternFill("solid", fgColor="2E75B6")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

wb = openpyxl.Workbook()

# ============================================================
# 行データ（アカウント / 商品 / 媒体） + サンプル入力値
# ============================================================
MEDIA = ["Amazon", "楽天", "自社サイト"]
# (アカウント, 商品/ブランド)
PRODUCTS = [
    ("メタアカウント", "スーツケース（トラベル）"),
    ("メタアカウント", "スーツケース（カタログ）"),
    ("ハンディファンアカウント", "ガジェティ"),
    ("ツヤリスアカウント", "ツヤリス"),
]
# サンプル入力（広告費, CPA, アクセス単価, 予想転換率, 客単価）媒体別
SAMPLE = {
    "スーツケース（トラベル）": {
        "Amazon":   (300000, 6000, 60, 0.010, 15000),
        "楽天":     (250000, 8000, 70, 0.006, 15000),
        "自社サイト": (150000, 7000, 55, 0.012, 16000),
    },
    "スーツケース（カタログ）": {
        "Amazon":   (200000, 6500, 65, 0.009, 14000),
        "楽天":     (180000, 8500, 75, 0.005, 14000),
        "自社サイト": (120000, 7200, 58, 0.011, 15000),
    },
    "ガジェティ": {
        "Amazon":   (180000, 1500, 40, 0.020, 3200),
        "楽天":     (150000, 1800, 45, 0.015, 3200),
        "自社サイト": (90000, 1600, 38, 0.022, 3500),
    },
    "ツヤリス": {
        "Amazon":   (160000, 2200, 45, 0.018, 4200),
        "楽天":     (140000, 2600, 50, 0.012, 4200),
        "自社サイト": (100000, 2400, 42, 0.020, 4500),
    },
}

# ============================================================
# シート1：入力＆予想（メイン）
# ============================================================
ws = wb.active
ws.title = "入力＆予想"

headers = [
    ("アカウント", 20), ("商品/ブランド", 22), ("媒体", 12),
    ("広告費", 13), ("CPA\n(獲得単価)", 12), ("アクセス単価", 12),
    ("予想\n転換率", 10), ("客単価", 12),
    ("予想\nアクセス数", 13), ("予想獲得数\n(CPA基準)", 13),
    ("予想獲得数\n(ｱｸｾｽ×CVR)", 14), ("整合差異", 11),
    ("予想売上", 15), ("ROAS", 9),
]
# タイトル
ws.merge_cells("A1:N1")
ws["A1"] = "Libetee 広告費 → 売上予想シート（メタ広告アカウント × 媒体）"
ws["A1"].font = TITLE
ws.merge_cells("A2:N2")
ws["A2"] = "青字＝入力セル（黄色網掛け）。数値はサンプルです。実績・目標値に置き換えると、右側と各サマリーが自動計算されます。"
ws["A2"].font = SUB

HROW = 3
for j, (h, w) in enumerate(headers, start=1):
    c = ws.cell(row=HROW, column=j, value=h)
    c.font = HEADW; c.fill = FILL_HEAD; c.alignment = CEN; c.border = BORDER
    ws.column_dimensions[get_column_letter(j)].width = w
ws.row_dimensions[HROW].height = 32

r = HROW + 1
data_start = r
input_cols = {"D", "E", "F", "G", "H"}
for acc, prod in PRODUCTS:
    for m in MEDIA:
        adv, cpa, cpc, cvr, aov = SAMPLE[prod][m]
        ws.cell(row=r, column=1, value=acc)
        ws.cell(row=r, column=2, value=prod)
        ws.cell(row=r, column=3, value=m)
        ws.cell(row=r, column=4, value=adv)
        ws.cell(row=r, column=5, value=cpa)
        ws.cell(row=r, column=6, value=cpc)
        ws.cell(row=r, column=7, value=cvr)
        ws.cell(row=r, column=8, value=aov)
        # 計算列
        ws.cell(row=r, column=9,  value=f"=IFERROR(D{r}/F{r},0)")            # 予想アクセス数 = 広告費/アクセス単価
        ws.cell(row=r, column=10, value=f"=IFERROR(D{r}/E{r},0)")            # 予想獲得数(CPA) = 広告費/CPA
        ws.cell(row=r, column=11, value=f"=I{r}*G{r}")                       # 予想獲得数(アクセス×CVR)
        ws.cell(row=r, column=12, value=f"=J{r}-K{r}")                       # 整合差異
        ws.cell(row=r, column=13, value=f"=J{r}*H{r}")                       # 予想売上 = 獲得(CPA)×客単価
        ws.cell(row=r, column=14, value=f"=IFERROR(M{r}/D{r},0)")           # ROAS = 予想売上/広告費
        r += 1
data_end = r - 1

# 合計行
ws.cell(row=r, column=1, value="合計").font = BOLD
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
for col in (4, 9, 10, 11, 13):
    L = get_column_letter(col)
    ws.cell(row=r, column=col, value=f"=SUM({L}{data_start}:{L}{data_end})")
ws.cell(row=r, column=14, value=f"=IFERROR(M{r}/D{r},0)")
total_row = r

# 書式
for rr in range(data_start, data_end + 1):
    for col in range(1, 15):
        cell = ws.cell(row=rr, column=col)
        cell.border = BORDER
        L = get_column_letter(col)
        if L in input_cols:
            cell.font = BLUE; cell.fill = FILL_INPUT
        else:
            cell.font = BLACK
        if col == 1:
            cell.fill = FILL_ACC
    ws.cell(row=rr, column=1).alignment = LEFT
    ws.cell(row=rr, column=2).alignment = LEFT
    ws.cell(row=rr, column=3).alignment = CEN
    for col in (4, 5, 6, 8): ws.cell(row=rr, column=col).number_format = YEN
    ws.cell(row=rr, column=7).number_format = PCT
    ws.cell(row=rr, column=9).number_format = NUM
    ws.cell(row=rr, column=10).number_format = '#,##0.0'
    ws.cell(row=rr, column=11).number_format = '#,##0.0'
    ws.cell(row=rr, column=12).number_format = '#,##0.0;[RED]-#,##0.0'
    ws.cell(row=rr, column=13).number_format = YEN
    ws.cell(row=rr, column=14).number_format = XR
# 合計行書式
for col in range(1, 15):
    cell = ws.cell(row=total_row, column=col)
    cell.fill = FILL_TOTAL; cell.border = BORDER; cell.font = BOLD
ws.cell(row=total_row, column=4).number_format = YEN
ws.cell(row=total_row, column=9).number_format = NUM
ws.cell(row=total_row, column=10).number_format = '#,##0.0'
ws.cell(row=total_row, column=11).number_format = '#,##0.0'
ws.cell(row=total_row, column=13).number_format = YEN
ws.cell(row=total_row, column=14).number_format = XR
ws.freeze_panes = "D4"

MAIN = "入力＆予想"

# ============================================================
# シート2：商品別サマリー（各商品の合計値を可視化）
# ============================================================
ws2 = wb.create_sheet("商品別サマリー")
ws2.merge_cells("A1:H1")
ws2["A1"] = "商品別サマリー（メイン入力から自動集計）"
ws2["A1"].font = TITLE
sh2 = [("商品/ブランド", 24), ("広告費", 14), ("予想アクセス数", 14),
       ("予想獲得数", 13), ("予想売上", 16), ("ブレンドCPA", 13),
       ("ROAS", 10), ("売上構成比", 12)]
for j, (h, w) in enumerate(sh2, start=1):
    c = ws2.cell(row=2, column=j, value=h)
    c.font = HEADW; c.fill = FILL_HEAD; c.alignment = CEN; c.border = BORDER
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.row_dimensions[2].height = 24

prod_names = [p for _, p in PRODUCTS]
rr = 3
prod_start = rr
for p in prod_names:
    ws2.cell(row=rr, column=1, value=p).font = GREEN
    rng = lambda L: f"'{MAIN}'!{L}{data_start}:{L}{data_end}"
    crit = f"'{MAIN}'!$B${data_start}:$B${data_end}"
    ws2.cell(row=rr, column=2, value=f'=SUMIFS({rng("D")},{crit},$A{rr})')
    ws2.cell(row=rr, column=3, value=f'=SUMIFS({rng("I")},{crit},$A{rr})')
    ws2.cell(row=rr, column=4, value=f'=SUMIFS({rng("J")},{crit},$A{rr})')
    ws2.cell(row=rr, column=5, value=f'=SUMIFS({rng("M")},{crit},$A{rr})')
    ws2.cell(row=rr, column=6, value=f'=IFERROR(B{rr}/D{rr}*0+B{rr}/D{rr},0)')  # placeholder replaced below
    ws2.cell(row=rr, column=6, value=f'=IFERROR(B{rr}/D{rr},0)')  # ブレンドCPA=広告費/獲得
    ws2.cell(row=rr, column=6, value=f'=IFERROR(B{rr}/D{rr},0)')
    rr += 1
prod_end = rr - 1
# ブレンドCPA修正（広告費/獲得数）と ROAS/構成比
for i, rrow in enumerate(range(prod_start, prod_end + 1)):
    ws2.cell(row=rrow, column=6, value=f'=IFERROR(B{rrow}/D{rrow},0)')   # 広告費/獲得
    ws2.cell(row=rrow, column=7, value=f'=IFERROR(E{rrow}/B{rrow},0)')   # ROAS=売上/広告費
    ws2.cell(row=rrow, column=8, value=f'=IFERROR(E{rrow}/E${prod_end+1},0)')  # 構成比
# 合計行
trow = prod_end + 1
ws2.cell(row=trow, column=1, value="合計").font = BOLD
for col, L in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    ws2.cell(row=trow, column=col, value=f"=SUM({L}{prod_start}:{L}{prod_end})")
ws2.cell(row=trow, column=6, value=f'=IFERROR(B{trow}/D{trow},0)')
ws2.cell(row=trow, column=7, value=f'=IFERROR(E{trow}/B{trow},0)')
ws2.cell(row=trow, column=8, value=1)
# 書式
for rrow in range(prod_start, trow + 1):
    for col in range(1, 9):
        cell = ws2.cell(row=rrow, column=col); cell.border = BORDER
        if rrow == trow: cell.fill = FILL_TOTAL; cell.font = BOLD
        elif col != 1: cell.font = BLACK
    ws2.cell(row=rrow, column=2).number_format = YEN
    ws2.cell(row=rrow, column=3).number_format = NUM
    ws2.cell(row=rrow, column=4).number_format = '#,##0.0'
    ws2.cell(row=rrow, column=5).number_format = YEN
    ws2.cell(row=rrow, column=6).number_format = YEN
    ws2.cell(row=rrow, column=7).number_format = XR
    ws2.cell(row=rrow, column=8).number_format = '0.0%'

# ============================================================
# シート3：媒体別サマリー（Amazon/楽天/自社サイト）
# ============================================================
ws3 = wb.create_sheet("媒体別サマリー")
ws3.merge_cells("A1:G1")
ws3["A1"] = "媒体別サマリー（振り分け媒体：Amazon / 楽天 / 自社サイト）"
ws3["A1"].font = TITLE
sh3 = [("媒体", 16), ("広告費", 14), ("予想アクセス数", 14),
       ("予想獲得数", 13), ("予想売上", 16), ("ROAS", 10), ("売上構成比", 12)]
for j, (h, w) in enumerate(sh3, start=1):
    c = ws3.cell(row=2, column=j, value=h)
    c.font = HEADW; c.fill = FILL_HEAD; c.alignment = CEN; c.border = BORDER
    ws3.column_dimensions[get_column_letter(j)].width = w
ws3.row_dimensions[2].height = 24
rr = 3
med_start = rr
for m in MEDIA:
    rng = lambda L: f"'{MAIN}'!{L}{data_start}:{L}{data_end}"
    crit = f"'{MAIN}'!$C${data_start}:$C${data_end}"
    ws3.cell(row=rr, column=1, value=m).font = GREEN
    ws3.cell(row=rr, column=2, value=f'=SUMIFS({rng("D")},{crit},$A{rr})')
    ws3.cell(row=rr, column=3, value=f'=SUMIFS({rng("I")},{crit},$A{rr})')
    ws3.cell(row=rr, column=4, value=f'=SUMIFS({rng("J")},{crit},$A{rr})')
    ws3.cell(row=rr, column=5, value=f'=SUMIFS({rng("M")},{crit},$A{rr})')
    rr += 1
med_end = rr - 1
trow3 = med_end + 1
for rrow in range(med_start, med_end + 1):
    ws3.cell(row=rrow, column=6, value=f'=IFERROR(E{rrow}/B{rrow},0)')
    ws3.cell(row=rrow, column=7, value=f'=IFERROR(E{rrow}/E${trow3},0)')
ws3.cell(row=trow3, column=1, value="合計").font = BOLD
for col, L in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
    ws3.cell(row=trow3, column=col, value=f"=SUM({L}{med_start}:{L}{med_end})")
ws3.cell(row=trow3, column=6, value=f'=IFERROR(E{trow3}/B{trow3},0)')
ws3.cell(row=trow3, column=7, value=1)
for rrow in range(med_start, trow3 + 1):
    for col in range(1, 8):
        cell = ws3.cell(row=rrow, column=col); cell.border = BORDER
        if rrow == trow3: cell.fill = FILL_TOTAL; cell.font = BOLD
        elif col != 1: cell.font = BLACK
    ws3.cell(row=rrow, column=2).number_format = YEN
    ws3.cell(row=rrow, column=3).number_format = NUM
    ws3.cell(row=rrow, column=4).number_format = '#,##0.0'
    ws3.cell(row=rrow, column=5).number_format = YEN
    ws3.cell(row=rrow, column=6).number_format = XR
    ws3.cell(row=rrow, column=7).number_format = '0.0%'

# ============================================================
# シート4：ダッシュボード
# ============================================================
ws4 = wb.create_sheet("ダッシュボード")
for w, col in zip([26, 20, 20, 20], "ABCD"): ws4.column_dimensions[col].width = w
ws4.merge_cells("A1:D1")
ws4["A1"] = "ダッシュボード ── 月商1億 / 10億への距離"
ws4["A1"].font = TITLE
tot = f"'{MAIN}'!"

def kv(row, label, formula, fmt, note=""):
    ws4.cell(row=row, column=1, value=label).font = BOLD
    c = ws4.cell(row=row, column=2, value=formula); c.number_format = fmt; c.font = BLACK
    if note:
        ws4.cell(row=row, column=3, value=note).font = SUB

kv(3, "予想売上（この予想の合計）", f"={tot}M{total_row}", YEN)
kv(4, "広告費（合計）", f"={tot}D{total_row}", YEN)
kv(5, "予想獲得数（合計）", f"={tot}J{total_row}", '#,##0.0')
kv(6, "全体ROAS", f"=IFERROR(B3/B4,0)", XR, "＝予想売上 ÷ 広告費")
kv(7, "全体ブレンドCPA", f"=IFERROR(B4/B5,0)", YEN, "＝広告費 ÷ 獲得数")

ws4.cell(row=9, column=1, value="■ 月商換算と目標対比").font = Font(name=FONT, bold=True, color="FFFFFF")
ws4.cell(row=9, column=1).fill = FILL_SEC
for col in (2,3,4): ws4.cell(row=9, column=col).fill = FILL_SEC
ws4.cell(row=10, column=1, value="この予想の対象日数（入力）").font = BLUE
d = ws4.cell(row=10, column=2, value=1); d.font = BLUE; d.fill = FILL_INPUT; d.number_format = '#,##0'
ws4.cell(row=10, column=3, value="日次予想なら1、週次なら7 など").font = SUB
kv(11, "月商換算（30日ベース）", "=IFERROR(B3/B10*30,0)", YEN, "＝予想売上 ÷ 対象日数 × 30")
kv(12, "月商1億まで（達成率）", "=IFERROR(B11/100000000,0)", '0.0%')
kv(13, "月商10億まで（達成率）", "=IFERROR(B11/1000000000,0)", '0.0%')
for rrow in range(3, 14):
    for col in (1,2,3):
        ws4.cell(row=rrow, column=col).border = BORDER

# ============================================================
# シート5：使い方
# ============================================================
ws5 = wb.create_sheet("使い方")
ws5.column_dimensions["A"].width = 100
lines = [
    ("Libetee 広告費→予想シート 使い方", TITLE),
    ("", None),
    ("【毎回やること】メインシート「入力＆予想」の青字セル（黄色網掛け）だけを埋める。", BOLD),
    ("  ・広告費 … その媒体・その商品に投下する広告費", None),
    ("  ・CPA（獲得単価）… 1件獲得あたりの広告費（実績 or 目標）", None),
    ("  ・アクセス単価 … 1アクセスあたりの広告費（クリック単価に相当）", None),
    ("  ・予想転換率 … アクセスのうち購入に至る割合（%）", None),
    ("  ・客単価 … 1件あたりの平均売上", None),
    ("", None),
    ("【自動で出る数字（社長指定の計算式）】", BOLD),
    ("  ・予想アクセス数 ＝ 広告費 ÷ アクセス単価", None),
    ("  ・予想獲得数(CPA基準) ＝ 広告費 ÷ CPA", None),
    ("  ・予想獲得数(アクセス基準) ＝ 予想アクセス数 × 予想転換率", None),
    ("  ・整合差異 ＝ 上の2つの獲得数の差（大きくズレたら CPA か 転換率の前提を見直す合図）", None),
    ("  ・予想売上 ＝ 予想獲得数(CPA基準) × 客単価", None),
    ("  ・ROAS ＝ 予想売上 ÷ 広告費（＝客単価 ÷ CPA）", None),
    ("", None),
    ("【見るシート】", BOLD),
    ("  ・商品別サマリー … スーツケース(トラベル/カタログ)・ガジェティ・ツヤリスの合計を自動集計", None),
    ("  ・媒体別サマリー … Amazon / 楽天 / 自社サイト の合計を自動集計", None),
    ("  ・ダッシュボード … 全体ROASと、月商1億/10億への達成率", None),
    ("", None),
    ("【色の意味】 青字＝入力セル / 黒字＝自動計算 / 緑字＝他シート参照", SUB),
    ("※ 初期の数値はサンプルです。実績・目標に置き換えてお使いください。", SUB),
]
for i, (txt, fnt) in enumerate(lines, start=1):
    c = ws5.cell(row=i, column=1, value=txt)
    if fnt: c.font = fnt
    else: c.font = Font(name=FONT)
    c.alignment = Alignment(wrap_text=True, vertical="center")

out = "/tmp/claude-0/-home-user--/c7179f16-7eb7-52e8-aa4a-c082e5dbce5f/scratchpad/Libetee_広告予想シート.xlsx"
wb.save(out)
print("saved:", out)
print("data rows:", data_start, "-", data_end, "total row:", total_row)
