# -*- coding: utf-8 -*-
"""
Meta広告 クリエイター配信管理シート 生成スクリプト

案件依頼の運用ルール（1〜4）を、現場で「見て・選んで・チェックできる」形にした
Excel管理シートを生成します。セレクトボックス（ドロップダウン）と色分けを多用し、
入力ミス・言い間違い・報告漏れを防ぐことを狙いとしています。

実行:  python generate.py
出力:  Meta広告クリエイター管理.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

FONT = "Arial"
MAX_ROW = 200          # 入力可能な行数（ヘッダー除く）

# ---- 色 -------------------------------------------------------------------
NAVY      = "1F3864"   # 見出し帯
BLUE_HEAD = "2E75B6"   # 表ヘッダー
LIGHT     = "DDEBF7"   # 見本行の薄い青
GREEN     = "C6EFCE"   # 利用可 / 済
GREEN_TXT = "006100"
RED       = "FFC7CE"   # 利用不可 / NG
RED_TXT   = "9C0006"
YELLOW    = "FFEB9C"   # 確認中 / 要注意
YELLOW_TXT= "9C6500"
GREY      = "F2F2F2"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_cell(c):
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor=BLUE_HEAD)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


# ===========================================================================
#  ワークブック
# ===========================================================================
wb = Workbook()

# ---------------------------------------------------------------------------
#  シート1: クリエイター配信管理（メイン）
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "クリエイター配信管理"
ws.sheet_view.showGridLines = False

# --- タイトル帯 ---
ws.merge_cells("A1:L1")
t = ws["A1"]
t.value = "Meta広告  クリエイター配信管理シート"
t.font = Font(name=FONT, bold=True, size=16, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=NAVY)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:L2")
sub = ws["A2"]
sub.value = ("※「利用可／利用不可」は必ずこのシートで管理（トラブル防止）　"
             "／　投稿後3日以内に配信可否を報告　／　事前確認は「二次利用」と言わない")
sub.font = Font(name=FONT, size=9, color="1F3864", italic=True)
sub.fill = PatternFill("solid", fgColor="EAF1FB")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# --- ヘッダー行（4行目） ---
HEADER_ROW = 4
headers = [
    ("No.", 5),
    ("名前", 16),
    ("アカウント名", 20),
    ("タイアップ投稿", 14),
    ("事前確認\n(Meta広告で配信OKか)", 20),
    ("Meta広告\n利用可否", 13),
    ("投稿日", 13),
    ("配信可否\n報告期限\n(投稿+3日)", 14),
    ("配信可否\n報告済", 11),
    ("配信予定日", 13),
    ("配信済み", 11),
    ("備考", 30),
]
for i, (name, width) in enumerate(headers, start=1):
    col = get_column_letter(i)
    ws.column_dimensions[col].width = width
    cell = ws.cell(row=HEADER_ROW, column=i, value=name)
    style_header_cell(cell)
ws.row_dimensions[HEADER_ROW].height = 46

# 列インデックス（可読性のため）
C_NO, C_NAME, C_ACC, C_TIEUP, C_PRECHK, C_USABLE, C_POST, \
    C_DEADLINE, C_REPORTED, C_PLAN, C_DONE, C_NOTE = range(1, 13)

FIRST = HEADER_ROW + 1                 # 5
LAST  = HEADER_ROW + MAX_ROW           # 204

# --- 見本行（5行目） ---
sample = {
    C_NO: 1,
    C_NAME: "山田 花子",
    C_ACC: "@hanako_life",
    C_TIEUP: "投稿済",
    C_PRECHK: "OK",
    C_USABLE: "利用可",
    C_POST: "2026-08-01",
    C_REPORTED: "済",
    C_PLAN: "2026-08-10",
    C_DONE: "未",
    C_NOTE: "（見本行）事前確認OK・報告済み。配信待ち。",
}
for col, val in sample.items():
    ws.cell(row=FIRST, column=col, value=val)

# --- データ行の書式・連番・数式 ---
for r in range(FIRST, LAST + 1):
    for c in range(1, 13):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.font = Font(name=FONT, size=10)
        if c in (C_NO, C_TIEUP, C_PRECHK, C_USABLE, C_POST,
                 C_DEADLINE, C_REPORTED, C_PLAN, C_DONE):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # 連番（名前が入っている行だけ番号表示）
    ws.cell(row=r, column=C_NO).value = f'=IF($B{r}="","",ROW()-{HEADER_ROW})'
    # 報告期限 = 投稿日 + 3日（投稿日が入っていれば自動計算）
    ws.cell(row=r, column=C_DEADLINE).value = f'=IF($G{r}="","",$G{r}+3)'
    # 日付書式
    for c in (C_POST, C_DEADLINE, C_PLAN):
        ws.cell(row=r, column=c).number_format = "yyyy/mm/dd"

# 見本行を薄く着色して「例」と分かるように
for c in range(1, 13):
    ws.cell(row=FIRST, column=c).fill = PatternFill("solid", fgColor=LIGHT)

# --- ドロップダウン（セレクトボックス）---
def add_dv(formula, col, allow_blank=True):
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    dv.error = "リストから選んでください"
    dv.errorTitle = "入力できます値が違います"
    dv.prompt = "▼ から選択"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    letter = get_column_letter(col)
    dv.add(f"{letter}{FIRST}:{letter}{LAST}")

add_dv('"予定,依頼済,投稿済,見送り"', C_TIEUP)
add_dv('"未確認,確認中,OK,NG"',       C_PRECHK)
add_dv('"利用可,利用不可,確認中"',     C_USABLE)
add_dv('"済,未"',                     C_REPORTED)
add_dv('"済,未"',                     C_DONE)

# --- 条件付き書式（色分けで一目チェック）---
def paint(col, value, fill, txt):
    letter = get_column_letter(col)
    rng = f"{letter}{FIRST}:{letter}{LAST}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=[f'"{value}"'],
                   fill=PatternFill("solid", fgColor=fill),
                   font=Font(name=FONT, size=10, bold=True, color=txt)))

# Meta広告 利用可否：可＝緑 / 不可＝赤 / 確認中＝黄
paint(C_USABLE, "利用可",   GREEN, GREEN_TXT)
paint(C_USABLE, "利用不可", RED,   RED_TXT)
paint(C_USABLE, "確認中",   YELLOW, YELLOW_TXT)
# 事前確認
paint(C_PRECHK, "OK", GREEN, GREEN_TXT)
paint(C_PRECHK, "NG", RED,   RED_TXT)
paint(C_PRECHK, "確認中", YELLOW, YELLOW_TXT)
# 報告済／配信済み：済＝緑 / 未＝黄
for col in (C_REPORTED, C_DONE):
    paint(col, "済", GREEN, GREEN_TXT)
    paint(col, "未", YELLOW, YELLOW_TXT)
# タイアップ投稿
paint(C_TIEUP, "投稿済", GREEN, GREEN_TXT)
paint(C_TIEUP, "見送り", RED,   RED_TXT)

# 報告期限の「超過アラート」：報告未 かつ 期限を過ぎている行を赤く
letter = get_column_letter(C_DEADLINE)
ws.conditional_formatting.add(
    f"{letter}{FIRST}:{letter}{LAST}",
    CellIsRule(operator="lessThan", formula=["TODAY()"],
               fill=PatternFill("solid", fgColor=RED),
               font=Font(name=FONT, size=10, bold=True, color=RED_TXT)))
# ↑ 期限セルは「未報告なら残る」運用。報告済にしたら備考で消込む想定。

# ヘッダーで固定＆オートフィルタ
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A{HEADER_ROW}:L{LAST}"

# コメント（列の意味）
ws.cell(row=HEADER_ROW, column=C_PRECHK).comment = Comment(
    '依頼時の事前確認。文言は「投稿いただいた動画を弊社のMeta広告で配信しても'
    '問題ないでしょうか？」で確認する。※「二次利用」とは言わない。', "運用ルール")
ws.cell(row=HEADER_ROW, column=C_USABLE).comment = Comment(
    'Meta広告で使える／使えないを厳重管理する欄（トラブル防止）。'
    '事前確認OKのクリエイターだけ「利用可」にする。', "運用ルール")
ws.cell(row=HEADER_ROW, column=C_DEADLINE).comment = Comment(
    '投稿日＋3日。この期限までに配信可否を報告する。'
    '期限を過ぎると赤くなります（報告したら「配信可否 報告済」を「済」に）。', "運用ルール")


# ---------------------------------------------------------------------------
#  シート2: 運用ルール・凡例
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("運用ルール・凡例")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 80

ws2.merge_cells("B1:C1")
h = ws2["B1"]
h.value = "案件依頼 運用ルール ／ このシートの使い方"
h.font = Font(name=FONT, bold=True, size=15, color="FFFFFF")
h.fill = PatternFill("solid", fgColor=NAVY)
h.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

rules = [
    ("① タイアップ投稿", "案件依頼では タイアップ投稿を必須 とする。"),
    ("② 事前確認の言い方",
     "「二次利用」という表現は使わない。"
     "「投稿いただいた動画を弊社のMeta広告で配信しても問題ないでしょうか？」と確認する。"
     "→「二次利用」と言うと、HPや各種媒体への掲載まで含むと受け取られる恐れがあるため。"),
    ("③ 利用可否の管理",
     "Meta広告で利用可能／利用不可のクリエイターは、このシートで厳重に管理する（トラブル防止）。"),
    ("④ 配信可否の報告",
     "Meta広告で利用できる動画は、投稿後3日以内に配信可否を報告する。"
     "（シートの「配信可否 報告期限」に自動計算。期限超過は赤表示）"),
]
r = 3
for title, body in rules:
    tc = ws2.cell(row=r, column=2, value=title)
    tc.font = Font(name=FONT, bold=True, size=11, color=NAVY)
    tc.alignment = Alignment(vertical="top", wrap_text=True)
    tc.fill = PatternFill("solid", fgColor=GREY)
    bc = ws2.cell(row=r, column=3, value=body)
    bc.font = Font(name=FONT, size=10)
    bc.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.row_dimensions[r].height = 46
    r += 1

# --- 凡例（色と選択肢の意味） ---
r += 1
lh = ws2.cell(row=r, column=2, value="凡例（色・選択肢の意味）")
lh.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
lh.fill = PatternFill("solid", fgColor=BLUE_HEAD)
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
lh.alignment = Alignment(horizontal="center", vertical="center")
r += 1

legend = [
    ("利用可 / OK / 済", GREEN, GREEN_TXT, "問題なし。Meta広告で配信できる状態。"),
    ("利用不可 / NG / 見送り", RED, RED_TXT, "配信できない。広告に使わない。"),
    ("確認中 / 未", YELLOW, YELLOW_TXT, "まだ確認・対応が終わっていない。要フォロー。"),
]
for label, fill, txt, meaning in legend:
    lc = ws2.cell(row=r, column=2, value=label)
    lc.fill = PatternFill("solid", fgColor=fill)
    lc.font = Font(name=FONT, bold=True, color=txt, size=10)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = BORDER
    mc = ws2.cell(row=r, column=3, value=meaning)
    mc.font = Font(name=FONT, size=10)
    mc.alignment = Alignment(vertical="center", wrap_text=True)
    mc.border = BORDER
    ws2.row_dimensions[r].height = 22
    r += 1

# --- 入力の手引き ---
r += 1
gh = ws2.cell(row=r, column=2, value="入力の手引き")
gh.font = Font(name=FONT, bold=True, size=12, color="FFFFFF")
gh.fill = PatternFill("solid", fgColor=BLUE_HEAD)
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
gh.alignment = Alignment(horizontal="center", vertical="center")
r += 1

guides = [
    ("名前 / アカウント名", "手入力。アカウント名は @ 付きが分かりやすい。"),
    ("タイアップ投稿", "▼ 予定／依頼済／投稿済／見送り から選択。"),
    ("事前確認", "▼ 未確認／確認中／OK／NG。OK＝Meta広告で配信して良いと承諾を得た状態。"),
    ("Meta広告 利用可否", "▼ 利用可／利用不可／確認中。事前確認OKの人だけ「利用可」に。"),
    ("投稿日", "日付を入力（例 2026/08/01）。入力すると報告期限が自動で3日後に。"),
    ("配信可否 報告期限", "自動計算（投稿日＋3日）。期限を過ぎると赤くなる。手入力しない。"),
    ("配信可否 報告済 / 配信済み", "▼ 済／未 で選択。"),
    ("配信予定日", "配信する日を入力。"),
]
for label, meaning in guides:
    gc = ws2.cell(row=r, column=2, value=label)
    gc.font = Font(name=FONT, bold=True, size=10, color=NAVY)
    gc.alignment = Alignment(vertical="top", wrap_text=True)
    gc.border = BORDER
    mc = ws2.cell(row=r, column=3, value=meaning)
    mc.font = Font(name=FONT, size=10)
    mc.alignment = Alignment(vertical="top", wrap_text=True)
    mc.border = BORDER
    ws2.row_dimensions[r].height = 30
    r += 1

wb.save("Meta広告クリエイター管理.xlsx")
print("saved: Meta広告クリエイター管理.xlsx")
